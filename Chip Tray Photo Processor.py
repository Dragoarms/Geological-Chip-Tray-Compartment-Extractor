
"""
Chip Tray Extractor with OCR capabilities

A tool to extract individual compartment images from panoramic chip tray photos 
using ArUco markers for alignment and compartment boundary detection.
Additionally extracts text metadata using Tesseract OCR when available.

Features:
- Simple GUI for folder selection and processing options
- ArUco marker detection for precise compartment extraction
- OCR metadata extraction with Tesseract (optional)
- High-quality image processing to preserve detail
- Visual debug output showing detection steps

Author: George Symonds - (using Claude Sonnet 3.7)

"""
__version__ = "1.1" 

import certifi
print(certifi.where())
import ssl 
import os
import sys
import logging
import threading
import queue
import traceback
import importlib.util
import subprocess
import platform
import webbrowser
import re
import math
import shutil
from datetime import datetime
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Union, Callable, Set
import urllib.request
import tempfile

# Configure logging first
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Handle potentially problematic imports
try:
    from PIL import Image, ImageTk
except ImportError:
    logger.warning("PIL/Pillow library not found, some image operations may fail")

import cv2
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

try:
    import pytesseract
except ImportError:
    # Import failed — call fallback directly if class is defined below
    class _TesseractImportHelper:
        @staticmethod
        def get_installation_instructions():
            from platform import system
            instructions = {
                'Windows': 'https://github.com/UB-Mannheim/tesseract/wiki'
            }
            base = (
                "To use OCR features, you need to:\n\n"
                "1. Install Tesseract OCR for your platform\n"
                "2. Install the pytesseract Python package: pip install pytesseract\n\n"
            )
            sys_type = system()
            specific = f"For {sys_type} systems: {instructions.get(sys_type, 'See https://github.com/tesseract-ocr/tesseract')}"
            return base + specific

    print(_TesseractImportHelper.get_installation_instructions())

class UpdateChecker:
    """
    Checks for updates to the application by comparing local version with GitHub version.
    """
    
    def __init__(self, github_repo="https://github.com/Dragoarms/Geological-Chip-Tray-Compartment-Extractor.git", token=None):
        """
        Initialize the update checker.
        
        Args:
            github_repo: URL to the GitHub repository
            token: GitHub personal access token for private repositories
        """
        self.github_repo = github_repo
        self.token = token  # Store the token for authentication
        self.logger = logging.getLogger(__name__)
        
        # Extract owner and repo name from the URL
        match = re.search(r'github\.com/([^/]+)/([^/.]+)', github_repo)
        if match:
            self.owner = match.group(1)
            self.repo = match.group(2)
            self.logger.info(f"Parsed GitHub repo: owner={self.owner}, repo={self.repo}")
        else:
            self.owner = None
            self.repo = None
            self.logger.error(f"Could not parse GitHub repository URL: {github_repo}")
    
    def get_local_version(self) -> str:
        """
        Get the local version from the script's __version__ variable.
        
        Returns:
            Local version as a string
        """
        try:
            # Check for __version__ in main module
            import __main__
            if hasattr(__main__, '__version__'):
                return __main__.__version__
                
            # If not found in main, use the global __version__ if available
            if '__version__' in globals():
                return globals()['__version__']
                
            # Fallback to hardcoded version
            return "1.0"  # Default fallback version
                
        except Exception as e:
            self.logger.error(f"Error getting local version: {str(e)}")
            return "1.0"  # Default fallback version
    
    def get_github_version(self) -> str:
        """
        Get the latest version from GitHub.
        
        Returns:
            Latest version as a string, or "Unknown" if not found
        """
        try:
            if not self.owner or not self.repo:
                return "Unknown"
                
            # Try location for the version information
            possible_paths = [
                f"https://raw.githubusercontent.com/{self.owner}/{self.repo}/main/version.txt",
            ]
            
            # Create a context that doesn't verify SSL certificates
            context = ssl._create_unverified_context()
            
            # Try each path
            for raw_url in possible_paths:
                try:
                    self.logger.info(f"Trying URL: {raw_url}")
                    
                    # Create a request object
                    request = urllib.request.Request(raw_url)
                    
                    # Add authorization header if token is provided
                    if self.token:
                        request.add_header("Authorization", f"token {self.token}")
                    
                    # Make the request
                    response = urllib.request.urlopen(request, context=context)
                    content = response.read().decode('utf-8')
                    
                    # Parse version using regex
                    match = re.search(r'[Vv]ersion\s*=\s*["\']*(\d+\.\d+(?:\.\d+)?)["\']*', content)
                    if match:
                        return match.group(1)
                except Exception as path_error:
                    self.logger.warning(f"Failed with path {raw_url}: {str(path_error)}")
                    continue
            
            # If all paths fail, try parsing version directly from repository URL
            # Format like "V2.py" suggests version 2.0
            match = re.search(r'V(\d+)\.py', self.github_repo)
            if match:
                return f"{match.group(1)}.0"
                
            # Last resort - use current version
            return self.get_local_version()
        except Exception as e:
            self.logger.error(f"Error getting GitHub version: {str(e)}")
            return "Unknown"

    def compare_versions(self) -> dict:
        """
        Compare local and GitHub versions.

        Returns:
            Dictionary with comparison results
        """
        local_version = self.get_local_version()
        github_version = self.get_github_version()

        result = {
            'local_version': local_version,
            'github_version': github_version,
            'update_available': False,
            'error': None
        }

        if local_version == "Unknown" or github_version == "Unknown":
            result['error'] = "Could not determine versions"
            return result

        try:
            # Convert to tuples of integers for comparison
            local_parts = tuple(map(int, local_version.split('.')))
            github_parts = tuple(map(int, github_version.split('.')))

            # Pad with zeros if versions have different number of parts
            max_length = max(len(local_parts), len(github_parts))
            local_parts = local_parts + (0,) * (max_length - len(local_parts))
            github_parts = github_parts + (0,) * (max_length - len(github_parts))

            result['update_available'] = github_parts > local_parts

            return result
        except Exception as e:
            result['error'] = str(e)
            self.logger.error(f"Error comparing versions: {str(e)}")
            return result

    def download_and_replace_script(self, file_manager, script_name="Chip Tray Photo Processor.py"):
        """
        Downloads the updated script from GitHub, replaces the current file, and restarts.

        Args:
            file_manager: Instance of FileManager that provides the base directory
            script_name: The filename to fetch from GitHub and run
        """
        try:
            import sys
            import subprocess
            import tempfile

            # Target directory in Program Resources
            program_dir = os.path.join(file_manager.base_dir, "Program Resources")
            os.makedirs(program_dir, exist_ok=True)

            # Remote raw GitHub URL
            raw_url = f"https://raw.githubusercontent.com/{self.owner}/{self.repo}/main/{script_name.replace(' ', '%20')}"
            target_path = os.path.join(program_dir, script_name)

            # Download to a temporary file first
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".py")
            os.close(tmp_fd)

            context = ssl._create_unverified_context()

            with urllib.request.urlopen(raw_url, context=context) as response, open(tmp_path, 'wb') as out_file:
                shutil.copyfileobj(response, out_file)

            # Move updated file to Program Resources
            shutil.move(tmp_path, target_path)
            self.logger.info(f"✅ Downloaded updated script to: {target_path}")

            # Remove current running script
            current_script = os.path.abspath(sys.argv[0])
            self.logger.info(f"🧹 Deleting current script: {current_script}")
            try:
                os.remove(current_script)
            except Exception as delete_error:
                self.logger.warning(f"⚠️ Could not delete original script: {delete_error}")

            # Restart from new script
            self.logger.info("🔁 Restarting from updated script...")
            subprocess.Popen([sys.executable, target_path])
            sys.exit(0)

        except Exception as e:
            self.logger.error(f"❌ Update failed: {e}")
            messagebox.showerror("Update Failed", f"An error occurred while updating:\n{e}")


class TesseractManager:
    """
    Manages Tesseract OCR detection, installation guidance, and OCR functionality.
    Designed to gracefully handle cases where Tesseract is not installed.
    """
    
    def __init__(self):
        """Initialize the Tesseract manager with default paths and settings."""
        self.is_available = False
        self.version = None
        self.pytesseract = None
        self.install_instructions = {
            'Windows': 'https://github.com/UB-Mannheim/tesseract/wiki',
            'Darwin': 'https://brew.sh/ (then run: brew install tesseract)',
            'Linux': 'sudo apt-get install tesseract-ocr libtesseract-dev'
        }
        
        # Try to detect Tesseract
        self.detect_tesseract()
    
    def detect_tesseract(self) -> bool:
        """
        Detect if Tesseract OCR is installed and available in the system.
        
        Returns:
            bool: True if Tesseract is available, False otherwise
        """
        # First check if pytesseract can be imported
        try:
            # Try to import pytesseract
            if importlib.util.find_spec("pytesseract") is None:
                logger.warning("pytesseract package not found")
                return False
            
            # Import pytesseract for OCR functionality
            import pytesseract
            from pytesseract import Output
            self.pytesseract = pytesseract
            
            # Try various methods to find Tesseract executable
            if platform.system() == 'Windows':
                # Common installation locations on Windows
                possible_paths = [
                    r'C:\Program Files\Tesseract-OCR',
                    r'C:\Program Files (x86)\Tesseract-OCR',
                    os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Programs', 'Tesseract-OCR'),
                    os.path.join(os.environ.get('APPDATA', ''), 'Tesseract-OCR')
                ]
                
                # Find the first valid path
                tesseract_path = None
                for path in possible_paths:
                    exe_path = os.path.join(path, 'tesseract.exe')
                    if os.path.exists(exe_path):
                        tesseract_path = path
                        break
                
                if tesseract_path:
                    # Add to system PATH if it exists
                    os.environ['PATH'] = os.environ['PATH'] + os.pathsep + tesseract_path
                    pytesseract.pytesseract.tesseract_cmd = os.path.join(tesseract_path, 'tesseract.exe')
                    logger.info(f"Found Tesseract at: {tesseract_path}")
            
            # Test if Tesseract works by getting version
            try:
                self.version = pytesseract.get_tesseract_version()
                logger.info(f"Tesseract OCR version {self.version} detected")
                self.is_available = True
                return True
            except Exception as e:
                logger.warning(f"Tesseract is installed but not working correctly: {str(e)}")
                return False
                
        except ImportError:
            logger.warning("Failed to import pytesseract module")
            return False
        except Exception as e:
            logger.warning(f"Error detecting Tesseract: {str(e)}")
            return False
    
    def get_installation_instructions(self) -> str:
        """
        Get installation instructions for the current platform.
        
        Returns:
            str: Installation instructions for Tesseract OCR
        """
        system = platform.system()
        base_instructions = (
            "To use OCR features, you need to:\n\n"
            "1. Install Tesseract OCR for your platform\n"
            "2. Install the pytesseract Python package: pip install pytesseract\n\n"
        )
        
        if system in self.install_instructions:
            platform_instructions = f"For {system} systems: {self.install_instructions[system]}"
        else:
            platform_instructions = "Visit https://github.com/tesseract-ocr/tesseract for installation instructions."
        
        return base_instructions + platform_instructions
    
    def show_installation_dialog(self, parent: Optional[tk.Tk] = None) -> None:
        """
        Show a dialog with installation instructions for Tesseract OCR.
        
        Args:
            parent: Optional parent Tkinter window
        """
        instructions = self.get_installation_instructions()
        system = platform.system()
        
        # If no parent window, just log the instructions
        if parent is None:
            logger.info(f"Tesseract OCR not available. {instructions}")
            return
        
        # Create a custom dialog
        dialog = tk.Toplevel(parent)
        dialog.title("Tesseract OCR Required")
        dialog.geometry("500x400")
        dialog.grab_set()  # Make dialog modal
        
        # Add icon and header
        header_frame = ttk.Frame(dialog, padding="10")
        header_frame.pack(fill=tk.X)
        
        # Use system-specific icons
        if system == 'Windows':
            try:
                dialog.iconbitmap(default='')  # Default Windows icon
            except:
                pass
                
        # Header label
        header_label = ttk.Label(
            header_frame, 
            text="Tesseract OCR Required for Text Recognition",
            font=("Arial", 12, "bold")
        )
        header_label.pack(pady=10)
        
        # Instructions text
        text_frame = ttk.Frame(dialog, padding="10")
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        instructions_text = tk.Text(text_frame, wrap=tk.WORD, height=10)
        instructions_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        instructions_text.insert(tk.END, instructions)
        instructions_text.config(state=tk.DISABLED)
        
        # Scrollbar for text
        scrollbar = ttk.Scrollbar(instructions_text, command=instructions_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        instructions_text.config(yscrollcommand=scrollbar.set)
        
        # Action buttons
        button_frame = ttk.Frame(dialog, padding="10")
        button_frame.pack(fill=tk.X)
        
        # Platform-specific install button
        if system in self.install_instructions:
            install_url = self.install_instructions[system].split(' ')[0]
            install_button = ttk.Button(
                button_frame, 
                text=f"Download for {system}", 
                command=lambda: webbrowser.open(install_url)
            )
            install_button.pack(side=tk.LEFT, padx=5)
        
        close_button = ttk.Button(
            button_frame, 
            text="Continue without OCR", 
            command=dialog.destroy
        )
        close_button.pack(side=tk.RIGHT, padx=5)
    
    # Function to extract patterns from OCR
    def extract_metadata_with_composite(self, 
                                    image: np.ndarray, 
                                    markers: Dict[int, np.ndarray], 
                                    original_filename: Optional[str] = None,
                                    progress_queue: Optional[queue.Queue] = None) -> Dict[str, Any]:
        """
        Extract metadata using a composite approach - combining multiple preprocessing methods
        on the extracted label regions and voting on the most likely correct values.
        
        Args:
            image: Input image
            markers: Detected ArUco marker positions
            original_filename: Original image filename for debug purposes
            progress_queue: Optional queue for reporting progress
        
        Returns:
            Metadata dictionary with extracted information
        """

        
        try:
            # Get image dimensions
            h, w = image.shape[:2]
            
            # Check if we have necessary markers
            if not markers:
                logger.warning("No markers found for metadata region extraction")
                return {
                    'hole_id': None,
                    'depth_from': None,
                    'depth_to': None,
                    'confidence': 0.0,
                    'metadata_region': None
                }
            
            # Define metadata region using markers
            metadata_region = None  # Initialize the variable
            hole_id_region = None
            depth_region = None
            
            # Look for marker ID 24 which is between hole ID and depth labels
            if 24 in markers:
                # Use marker 24 as a reference point
                marker24_corners = markers[24]
                marker24_center_x = int(np.mean(marker24_corners[:, 0]))
                marker24_center_y = int(np.mean(marker24_corners[:, 1]))
                
                # The hole ID label is likely above marker 24
                hole_id_region_y1 = max(0, marker24_center_y - 150)  # 150px above marker center
                hole_id_region_y2 = max(0, marker24_center_y - 20)   # 20px above marker center
                
                # The depth label is likely below marker 24
                depth_region_y1 = min(h, marker24_center_y + 20)    # 20px below marker center
                depth_region_y2 = min(h, marker24_center_y + 150)   # 150px below marker center
                
                # Horizontal range for both regions (centered on marker 24)
                region_x1 = max(0, marker24_center_x - 100)
                region_x2 = min(w, marker24_center_x + 100)
                
                # Extract the hole ID and depth regions
                hole_id_region = image[hole_id_region_y1:hole_id_region_y2, region_x1:region_x2].copy()
                depth_region = image[depth_region_y1:depth_region_y2, region_x1:region_x2].copy()
                
                # Create visualization
                viz_image = image.copy()
                # Draw hole ID region in blue
                cv2.rectangle(viz_image, 
                            (region_x1, hole_id_region_y1), 
                            (region_x2, hole_id_region_y2), 
                            (255, 0, 0), 2)
                # Draw marker 24 in purple
                cv2.circle(viz_image, (marker24_center_x, marker24_center_y), 10, (255, 0, 255), -1)
                # Draw depth region in green
                cv2.rectangle(viz_image, 
                            (region_x1, depth_region_y1), 
                            (region_x2, depth_region_y2), 
                            (0, 255, 0), 2)
                            
                # Create a metadata region by combining both regions
                metadata_region = image[hole_id_region_y1:depth_region_y2, region_x1:region_x2].copy()
                metadata_region_viz = viz_image
                
            else:
                # Fallback: use the bottom portion of the image
                logger.warning("Marker 24 not found, using fallback region detection")
                
                # Define a larger metadata region at the bottom of the image
                metadata_top = int(h * 0.65)  # Start at 65% down the image
                metadata_left = int(w * 0.03)  # 3% from left edge
                metadata_right = int(w * 0.5)  # Use left half of the image
                metadata_bottom = h
                
                # Extract the whole metadata region
                metadata_region = image[
                    metadata_top:metadata_bottom, 
                    metadata_left:metadata_right
                ].copy()
                
                # Split this region for hole ID (top) and depth (bottom)
                middle_y = (metadata_bottom - metadata_top) // 2 + metadata_top
                
                hole_id_region = image[metadata_top:middle_y, metadata_left:metadata_right].copy()
                depth_region = image[middle_y:metadata_bottom, metadata_left:metadata_right].copy()
                
                # Create visualization
                viz_image = image.copy()
                cv2.rectangle(viz_image, 
                            (metadata_left, metadata_top), 
                            (metadata_right, metadata_bottom), 
                            (0, 255, 0), 2)
                cv2.line(viz_image,
                        (metadata_left, middle_y),
                        (metadata_right, middle_y),
                        (255, 0, 0), 2)
                        
                metadata_region_viz = viz_image
            
            # ALWAYS save visualization to memory cache regardless of debug settings
            # This is crucial for duplicate detection dialog
            if hasattr(self, 'extractor') and hasattr(self.extractor, 'visualization_cache'):
                # We'll create a preliminary cache entry with what we have so far
                # The final values will be updated later once we have hole_id and depths
                self.extractor.visualization_cache['current_processing'] = {
                    'metadata_region': metadata_region,
                    'metadata_region_viz': metadata_region_viz,
                    'hole_id_region': hole_id_region,
                    'depth_region': depth_region
                }
            
            # Save metadata region visualization if debug is enabled
            if original_filename and hasattr(self, 'config') and self.config.get('save_debug_images', False):
                file_manager = getattr(self, 'file_manager', None)
                if file_manager is not None:
                    file_manager.save_temp_debug_image(
                        metadata_region_viz,
                        original_filename,
                        "ocr_regions"
                    )
                    
                    # Save individual regions for reference
                    if hole_id_region is not None:
                        file_manager.save_temp_debug_image(
                            hole_id_region,
                            original_filename,
                            "hole_id_region"
                        )
                    
                    if depth_region is not None:
                        file_manager.save_temp_debug_image(
                            depth_region,
                            original_filename,
                            "depth_region"
                        )
            
            # Process hole ID and depth regions separately
            hole_id_results = []
            depth_results = []
            
            if hole_id_region is not None and depth_region is not None:
                # Create preprocessing methods specifically tailored for label text
                # From your images, we can see both original color and various binary versions
                preprocessing_methods = [
                    ("original", lambda img: img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)),
                    ("threshold1", lambda img: cv2.threshold(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        127, 255, cv2.THRESH_BINARY)[1]),
                    ("threshold2", lambda img: cv2.threshold(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        100, 255, cv2.THRESH_BINARY)[1]),
                    ("threshold3", lambda img: cv2.threshold(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        150, 255, cv2.THRESH_BINARY)[1]),
                    ("adaptive", lambda img: cv2.adaptiveThreshold(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)),
                    ("otsu", lambda img: cv2.threshold(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]),
                    ("canny", lambda img: cv2.Canny(
                        img if len(img.shape) == 2 else cv2.cvtColor(img, cv2.COLOR_BGR2GRAY), 
                        50, 150))
                ]
                
                # Create composite images for hole ID and depth
                hole_id_composite_images = []
                depth_composite_images = []
                
                # Process hole ID region
                for method_name, method_func in preprocessing_methods:
                    try:
                        # Process image
                        processed_img = method_func(hole_id_region)
                        
                        # Run OCR specifically for hole ID pattern
                        # Use different PSM modes that are optimal for single line text (7) or single word (8)
                        for psm in [7, 8, 6]:
                            config = f"--psm {psm} --oem 3 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
                            text = self.pytesseract.image_to_string(processed_img, config=config).strip()
                            
                            # Try to extract hole ID pattern
                            match = re.search(r'([A-Z]{2}\d{4})', text.replace(" ", ""))
                            hole_id = match.group(1) if match else None
                            
                            if hole_id:
                                hole_id_results.append({
                                    'hole_id': hole_id,
                                    'method': method_name,
                                    'psm': psm,
                                    'text': text,
                                    'processed_image': processed_img
                                })
                                
                                # Add image to composite with text overlay
                                labeled_img = processed_img.copy()
                                if len(labeled_img.shape) == 2:
                                    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_GRAY2BGR)
                                
                                # Add text overlay
                                cv2.putText(
                                    labeled_img,
                                    f"{method_name} PSM{psm}: {hole_id}",
                                    (10, 30),
                                    cv2.FONT_HERSHEY_SIMPLEX,
                                    0.7,
                                    (0, 0, 255),
                                    2
                                )
                                
                                hole_id_composite_images.append(labeled_img)
                                
                                # If we found a valid hole ID, no need to try other PSM modes
                                break
                        
                        # Save individual processed images if debug is enabled
                        if original_filename and hasattr(self, 'config') and self.config.get('save_debug_images', False):
                            file_manager = getattr(self, 'file_manager', None)
                            if file_manager is not None:
                                file_manager.save_temp_debug_image(
                                    processed_img,
                                    original_filename,
                                    f"hole_id_{method_name}"
                                )
                    except Exception as e:
                        logger.error(f"Error processing hole ID with {method_name}: {str(e)}")
                
                # Process depth region
                for method_name, method_func in preprocessing_methods:
                    try:
                        # Process image
                        processed_img = method_func(depth_region)
                        
                        # Run OCR specifically for depth pattern
                        # Try different PSM modes
                        for psm in [7, 6, 3]:
                            config = f"--psm {psm} --oem 3 -c tessedit_char_whitelist=0123456789-."
                            text = self.pytesseract.image_to_string(processed_img, config=config).strip()
                            
                            # Try to extract depth pattern
                            match = re.search(r'(\d+)[\s\-–—to]+(\d+)', text)
                            if match:
                                depth_from = float(match.group(1))
                                depth_to = float(match.group(2))
                                
                                depth_results.append({
                                    'depth_from': depth_from,
                                    'depth_to': depth_to,
                                    'method': method_name,
                                    'psm': psm,
                                    'text': text,
                                    'processed_image': processed_img
                                })
                                
                                # Add image to composite with text overlay
                                labeled_img = processed_img.copy()
                                if len(labeled_img.shape) == 2:
                                    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_GRAY2BGR)
                                
                                # Add text overlay
                                cv2.putText(
                                    labeled_img,
                                    f"{method_name} PSM{psm}: {depth_from}-{depth_to}",
                                    (10, 30),
                                    cv2.FONT_HERSHEY_SIMPLEX,
                                    0.7,
                                    (0, 0, 255),
                                    2
                                )
                                
                                depth_composite_images.append(labeled_img)
                                
                                # If we found valid depths, no need to try other PSM modes
                                break
                        
                        # Save individual processed images if debug is enabled
                        if original_filename and hasattr(self, 'config') and self.config.get('save_debug_images', False):
                            file_manager = getattr(self, 'file_manager', None)
                            if file_manager is not None:
                                file_manager.save_temp_debug_image(
                                    processed_img,
                                    original_filename,
                                    f"depth_{method_name}"
                                )
                    except Exception as e:
                        logger.error(f"Error processing depth with {method_name}: {str(e)}")
                
                # Create and save composite images
                if hole_id_composite_images and hasattr(self, 'config') and self.config.get('save_debug_images', False):
                    file_manager = getattr(self, 'file_manager', None)
                    if file_manager is not None:
                        # Create a grid layout for hole ID images
                        grid_size = math.ceil(math.sqrt(len(hole_id_composite_images)))
                        grid_h = grid_size
                        grid_w = grid_size
                        
                        # Find max dimensions
                        max_h = max(img.shape[0] for img in hole_id_composite_images)
                        max_w = max(img.shape[1] for img in hole_id_composite_images)
                        
                        # Create grid
                        grid = np.ones((max_h * grid_h, max_w * grid_w, 3), dtype=np.uint8) * 255
                        
                        # Place images
                        for i, img in enumerate(hole_id_composite_images):
                            if i >= grid_h * grid_w:
                                break
                                
                            row = i // grid_w
                            col = i % grid_w
                            
                            y = row * max_h
                            x = col * max_w
                            
                            h, w = img.shape[:2]
                            grid[y:y+h, x:x+w] = img
                        
                        # Save composite
                        file_manager.save_temp_debug_image(
                            grid,
                            original_filename,
                            "hole_id_composite"
                        )
                
                # Same for depth composite
                if depth_composite_images and hasattr(self, 'config') and self.config.get('save_debug_images', False):
                    file_manager = getattr(self, 'file_manager', None)
                    if file_manager is not None:
                        # Create a grid layout for depth images
                        grid_size = math.ceil(math.sqrt(len(depth_composite_images)))
                        grid_h = grid_size
                        grid_w = grid_size
                        
                        # Find max dimensions
                        max_h = max(img.shape[0] for img in depth_composite_images)
                        max_w = max(img.shape[1] for img in depth_composite_images)
                        
                        # Create grid
                        grid = np.ones((max_h * grid_h, max_w * grid_w, 3), dtype=np.uint8) * 255
                        
                        # Place images
                        for i, img in enumerate(depth_composite_images):
                            if i >= grid_h * grid_w:
                                break
                                
                            row = i // grid_w
                            col = i % grid_w
                            
                            y = row * max_h
                            x = col * max_w
                            
                            h, w = img.shape[:2]
                            grid[y:y+h, x:x+w] = img
                        
                        # Save composite
                        file_manager.save_temp_debug_image(
                            grid,
                            original_filename,
                            "depth_composite"
                        )
            
            # Vote on results
            # For hole ID
            hole_id_votes = {}
            for result in hole_id_results:
                hole_id = result.get('hole_id')
                if hole_id:
                    hole_id_votes[hole_id] = hole_id_votes.get(hole_id, 0) + 1
            
            # For depths
            depth_from_votes = {}
            depth_to_votes = {}
            for result in depth_results:
                depth_from = result.get('depth_from')
                depth_to = result.get('depth_to')
                
                if depth_from is not None:
                    # Round to nearest integer for voting
                    depth_from_int = int(round(depth_from))
                    depth_from_votes[depth_from_int] = depth_from_votes.get(depth_from_int, 0) + 1
                
                if depth_to is not None:
                    # Round to nearest integer for voting
                    depth_to_int = int(round(depth_to))
                    depth_to_votes[depth_to_int] = depth_to_votes.get(depth_to_int, 0) + 1
            
            # Get the most voted values
            final_hole_id = None
            if hole_id_votes:
                final_hole_id = max(hole_id_votes.items(), key=lambda x: x[1])[0]
            
            final_depth_from = None
            if depth_from_votes:
                final_depth_from = max(depth_from_votes.items(), key=lambda x: x[1])[0]
            
            final_depth_to = None
            if depth_to_votes:
                final_depth_to = max(depth_to_votes.items(), key=lambda x: x[1])[0]
            
            # Additional validation for depths
            if final_depth_from is not None and final_depth_to is not None:
                # Make sure depth_to > depth_from
                if final_depth_to <= final_depth_from:
                    logger.warning(f"Invalid depth range detected: {final_depth_from}-{final_depth_to}")
                    
                    # Try to correct: standard pattern is 20m intervals
                    if abs(final_depth_to - final_depth_from) < 5:
                        # Likely a small OCR error, set depth_to to depth_from + 20
                        final_depth_to = final_depth_from + 20
                        logger.info(f"Corrected depth range to: {final_depth_from}-{final_depth_to}")
            
            # Calculate confidence based on vote consistency
            confidence = 0.0
            
            # Hole ID confidence
            if final_hole_id and hole_id_votes:
                # Percentage of agreement
                agreement = min(100.0, hole_id_votes[final_hole_id] / len(hole_id_results) * 100 if hole_id_results else 0)
                confidence += min(agreement, 50.0)  # Max 50 points from hole ID

            # Depth confidence
            if final_depth_from is not None and final_depth_to is not None and depth_from_votes and depth_to_votes:
                # Percentage of agreement for each value
                # Make sure the keys exist in the dictionaries
                from_agreement = min(100.0, depth_from_votes.get(final_depth_from, 0) / len(depth_results) * 100 if depth_results else 0)
                to_agreement = min(100.0, depth_to_votes.get(final_depth_to, 0) / len(depth_results) * 100 if depth_results else 0)
                
                # Average the agreements
                avg_agreement = (from_agreement + to_agreement) / 2
                confidence += min(avg_agreement, 50.0)  # Max 50 points from depths

            # Ensure final confidence never exceeds 100%
            confidence = min(confidence, 100.0)
            
            # Final result
            result = {
                'hole_id': final_hole_id,
                'depth_from': float(final_depth_from) if final_depth_from is not None else None,
                'depth_to': float(final_depth_to) if final_depth_to is not None else None,
                'confidence': confidence,
                'metadata_region': metadata_region,
                'metadata_region_viz': metadata_region_viz,
                'ocr_text': f"Composite OCR: ID={final_hole_id}, Depth={final_depth_from}-{final_depth_to}"
            }
            
            # Log final results
            logger.info("------------------------------------------------")
            logger.info("OCR results:")
            if final_hole_id:
                logger.info(f"Hole ID: {final_hole_id} (votes: {hole_id_votes.get(final_hole_id, 0)}/{len(hole_id_results) or 1})")
            else:
                logger.info("Hole ID: None")
                
            if final_depth_from is not None and final_depth_to is not None:
                logger.info(f"Depth range: {final_depth_from}-{final_depth_to} "
                        f"(votes: {depth_from_votes.get(final_depth_from, 0)}/{len(depth_results) or 1}, "
                        f"{depth_to_votes.get(final_depth_to, 0)}/{len(depth_results) or 1})")
            else:
                logger.info("Depth range: None")
                
            logger.info(f"Confidence: {confidence:.1f}%")
            logger.info("------------------------------------------------")
            
            return result
            
        except Exception as e:
            logger.error(f"Error in composite OCR extraction: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'hole_id': None,
                'depth_from': None,
                'depth_to': None,
                'confidence': 0.0,
                'metadata_region': None
            }
    
    def _remove_borders(self, gray_image: np.ndarray) -> Optional[np.ndarray]:
        """
        Detect and remove black borders around chip tray labels.
        
        Args:
            gray_image: Grayscale input image
        
        Returns:
            Border-removed image or None if no borders detected
        """
        try:
            # Create binary image to identify potential borders
            _, border_thresh = cv2.threshold(gray_image, 50, 255, cv2.THRESH_BINARY)
            
            # Find contours in the thresholded image
            contours, _ = cv2.findContours(border_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            # If we have contours, find the largest one (likely the label boundary)
            if contours:
                # Sort contours by area, largest first
                contours = sorted(contours, key=cv2.contourArea, reverse=True)
                
                # Get the largest contour
                largest_contour = contours[0]
                
                # Check if this contour is sufficiently large and looks like a border
                # (covers at least 70% of the perimeter of the image)
                image_perimeter = 2 * (gray_image.shape[0] + gray_image.shape[1])
                contour_perimeter = cv2.arcLength(largest_contour, True)
                perimeter_ratio = contour_perimeter / image_perimeter
                
                if perimeter_ratio > 0.7:
                    logger.info(f"Detected potential border (perimeter ratio: {perimeter_ratio:.2f})")
                    
                    # Get bounding rectangle
                    x, y, w, h = cv2.boundingRect(largest_contour)
                    
                    # Ensure bounding rect covers most of the image (indicating a border)
                    image_area = gray_image.shape[0] * gray_image.shape[1]
                    rect_area = w * h
                    area_ratio = rect_area / image_area
                    
                    if area_ratio > 0.8:
                        logger.info(f"Confirmed border detection (area ratio: {area_ratio:.2f})")
                        
                        # Create a mask for everything inside the contour
                        mask = np.zeros_like(gray_image)
                        cv2.drawContours(mask, [largest_contour], 0, 255, -1)
                        
                        # Create a smaller mask to exclude the border itself
                        # Use adaptive kernel size based on image dimensions
                        border_width = min(max(3, min(w, h) // 30), 10)  # Between 3 and 10 pixels
                        kernel = np.ones((border_width, border_width), np.uint8)
                        eroded_mask = cv2.erode(mask, kernel, iterations=1)
                        
                        # Extract the label content without the border
                        inner_content = np.ones_like(gray_image) * 255  # White background
                        inner_content[eroded_mask > 0] = gray_image[eroded_mask > 0]
                        
                        return inner_content
            
            # No borders detected or removed
            return None
            
        except Exception as e:
            logger.error(f"Error during border removal: {str(e)}")
            return None

class MetadataInputDialog:
    """
    Dialog for collecting metadata when OCR fails or needs confirmation.
    Shows the metadata region of the image and provides fields for entering hole ID and depth range.
    """
    
    def __init__(self, parent: Optional[tk.Tk], image: Optional[np.ndarray] = None, 
                metadata: Optional[Dict[str, Any]] = None,
                tesseract_manager: Optional[Any] = None):
        """
        Initialize the metadata input dialog.
        
        Args:
            parent: Parent Tkinter window
            image: Optional image to display (metadata region)
            metadata: Optional pre-filled metadata from OCR
            tesseract_manager: Optional TesseractManager instance for validation
        """
        self.parent = parent
        self.image = image
        self.metadata = metadata or {}
        self.tesseract_manager = tesseract_manager
        
        # Result values
        self.hole_id = tk.StringVar(value=self.metadata.get('hole_id', ''))
        self.depth_from = tk.StringVar(value=str(self.metadata.get('depth_from', '')))
        self.depth_to = tk.StringVar(value=str(self.metadata.get('depth_to', '')))
        self.result = None
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Enter Chip Tray Metadata")
        self.dialog.grab_set()  # Make dialog modal
        
        # Set a larger default size (width x height)
        self.dialog.geometry("1200x900")
        
        # Ensure the dialog can be resized
        self.dialog.resizable(True, True)
        
        self._create_widgets()
    
    def _create_widgets(self) -> None:
        """Create all widgets for the dialog."""
        # Main frame with padding
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_label = ttk.Label(
            main_frame, 
            text="Enter Chip Tray Metadata",
            font=("Arial", 12, "bold")
        )
        header_label.pack(pady=(0, 10))
        
        # Image display (if provided)
        if self.image is not None:
            self._add_image_display(main_frame)
        
        # OCR result display (if available)
        if self.metadata.get('ocr_text'):
            self._add_ocr_result_display(main_frame)
        
        # Input fields
        self._add_input_fields(main_frame)
        
        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ok_button = ttk.Button(
            button_frame, 
            text="OK", 
            command=self._on_ok
        )
        ok_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        cancel_button = ttk.Button(
            button_frame, 
            text="Cancel", 
            command=self._on_cancel
        )
        cancel_button.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(5, 0))
    
    
    
    def _add_image_display(self, parent: ttk.Frame) -> None:
        """
        Add image display widget to the dialog with specialized layout for chip tray labels.
        
        Args:
            parent: Parent frame to add the widget to
        """
        # Create a frame to hold all images
        image_frame = ttk.LabelFrame(parent, text="Image Analysis", padding="10")
        image_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        try:
            # Create a canvas for scrollable content
            canvas = tk.Canvas(image_frame, borderwidth=0)
            scrollbar = ttk.Scrollbar(image_frame, orient=tk.VERTICAL, command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Display order and captions
            display_items = [
                ('metadata_region_viz', 'Full Image with Detected Region'),
                ('metadata_region', 'Extracted Metadata Region'),
                ('preprocessed_ocr', 'OCR Processing Result')
            ]
            
            for item_key, item_title in display_items:
                if item_key in self.metadata and self.metadata[item_key] is not None:
                    # Create container for this image
                    item_frame = ttk.Frame(scrollable_frame)
                    item_frame.pack(pady=10, fill=tk.X)
                    
                    # Add title above image
                    title_label = ttk.Label(
                        item_frame, 
                        text=item_title,
                        font=("Arial", 10, "bold")
                    )
                    title_label.pack(pady=(0, 5))
                    
                    # Prepare image
                    img = self.metadata[item_key]
                    h, w = img.shape[:2]
                    
                    # Calculate display dimensions - preserve aspect ratio
                    max_width = 1000  # Maximum width for display
                    max_height = 600  # Maximum height for display
                    
                    # Calculate scale factor to fit within limits
                    width_scale = max_width / w if w > max_width else 1
                    height_scale = max_height / h if h > max_height else 1
                    scale = min(width_scale, height_scale)
                    
                    # Apply scaling
                    if scale < 1:
                        new_width = int(w * scale)
                        new_height = int(h * scale)
                        resized = cv2.resize(img, (new_width, new_height), interpolation=cv2.INTER_AREA)
                    else:
                        resized = img.copy()
                    
                    # Convert to RGB if needed
                    if len(resized.shape) == 3:
                        if resized.shape[2] == 3:
                            img_rgb = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
                        else:
                            img_rgb = resized
                    else:
                        img_rgb = cv2.cvtColor(resized, cv2.COLOR_GRAY2RGB)
                    
                    # Convert to Tkinter compatible image
                    from PIL import Image, ImageTk
                    pil_img = Image.fromarray(img_rgb)
                    tk_img = ImageTk.PhotoImage(image=pil_img)
                    
                    # Create image display
                    img_label = ttk.Label(item_frame, image=tk_img)
                    img_label.image = tk_img  # Keep reference
                    img_label.pack()
                    
                    # Add dimensions info
                    size_label = ttk.Label(
                        item_frame,
                        text=f"Original size: {w}x{h} pixels",
                        font=("Arial", 8),
                        foreground="gray"
                    )
                    size_label.pack(pady=(5, 0))
            
            # Add help text
            help_text = ttk.Label(
                scrollable_frame, 
                text="Examine these images to verify hole ID and depth range.",
                font=("Arial", 9)
            )
            help_text.pack(pady=10)
            
        except Exception as e:
            # If image display fails, show error message
            logger.error(f"Error displaying metadata images: {str(e)}")
            logger.error(traceback.format_exc())
            error_label = ttk.Label(
                image_frame, 
                text=f"Error displaying images: {str(e)}",
                foreground="red"
            )
            error_label.pack(pady=10)
    
    def _add_ocr_result_display(self, parent: ttk.Frame) -> None:
        """
        Add OCR result display widget to the dialog.
        
        Args:
            parent: Parent frame to add the widget to
        """
        ocr_frame = ttk.LabelFrame(parent, text="OCR Result", padding="10")
        ocr_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Display OCR text
        ocr_text = tk.Text(ocr_frame, height=3, wrap=tk.WORD)
        ocr_text.pack(fill=tk.X, expand=True)
        ocr_text.insert(tk.END, self.metadata.get('ocr_text', ''))
        ocr_text.config(state=tk.DISABLED)
        
        # Add confidence info
        confidence = self.metadata.get('confidence', 0)
        confidence_label = ttk.Label(
            ocr_frame, 
            text=f"OCR Confidence: {confidence:.1f}%",
            font=("Arial", 9)
        )
        confidence_label.pack(anchor='e', pady=(5, 0))
    
    def _add_input_fields(self, parent: ttk.Frame) -> None:
        """
        Add input fields for metadata.
        
        Args:
            parent: Parent frame to add the widgets to
        """
        fields_frame = ttk.LabelFrame(parent, text="Enter Metadata", padding="10")
        fields_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Hole ID field
        hole_id_frame = ttk.Frame(fields_frame)
        hole_id_frame.pack(fill=tk.X, pady=5)
        
        hole_id_label = ttk.Label(
            hole_id_frame, 
            text="Hole ID:",
            width=10,
            font=("Arial", 12, "bold")  # Larger, bold font
        )
        hole_id_label.pack(side=tk.LEFT)
        
        # Custom font for entry
        custom_font = ("Arial", 14, "bold")  # Even larger, bold font for entry
        
        hole_id_entry = ttk.Entry(
            hole_id_frame,
            textvariable=self.hole_id,
            font=custom_font,  # Apply custom font
            width=15  # Wider field for better visibility
        )
        hole_id_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Format help
        hole_id_help = ttk.Label(
            hole_id_frame,
            text="Format: XX0000",
            font=("Arial", 10),  # Slightly larger font for help text
            foreground="gray"
        )
        hole_id_help.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Depth range fields
        depth_frame = ttk.Frame(fields_frame)
        depth_frame.pack(fill=tk.X, pady=5)
        
        depth_label = ttk.Label(
            depth_frame, 
            text="Depth:",
            width=10,
            font=("Arial", 12, "bold")  # Larger, bold font
        )
        depth_label.pack(side=tk.LEFT)
        
        depth_from_entry = ttk.Entry(
            depth_frame,
            textvariable=self.depth_from,
            width=8,
            font=custom_font  # Apply custom font
        )
        depth_from_entry.pack(side=tk.LEFT)
        
        depth_separator = ttk.Label(
            depth_frame, 
            text="-",
            font=custom_font  # Apply same font to separator
        )
        depth_separator.pack(side=tk.LEFT, padx=5)
        
        depth_to_entry = ttk.Entry(
            depth_frame,
            textvariable=self.depth_to,
            width=8,
            font=custom_font  # Apply custom font
        )
        depth_to_entry.pack(side=tk.LEFT)
        
        # Format help
        depth_help = ttk.Label(
            depth_frame,
            text="Format: 0.0-0.0",
            font=("Arial", 10),  # Slightly larger font for help text
            foreground="gray"
        )
        depth_help.pack(side=tk.RIGHT, padx=(5, 0))
    
    def _on_ok(self) -> None:
        """Handle OK button click."""
        # Validate input
        try:
            hole_id = self.hole_id.get().strip()
            depth_from_str = self.depth_from.get().strip()
            depth_to_str = self.depth_to.get().strip()
            
            # Validate hole ID - must be 2 letters followed by 4 digits
            if not hole_id:
                messagebox.showerror("Validation Error", "Hole ID is required")
                return
            
            # Validate hole ID format using regex - 2 letters followed by 4 digits
            if not re.match(r'^[A-Za-z]{2}\d{4}$', hole_id):
                messagebox.showerror("Validation Error", 
                                    "Hole ID must be 2 letters followed by 4 digits (e.g., AB1234)")
                return
            
            # Check if the hole ID prefix is in the list of valid prefixes
            if hasattr(self, 'tesseract_manager') and hasattr(self.tesseract_manager, 'config'):
                config = self.tesseract_manager.config
                if config.get('enable_prefix_validation', False):
                    valid_prefixes = config.get('valid_hole_prefixes', [])
                    if valid_prefixes:
                        prefix = hole_id[:2].upper()
                        if prefix not in valid_prefixes:
                            if not messagebox.askyesno("Prefix Validation Warning", 
                                                    f"The prefix '{prefix}' is not in the list of valid prefixes: {', '.join(valid_prefixes)}.\n\nDo you want to continue anyway?",
                                                    icon='warning'):
                                return
            
            # Validate depth range if provided - must be whole numbers
            depth_from = None
            depth_to = None
            
            if depth_from_str:
                try:
                    depth_from = float(depth_from_str)
                    # Validate as a whole number
                    if depth_from != int(depth_from):
                        messagebox.showerror("Validation Error", "Depth From must be a whole number")
                        return
                    # Convert to integer
                    depth_from = int(depth_from)
                except ValueError:
                    messagebox.showerror("Validation Error", "Depth From must be a number")
                    return
            
            if depth_to_str:
                try:
                    depth_to = float(depth_to_str)
                    # Validate as a whole number
                    if depth_to != int(depth_to):
                        messagebox.showerror("Validation Error", "Depth To must be a whole number")
                        return
                    # Convert to integer
                    depth_to = int(depth_to)
                except ValueError:
                    messagebox.showerror("Validation Error", "Depth To must be a number")
                    return
            
            # Validate that depth_to is greater than depth_from
            if depth_from is not None and depth_to is not None:
                if depth_to <= depth_from:
                    messagebox.showerror("Validation Error", "Depth To must be greater than Depth From")
                    return
                
                # Validate that depth intervals are sensible
                if depth_to - depth_from > 40:
                    if not messagebox.askyesno("Validation Warning", 
                                            f"Depth range ({depth_from}-{depth_to}) seems unusually large. Continue anyway?",
                                            icon='warning'):
                        return
            
            # Set result
            self.result = {
                'hole_id': hole_id,
                'depth_from': depth_from,
                'depth_to': depth_to
            }
            
            # Close dialog
            self.dialog.destroy()
            
        except Exception as e:
            logger.error(f"Error validating metadata input: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    def _on_cancel(self) -> None:
        """Handle Cancel button click."""
        self.result = None
        self.dialog.destroy()
    
    def show(self) -> Optional[Dict[str, Any]]:
        """
        Show the dialog and wait for user input.
        
        Returns:
            Dictionary with entered metadata or None if canceled
        """
        # Wait for dialog to be closed
        self.dialog.wait_window()
        return self.result


class ColorButton(tk.Button):
    """Custom button with colored background and bold text."""
    
    def __init__(self, parent, text, background, foreground="black", command=None, **kwargs):
        """
        Initialize a custom colored button.
        
        Args:
            parent: Parent widget
            text: Button text
            background: Background color
            foreground: Text color
            command: Button command
            **kwargs: Additional arguments for tk.Button
        """
        super().__init__(
            parent,
            text=text,
            background=background,
            foreground=foreground,
            font=("Arial", 12, "bold"),
            command=command,
            relief=tk.RAISED,
            borderwidth=2,
            **kwargs
        )


class QAQCManager:
    """
    Manages quality assurance and quality control of extracted compartment images.
    Provides a GUI for reviewing images and approving/rejecting them.
    """
    
    def __init__(self, root, file_manager, extractor):
        """
        Initialize the QAQC Manager.
        
        Args:
            root: tkinter root window
            file_manager: FileManager instance for handling file operations
            extractor: ChipTrayExtractor instance
        """
        self.root = root
        self.file_manager = file_manager
        self.extractor = extractor
        self.logger = logging.getLogger(__name__)
        
        # Queue for pending trays to review
        self.pending_trays = []
        
        # Current tray being reviewed
        self.current_tray = None
        
        # Current compartment being reviewed
        self.current_compartment_index = 0
        
        # Status data for compartments
        self.compartment_statuses = {}
        
        # Review window
        self.review_window = None
        
        # OneDrive path manager with root for dialogs
        self.onedrive_manager = OneDrivePathManager(root)
        
        # Constants for status values
        self.STATUS_OK = "OK"
        self.STATUS_BLURRY = "Blurry"
        self.STATUS_DAMAGED = "Damaged"
        self.STATUS_MISSING = "Missing"
        
    def add_tray_for_review(self, hole_id, depth_from, depth_to, original_path, compartments):
        """
        Add a tray to the review queue.
        
        Args:
            hole_id: The hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            original_path: Path to the original image file
            compartments: List of extracted compartment images
        """
        self.pending_trays.append({
            'hole_id': hole_id,
            'depth_from': depth_from,
            'depth_to': depth_to,
            'original_path': original_path,
            'compartments': compartments.copy(),
            'temp_paths': [],  # Will store temporary saved compartment paths
            'compartment_statuses': {}  # Will store status for each compartment
        })
        
        self.logger.info(f"Added tray for review: {hole_id} {depth_from}-{depth_to}m")
        
    def start_review_process(self):
        """Start the review process for all pending trays."""
        if not self.pending_trays:
            messagebox.showinfo("Review Complete", "No trays to review.")
            return
        
        # Process the first tray
        self._review_next_tray()
        
    def _review_next_tray(self):
        """Display the next tray for review."""
        if not self.pending_trays:
            messagebox.showinfo("Review Complete", "All trays have been reviewed.")
            return
        
        # Get the next tray
        self.current_tray = self.pending_trays.pop(0)
        self.current_compartment_index = 0
        
        # Save compartment images to temporary location
        self._save_temp_compartments()
        
        # Initialize compartment statuses
        self.current_tray['compartment_statuses'] = {
            i: self.STATUS_OK for i in range(len(self.current_tray['compartments']))
        }
        
        # Create review window
        self._create_review_window()
        
        # Show first compartment
        self._show_current_compartment()
        
    def _save_temp_compartments(self):
        """Save compartment images to a temporary location for review."""
        if not self.current_tray:
            return
        
        # Create a temporary directory
        temp_dir = os.path.join(self.file_manager.processed_dir, "Temp_Review")
        os.makedirs(temp_dir, exist_ok=True)
        
        # Clear any previous temp paths
        self.current_tray['temp_paths'] = []
        
        # Save each compartment
        for i, compartment in enumerate(self.current_tray['compartments']):
            try:
                # Calculate compartment depth
                depth_from = self.current_tray['depth_from']
                depth_increment = self.extractor.config['compartment_interval']
                comp_depth_from = depth_from + (i * depth_increment)
                comp_depth_to = comp_depth_from + depth_increment
                compartment_depth = int(comp_depth_to)
                
                # Save temporarily
                filename = f"{self.current_tray['hole_id']}_CC_{compartment_depth}_temp.png"
                file_path = os.path.join(temp_dir, filename)
                
                cv2.imwrite(file_path, compartment)
                self.current_tray['temp_paths'].append(file_path)
                
            except Exception as e:
                self.logger.error(f"Error saving temporary compartment {i+1}: {str(e)}")
                
    def _create_review_window(self):
        """Create a window for reviewing compartments one by one."""
        # Close existing window if open
        if self.review_window and self.review_window.winfo_exists():
            self.review_window.destroy()
        
        # Create new window
        self.review_window = tk.Toplevel(self.root)
        self.review_window.title(f"Review Tray: {self.current_tray['hole_id']} {self.current_tray['depth_from']}-{self.current_tray['depth_to']}m")
        self.review_window.geometry("1000x800")
        self.review_window.protocol("WM_DELETE_WINDOW", self._on_review_window_close)
        
        # Main frame
        main_frame = ttk.Frame(self.review_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title and info
        ttk.Label(
            main_frame,
            text=f"Reviewing: {self.current_tray['hole_id']} {self.current_tray['depth_from']}-{self.current_tray['depth_to']}m",
            font=("Arial", 16, "bold")
        ).pack(pady=10)
        
        ttk.Label(
            main_frame,
            text=f"Original image: {os.path.basename(self.current_tray['original_path'])}",
            font=("Arial", 10)
        ).pack(pady=5)
        
        # Progress label
        self.progress_label = ttk.Label(
            main_frame,
            text="Compartment 1 of X",
            font=("Arial", 12)
        )
        self.progress_label.pack(pady=5)
        
        # Compartment frame
        self.compartment_frame = ttk.Frame(main_frame)
        self.compartment_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Status frame
        status_frame = ttk.LabelFrame(main_frame, text="Compartment Status", padding=10)
        status_frame.pack(fill=tk.X, pady=10)
        
        # Status radio buttons
        self.status_var = tk.StringVar(value=self.STATUS_OK)
        
        status_options = [
            (self.STATUS_OK, "OK - Good quality image"),
            (self.STATUS_BLURRY, "Blurry - Image is not clear"),
            (self.STATUS_DAMAGED, "Damaged - Physical damage visible"),
            (self.STATUS_MISSING, "Missing - No compartment available")
        ]
        
        for status, text in status_options:
            radio = ttk.Radiobutton(
                status_frame,
                text=text,
                value=status,
                variable=self.status_var,
                command=self._on_status_change
            )
            radio.pack(anchor='w', pady=2)
        
        # Navigation buttons frame
        nav_frame = ttk.Frame(main_frame)
        nav_frame.pack(fill=tk.X, pady=10)
        
        # Previous button
        self.prev_button = ttk.Button(
            nav_frame,
            text="← Previous",
            command=self._on_previous
        )
        self.prev_button.pack(side=tk.LEFT, padx=5)
        
        # Next button
        self.next_button = ttk.Button(
            nav_frame,
            text="Next →",
            command=self._on_next
        )
        self.next_button.pack(side=tk.RIGHT, padx=5)
        
        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # Approve button - green
        self.approve_button = tk.Button(
            button_frame,
            text="APPROVE ALL COMPARTMENTS",
            background="green",
            foreground="white",
            font=("Arial", 14, "bold"),
            command=self._on_approve,
            state=tk.DISABLED  # Initially disabled until all compartments reviewed
        )
        self.approve_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=10)
        
        # Reject button - red
        reject_button = tk.Button(
            button_frame,
            text="REJECT ALL",
            background="red",
            foreground="white",
            font=("Arial", 14, "bold"),
            command=self._on_reject
        )
        reject_button.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=5, pady=10)
    
    def _show_current_compartment(self):
        """Display the current compartment for review."""
        # Clear previous compartment display
        for widget in self.compartment_frame.winfo_children():
            widget.destroy()
        
        # Update progress label
        total_compartments = len(self.current_tray['compartments'])
        self.progress_label.config(
            text=f"Compartment {self.current_compartment_index + 1} of {total_compartments}"
        )
        
        # Update status variable
        current_status = self.current_tray['compartment_statuses'].get(
            self.current_compartment_index, self.STATUS_OK
        )
        self.status_var.set(current_status)
        
        # Update navigation buttons
        self.prev_button.config(state=tk.NORMAL if self.current_compartment_index > 0 else tk.DISABLED)
        self.next_button.config(state=tk.NORMAL if self.current_compartment_index < total_compartments - 1 else tk.DISABLED)
        
        # Calculate depth for this compartment
        depth_from = self.current_tray['depth_from']
        depth_increment = self.extractor.config['compartment_interval']
        comp_depth_from = depth_from + (self.current_compartment_index * depth_increment)
        comp_depth_to = comp_depth_from + depth_increment
        
        # Check if this compartment interval exists in database
        existing_image = self._check_for_existing_compartment(
            self.current_tray['hole_id'], 
            int(comp_depth_to)
        )
        
        # Depth label
        ttk.Label(
            self.compartment_frame,
            text=f"Depth: {int(comp_depth_from)}-{int(comp_depth_to)}m",
            font=("Arial", 14, "bold")
        ).pack(pady=(0, 10))
        
        # Image display frame
        if existing_image:
            # Show side-by-side display
            image_frame = ttk.Frame(self.compartment_frame)
            image_frame.pack(fill=tk.BOTH, expand=True)
            
            # Original image column
            orig_col = ttk.Frame(image_frame)
            orig_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
            
            ttk.Label(
                orig_col,
                text="Existing Image",
                font=("Arial", 12, "bold")
            ).pack(pady=(0, 5))
            
            # Load and display existing image
            existing_img = self._load_image_for_display(existing_image, max_size=(400, 400))
            if existing_img:
                ttk.Label(orig_col, image=existing_img).pack(pady=5)
                
            # Keep original option
            self.keep_original_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                orig_col,
                text="Keep Original Image",
                variable=self.keep_original_var
            ).pack(pady=10)
            
            # New image column
            new_col = ttk.Frame(image_frame)
            new_col.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)
            
            ttk.Label(
                new_col,
                text="New Image",
                font=("Arial", 12, "bold")
            ).pack(pady=(0, 5))
            
            # Load and display new image
            if self.current_compartment_index < len(self.current_tray['temp_paths']):
                new_path = self.current_tray['temp_paths'][self.current_compartment_index]
                new_img = self._load_image_for_display(new_path, max_size=(400, 400))
                if new_img:
                    ttk.Label(new_col, image=new_img).pack(pady=5)
        else:
            # Only show new image
            # Load and display the compartment image
            if self.current_compartment_index < len(self.current_tray['temp_paths']):
                new_path = self.current_tray['temp_paths'][self.current_compartment_index]
                new_img = self._load_image_for_display(new_path, max_size=(600, 600))
                if new_img:
                    ttk.Label(self.compartment_frame, image=new_img).pack(pady=5)
    
    def _load_image_for_display(self, path, max_size=(400, 400)):
        """Load an image and resize it for display."""
        try:
            from PIL import Image, ImageTk
            
            # Load image
            img = Image.open(path)
            
            # Resize for display while preserving aspect ratio
            img.thumbnail(max_size, Image.LANCZOS)
            
            # Convert to Tkinter-compatible image
            tk_img = ImageTk.PhotoImage(img)
            
            # Keep a reference to prevent garbage collection
            if not hasattr(self, 'image_references'):
                self.image_references = []
            self.image_references.append(tk_img)
            
            return tk_img
        except Exception as e:
            self.logger.error(f"Error loading image {path}: {str(e)}")
            return None
    
    def _check_for_existing_compartment(self, hole_id, compartment_depth):
        """
        Check if a compartment already exists in the output directory.
        
        Args:
            hole_id: Hole ID
            compartment_depth: Compartment depth identifier
            
        Returns:
            Path to existing image if found, None otherwise
        """
        try:
            # Get the compartments directory
            compartments_dir = os.path.join(
                self.file_manager.dir_structure["chip_compartments"], 
                hole_id
            )
            
            if not os.path.exists(compartments_dir):
                return None
            
            # Look for matching compartment file
            for file in os.listdir(compartments_dir):
                if file.endswith(('.png', '.jpg')) and f"{hole_id}_CC_{compartment_depth}" in file:
                    return os.path.join(compartments_dir, file)
            
            return None
        except Exception as e:
            self.logger.error(f"Error checking for existing compartment: {str(e)}")
            return None
    
    def _on_status_change(self):
        """Handle status change for the current compartment."""
        # Save the status for the current compartment
        self.current_tray['compartment_statuses'][self.current_compartment_index] = self.status_var.get()
        
        # Check if all compartments have been reviewed
        all_reviewed = len(self.current_tray['compartment_statuses']) == len(self.current_tray['compartments'])
        
        # Enable the approve button if all compartments have been reviewed
        if all_reviewed:
            self.approve_button.config(state=tk.NORMAL)
    
    def _on_previous(self):
        """Show the previous compartment."""
        # Save current status
        self.current_tray['compartment_statuses'][self.current_compartment_index] = self.status_var.get()
        
        # Move to previous compartment
        if self.current_compartment_index > 0:
            self.current_compartment_index -= 1
            self._show_current_compartment()
    
    def _on_next(self):
        """Show the next compartment."""
        # Save current status
        self.current_tray['compartment_statuses'][self.current_compartment_index] = self.status_var.get()
        
        # Move to next compartment
        if self.current_compartment_index < len(self.current_tray['compartments']) - 1:
            self.current_compartment_index += 1
            self._show_current_compartment()
        else:
            # We've reached the last compartment, enable the approve button
            self.approve_button.config(state=tk.NORMAL)
    
    def _on_approve(self):
        """Handle approve button click."""
        if not self.current_tray:
            return
        
        try:
            # Save compartments using FileManager and record their status
            self._save_approved_compartments()
            
            # Update Excel register with individual compartment entries
            self._update_excel_register()
            
            # Move original file to Processed Originals
            self._move_original_file(is_approved=True)
            
            # Close review window
            self.review_window.destroy()
            
            # Process next tray
            self._review_next_tray()
            
        except Exception as e:
            self.logger.error(f"Error during approval: {str(e)}")
            self.logger.error(traceback.format_exc())
            messagebox.showerror("Error", f"An error occurred during approval: {str(e)}")
    
    def _on_reject(self):
        """Handle reject button click."""
        if not self.current_tray:
            return
        
        try:
            # Ask for confirmation
            if not messagebox.askyesno("Confirm Rejection", 
                                     "Are you sure you want to reject this tray? "
                                     "The compartment images will be discarded."):
                return
            
            # Move original file to Failed and Skipped Originals
            self._move_original_file(is_approved=False)
            
            # Close review window
            self.review_window.destroy()
            
            # Process next tray
            self._review_next_tray()
            
        except Exception as e:
            self.logger.error(f"Error during rejection: {str(e)}")
            messagebox.showerror("Error", f"An error occurred during rejection: {str(e)}")
    
    def _on_review_window_close(self):
        """Handle review window close event."""
        if messagebox.askyesno("Confirm", "Are you sure you want to exit the review process? "
                             "All remaining trays will be skipped."):
            self.review_window.destroy()
            self.pending_trays = []
    
    def _save_approved_compartments(self):
        """Save the approved compartment images using FileManager."""
        if not self.current_tray:
            return
        
        hole_id = self.current_tray['hole_id']
        depth_from = self.current_tray['depth_from']
        depth_to = self.current_tray['depth_to']
        compartment_interval = self.extractor.config['compartment_interval']
        
        # Count how many compartments were processed
        processed_count = 0
        
        # Save each compartment based on its status
        for i, compartment in enumerate(self.current_tray['compartments']):
            try:
                # Calculate compartment depth
                comp_depth_from = depth_from + (i * compartment_interval)
                comp_depth_to = comp_depth_from + compartment_interval
                compartment_depth = int(comp_depth_to)
                
                # Get status for this compartment
                status = self.current_tray['compartment_statuses'].get(i, self.STATUS_OK)
                
                # Check if we should keep the original for this compartment
                existing_image = self._check_for_existing_compartment(hole_id, compartment_depth)
                keep_original = False
                
                # If there's a comparison being made between existing and new
                if existing_image and hasattr(self, 'keep_original_var') and self.current_compartment_index == i:
                    keep_original = self.keep_original_var.get()
                
                # Don't save if we're keeping the original or if marked as missing
                if keep_original or status == self.STATUS_MISSING:
                    self.logger.info(f"Skipping compartment {i+1}: keep_original={keep_original}, status={status}")
                    continue
                
                # Save using FileManager
                self.file_manager.save_compartment(
                    compartment,
                    hole_id,
                    compartment_depth,
                    False,  # has_data
                    self.extractor.config['output_format']
                )
                processed_count += 1
                
                # Also copy to OneDrive approved folder if path found
                self._copy_to_onedrive_approved(
                    compartment,
                    hole_id,
                    compartment_depth
                )
                
            except Exception as e:
                self.logger.error(f"Error saving approved compartment {i+1}: {str(e)}")
        
        # Log summary
        self.logger.info(f"Saved {processed_count} compartments out of {len(self.current_tray['compartments'])} total")
    
    def _move_original_file(self, is_approved: bool):
        """
        Move the original file to the appropriate directory.
        
        Args:
            is_approved: Whether the tray was approved
        """
        if not self.current_tray:
            return
        
        # Get original file path
        original_path = self.current_tray['original_path']
        if not os.path.exists(original_path):
            self.logger.warning(f"Original file not found: {original_path}")
            return
        
        # Move using FileManager
        self.file_manager.move_original_file(
            original_path,
            self.current_tray['hole_id'],
            self.current_tray['depth_from'],
            self.current_tray['depth_to'],
            is_processed=is_approved
        )
    
    def _copy_to_onedrive_approved(self, image, hole_id, compartment_depth):
        """
        Copy a compartment image to the OneDrive approved folder.
        
        Args:
            image: The compartment image
            hole_id: Hole ID
            compartment_depth: Compartment depth
        """
        try:
            # Get the OneDrive approved folder path
            approved_path = self.onedrive_manager.get_approved_folder_path()
            if not approved_path:
                # Error message already shown by OneDrivePathManager
                return
            
            # Check for hole-specific subfolder within the approved folder
            hole_folder = os.path.join(approved_path, hole_id)
            if not os.path.exists(hole_folder):
                # Prompt with clear instructions
                if self.root is not None:
                    messagebox.showinfo(
                        "Subfolder Not Found", 
                        f"The subfolder for '{hole_id}' was not found in the Approved folder. "
                        f"Please locate it or select the main Approved folder."
                    )
                    
                    from tkinter import filedialog
                    selected_path = filedialog.askdirectory(
                        title=f"Select folder for hole '{hole_id}' or the main Approved folder"
                    )
                    
                    if selected_path:
                        hole_folder = selected_path
                    else:
                        return
                else:
                    # In non-GUI mode, use the main folder
                    hole_folder = approved_path
                    self.logger.warning(f"Hole subfolder not found, using main approved folder: {approved_path}")
            
            # Create filename
            filename = f"{hole_id}_CC_{compartment_depth}.png"
            output_path = os.path.join(hole_folder, filename)
            
            # If file already exists, add a counter to the filename
            counter = 1
            base_name = os.path.splitext(filename)[0]
            ext = os.path.splitext(filename)[1]
            while os.path.exists(output_path):
                new_filename = f"{base_name}_{counter}{ext}"
                output_path = os.path.join(hole_folder, new_filename)
                counter += 1
            
            # Save the image
            cv2.imwrite(output_path, image)
            self.logger.info(f"Copied to OneDrive approved folder: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Error copying to OneDrive: {str(e)}")
            messagebox.showerror(
                "Error", 
                f"Could not copy image to OneDrive approved folder: {str(e)}"
            )

    def _copy_original_to_onedrive(self, original_path, hole_id, depth_from, depth_to):
        """
        Copy the original image to the OneDrive Processed Originals folder.
        
        Args:
            original_path: Path to the original image file
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
        """
        try:
            # Get the OneDrive Processed Originals folder path
            processed_path = self.onedrive_manager.get_processed_originals_path()
            if not processed_path:
                # Error message already shown by OneDrivePathManager
                return
            
            # Check for hole-specific subfolder within the processed folder
            hole_folder = os.path.join(processed_path, hole_id)
            if not os.path.exists(hole_folder):
                # Prompt with clear instructions
                if self.root is not None:
                    messagebox.showinfo(
                        "Subfolder Not Found", 
                        f"The subfolder for '{hole_id}' was not found in the Processed Originals folder. "
                        f"Please locate it or select the main Processed Originals folder."
                    )
                    
                    from tkinter import filedialog
                    selected_path = filedialog.askdirectory(
                        title=f"Select folder for hole '{hole_id}' or the main Processed Originals folder"
                    )
                    
                    if selected_path:
                        hole_folder = selected_path
                    else:
                        return
                else:
                    # In non-GUI mode, use the main folder
                    hole_folder = processed_path
                    self.logger.warning(f"Hole subfolder not found, using main processed folder: {processed_path}")
            
            # Get original file extension
            _, ext = os.path.splitext(original_path)
            
            # Create new filename
            new_filename = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Original{ext}"
            target_path = os.path.join(hole_folder, new_filename)
            
            # Check if file already exists, add number if needed
            counter = 1
            base_name = new_filename
            while os.path.exists(target_path):
                name_parts = os.path.splitext(base_name)
                new_filename = f"{name_parts[0]}_{counter}{name_parts[1]}"
                target_path = os.path.join(hole_folder, new_filename)
                counter += 1
            
            # Copy the file
            shutil.copy2(original_path, target_path)
            self.logger.info(f"Copied original file to OneDrive: {target_path}")
            
        except Exception as e:
            self.logger.error(f"Error copying original file to OneDrive: {str(e)}")
            messagebox.showerror(
                "Error", 
                f"Could not copy original file to OneDrive: {str(e)}"
            )

    def _update_excel_register(self):
        """Update the Excel register with the approved compartments."""
        try:
            # Find the Excel register
            register_path = self.onedrive_manager.get_register_path()
            if not register_path:
                # Error message already shown by OneDrivePathManager
                # Fall back to local storage
                self._save_entries_to_local_file()
                return
            
            # Prepare data to write
            register_entries = []
            
            hole_id = self.current_tray['hole_id']
            depth_from = self.current_tray['depth_from']
            depth_to = self.current_tray['depth_to']
            original_filename = os.path.basename(self.current_tray['original_path'])
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = os.getenv("USERNAME") or "Unknown User"
            
            # Create an entry for each compartment
            compartment_interval = self.extractor.config['compartment_interval']
            
            for i, compartment in enumerate(self.current_tray['compartments']):
                # Calculate compartment depth
                comp_depth_from = depth_from + (i * compartment_interval)
                comp_depth_to = comp_depth_from + compartment_interval
                
                # Get status for this compartment
                status = self.current_tray['compartment_statuses'].get(i, self.STATUS_OK)
                
                # Add to entries list
                register_entries.append({
                    'HoleID': hole_id,
                    'From': int(comp_depth_from),
                    'To': int(comp_depth_to),
                    'Status': status,
                    'Approved Date': timestamp,
                    'Approved By': username
                })
            
            # First, try to process any pending entries from previous runs
            self._process_pending_entries(register_path)
            
            # Try to update the Excel file with retry logic for file locking
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    # First try to open the workbook
                    workbook = None
                    try:
                        from openpyxl import load_workbook
                        workbook = load_workbook(register_path)
                    except Exception as open_error:
                        self.logger.warning(f"Error opening Excel file (attempt {attempt+1}): {str(open_error)}")
                        
                        # Check if it's a file access error and try to close Excel via system command
                        if "Permission denied" in str(open_error) or "being used by another process" in str(open_error):
                            self._try_close_excel()
                            # Wait a moment before trying again
                            import time
                            time.sleep(2)
                            continue
                    
                    if workbook is None:
                        # Try with pandas as a fallback
                        import pandas as pd
                        
                        # Check if file exists
                        if not os.path.exists(register_path):
                            # Create a new DataFrame with the required columns
                            df = pd.DataFrame(columns=[
                                'HoleID', 'From', 'To', 'Status', 'Approved Date', 'Approved By'
                            ])
                        else:
                            # Read existing data
                            df = pd.read_excel(register_path, sheet_name='Processed Images')
                        
                        # Append new data
                        for entry in register_entries:
                            df = pd.concat([df, pd.DataFrame([entry])], ignore_index=True)
                        
                        # Save to Excel
                        with pd.ExcelWriter(register_path, engine='openpyxl') as writer:
                            df.to_excel(writer, sheet_name='Processed Images', index=False)
                            
                        self.logger.info(f"Updated Excel register using pandas: {register_path}")
                        break
                    
                    # Check if "Processed Images" sheet exists
                    if "Processed Images" not in workbook.sheetnames:
                        # Create the sheet
                        sheet = workbook.create_sheet("Processed Images")
                        
                        # Add header row
                        headers = ['HoleID', 'From', 'To', 'Status', 'Approved Date', 'Approved By']
                        for col, header in enumerate(headers, start=1):
                            sheet.cell(row=1, column=col).value = header
                    else:
                        # Get sheet
                        sheet = workbook["Processed Images"]
                    
                    # Find next empty row
                    next_row = sheet.max_row + 1
                    if sheet.cell(row=1, column=1).value is None:  # Empty sheet
                        next_row = 1
                        # Add header row
                        headers = ['HoleID', 'From', 'To', 'Status', 'Approved Date', 'Approved By']
                        for col, header in enumerate(headers, start=1):
                            sheet.cell(row=1, column=col).value = header
                        next_row = 2
                    
                    # Add each entry to sheet
                    for entry in register_entries:
                        sheet.cell(row=next_row, column=1).value = entry['HoleID']
                        sheet.cell(row=next_row, column=2).value = entry['From']
                        sheet.cell(row=next_row, column=3).value = entry['To']
                        sheet.cell(row=next_row, column=4).value = entry['Status']
                        sheet.cell(row=next_row, column=5).value = entry['Approved Date']
                        sheet.cell(row=next_row, column=6).value = entry['Approved By']
                        next_row += 1
                    
                    # Save workbook
                    workbook.save(register_path)
                    self.logger.info(f"Updated Excel register: {register_path}")
                    break
                    
                except Exception as e:
                    self.logger.error(f"Error updating Excel register (attempt {attempt+1}): {str(e)}")
                    if attempt == max_retries - 1:
                        # Last attempt failed, save entries to local file
                        self._save_entries_to_local_file(register_entries)
                        
                        messagebox.showwarning(
                            "Excel Register Update", 
                            "Could not update the Excel register on OneDrive. Entries have been saved locally and will be applied during the next run."
                        )
                    else:
                        # Wait before retrying
                        import time
                        time.sleep(2)
                
        except Exception as e:
            self.logger.error(f"Error in Excel register update: {str(e)}")
            self.logger.error(traceback.format_exc())
            
            # Save entries to local file as a fallback
            self._save_entries_to_local_file(register_entries)
            
            messagebox.showwarning(
                "Excel Register Update", 
                f"An error occurred while updating the Excel register. Entries have been saved locally and will be applied during the next run."
            )
    
    def _try_close_excel(self):
        """Attempt to close Excel processes that might be locking the file."""
        try:
            import platform
            import subprocess
            
            system = platform.system()
            if system == "Windows":
                # Try to close Excel using taskkill
                subprocess.run(["taskkill", "/f", "/im", "excel.exe"], 
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE,
                            shell=True)
                self.logger.info("Attempted to close Excel processes")
            elif system == "Darwin":  # macOS
                subprocess.run(["pkill", "-x", "Microsoft Excel"], 
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE)
                self.logger.info("Attempted to close Excel processes on macOS")
            else:
                self.logger.info(f"No Excel termination support for {system}")
                
        except Exception as e:
            self.logger.error(f"Error trying to close Excel: {str(e)}")

    def _save_entries_to_local_file(self, entries=None):
        """
        Save register entries to a local CSV file when OneDrive update fails.
        
        Args:
            entries: List of entry dictionaries to save (uses current tray if None)
        """
        try:
            # Use entries provided or create from current tray
            if entries is None:
                entries = []
                # Current tray data
                hole_id = self.current_tray['hole_id']
                depth_from = self.current_tray['depth_from']
                depth_to = self.current_tray['depth_to']
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                username = os.getenv("USERNAME") or "Unknown User"
                
                # Create entries for each compartment
                compartment_interval = self.extractor.config['compartment_interval']
                for i, compartment in enumerate(self.current_tray['compartments']):
                    comp_depth_from = depth_from + (i * compartment_interval)
                    comp_depth_to = comp_depth_from + compartment_interval
                    status = self.current_tray['compartment_statuses'].get(i, self.STATUS_OK)
                    
                    entries.append({
                        'HoleID': hole_id,
                        'From': int(comp_depth_from),
                        'To': int(comp_depth_to),
                        'Status': status,
                        'Approved Date': timestamp,
                        'Approved By': username
                    })
            
            # Create local path
            local_dir = self.file_manager.processed_dir
            pending_file = os.path.join(local_dir, "pending_register_entries.csv")
            
            # Convert to DataFrame
            import pandas as pd
            new_df = pd.DataFrame(entries)
            
            # Check if file exists and append
            if os.path.exists(pending_file):
                try:
                    existing_df = pd.read_csv(pending_file)
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                except Exception as e:
                    self.logger.error(f"Error reading existing pending file: {str(e)}")
                    combined_df = new_df
            else:
                combined_df = new_df
            
            # Save to CSV
            combined_df.to_csv(pending_file, index=False)
            self.logger.info(f"Saved {len(entries)} entries to local pending file: {pending_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving entries to local file: {str(e)}")
            self.logger.error(traceback.format_exc())

    def _process_pending_entries(self, register_path):
        """
        Process pending entries from local CSV file and apply to OneDrive register.
        
        Args:
            register_path: Path to the OneDrive Excel register
        """
        # Check for pending entries file
        local_dir = self.file_manager.processed_dir
        pending_file = os.path.join(local_dir, "pending_register_entries.csv")
        
        if not os.path.exists(pending_file):
            return  # No pending entries
        
        try:
            # Read pending entries
            import pandas as pd
            pending_df = pd.read_csv(pending_file)
            
            if pending_df.empty:
                self.logger.info("Pending entries file exists but is empty")
                return
            
            self.logger.info(f"Found {len(pending_df)} pending entries to process")
            
            # Try to update Excel with these entries
            try:
                # Read existing Excel
                excel_df = pd.read_excel(register_path, sheet_name='Processed Images')
                
                # Combine with pending entries
                combined_df = pd.concat([excel_df, pending_df], ignore_index=True)
                
                # Write back to Excel
                with pd.ExcelWriter(register_path, engine='openpyxl') as writer:
                    combined_df.to_excel(writer, sheet_name='Processed Images', index=False)
                
                # Success! Delete the pending file
                os.remove(pending_file)
                self.logger.info(f"Successfully processed {len(pending_df)} pending entries and updated Excel register")
                
            except Exception as e:
                self.logger.error(f"Error processing pending entries: {str(e)}")
                # Keep the pending file for next attempt
        
        except Exception as e:
            self.logger.error(f"Error reading pending entries file: {str(e)}")



class OneDrivePathManager:
    """
    Manages finding OneDrive paths for the project.
    """
    
    def __init__(self, root=None):
        """
        Initialize the OneDrive path manager.
        
        Args:
            root: Optional tkinter root for dialogs
        """
        self.logger = logging.getLogger(__name__)
        self.root = root
        
        # Path patterns to check for OneDrive
        self.onedrive_patterns = [
            # Standard OneDrive pattern
            os.path.join(os.path.expanduser("~"), "OneDrive - Fortescue Metals Group"),
            # Alternative pattern
            os.path.join(os.path.expanduser("~"), "Fortescue Metals Group"),
            # Direct OneDrive pattern
            os.path.join(os.path.expanduser("~"), "OneDrive")
        ]
        
        # Base path for project files
        self.project_path = "Gabon - Belinga - Exploration Drilling"
        self.project_path_alt = os.path.join("Shared Documents", "Exploration Drilling")
        
        # Cache found paths to avoid repetitive searching
        self._approved_folder_path = None
        self._register_path = None
        self._chip_tray_folder_path = None
        self._processed_originals_path = None
        

    def get_processed_originals_path(self) -> Optional[str]:
        """
        Get the path to the Processed Originals folder in OneDrive.
        
        Returns:
            Path to the Processed Originals folder, or None if not found
        """
        if hasattr(self, '_processed_originals_path') and self._processed_originals_path is not None:
            # Use the custom path if it's been set
            if os.path.exists(self._processed_originals_path):
                return self._processed_originals_path
            elif self.root is not None:
                # Prompt user if the custom path doesn't exist
                if messagebox.askyesno(
                    "Path Not Found",
                    f"The custom Processed Originals path does not exist: {self._processed_originals_path}\n\n"
                    "Would you like to browse for the correct location?",
                    icon='warning'
                ):
                    from tkinter import filedialog
                    selected_path = filedialog.askdirectory(
                        title="Select Processed Originals folder"
                    )
                    if selected_path:
                        self._processed_originals_path = selected_path
                        return selected_path
                
        # First make sure we have the Chip Tray Photos folder
        chip_tray_folder = self.get_chip_tray_folder_path()
        if not chip_tray_folder:
            return None
                
        # Found it or selected it!
        self._processed_originals_path = processed_folder
        self.logger.info(f"Found/selected OneDrive Processed Originals folder: {processed_folder}")
        return processed_folder

    def get_chip_tray_folder_path(self) -> Optional[str]:
        """
        Get the path to the main Chip Tray Photos folder in OneDrive.
        If not found, asks the user to select it.
        
        Returns:
            Path to the Chip Tray Photos folder, or None if not found/selected
        """
        if self._chip_tray_folder_path is not None:
            return self._chip_tray_folder_path
            
        # Potential paths based on provided information
        relative_paths = [
            os.path.join(self.project_path, "03 - Reverse Circulation", "Chip Tray Photos"),
            os.path.join(self.project_path_alt, "03 - Reverse Circulation", "Chip Tray Photos")
        ]
        
        # Try to find a valid path automatically
        for base_path in self.onedrive_patterns:
            for rel_path in relative_paths:
                full_path = os.path.join(base_path, rel_path)
                if os.path.exists(full_path):
                    self._chip_tray_folder_path = full_path
                    self.logger.info(f"Found OneDrive Chip Tray Photos folder: {full_path}")
                    return full_path
        
        # If not found automatically, prompt the user
        if self.root is not None:
            try:
                messagebox.showinfo(
                    "Select Folder", 
                    "Please select the 'Chip Tray Photos' folder in OneDrive"
                )
                
                from tkinter import filedialog
                selected_path = filedialog.askdirectory(
                    title="Select 'Chip Tray Photos' folder in OneDrive"
                )
                
                if selected_path:
                    # Verify selected folder name
                    folder_name = os.path.basename(selected_path)
                    if folder_name != "Chip Tray Photos":
                        if not messagebox.askyesno(
                            "Folder Name Mismatch",
                            f"Selected folder name '{folder_name}' doesn't match 'Chip Tray Photos'. Proceed anyway?"
                        ):
                            return None
                    
                    self._chip_tray_folder_path = selected_path
                    self.logger.info(f"User selected Chip Tray Photos folder: {selected_path}")
                    return selected_path
                else:
                    self.logger.warning("User cancelled folder selection")
                    return None
            except Exception as e:
                self.logger.error(f"Error during folder selection: {str(e)}")
                messagebox.showerror(
                    "Error", 
                    f"Could not select Chip Tray Photos folder: {str(e)}"
                )
                return None
        
        self.logger.warning("OneDrive Chip Tray Photos folder not found and no UI to prompt user")
        return None
        
    def get_approved_folder_path(self) -> Optional[str]:
        """
        Get the path to the approved folder in OneDrive.
        
        Returns:
            Path to the approved folder, or None if not found
        """
        if self._approved_folder_path is not None:
            # Use the custom path if it's been set
            if os.path.exists(self._approved_folder_path):
                return self._approved_folder_path
            elif self.root is not None:
                # Prompt user if the custom path doesn't exist
                if messagebox.askyesno(
                    "Path Not Found",
                    f"The custom Approved Images folder path does not exist: {self._approved_folder_path}\n\n"
                    "Would you like to browse for the correct location?",
                    icon='warning'
                ):
                    from tkinter import filedialog
                    selected_path = filedialog.askdirectory(
                        title="Select Approved Images Folder (Chip Tray Photos --> Chip Tray Register and Images --> Approved Compartment Images)"
                    )
                    if selected_path:
                        self._approved_folder_path = selected_path
                        return selected_path
                
        # First make sure we have the Chip Tray Photos folder
        chip_tray_folder = self.get_chip_tray_folder_path()
        if not chip_tray_folder:
            return None
        
    def get_register_path(self) -> Optional[str]:
        """
        Get the path to the Excel register in OneDrive.
        
        Returns:
            Path to the Excel register, or None if not found
        """
        if self._register_path is not None:
            # Use the custom path if it's been set
            if os.path.exists(self._register_path):
                return self._register_path
            elif self.root is not None:
                # Prompt user if the custom path doesn't exist
                if messagebox.askyesno(
                    "Path Not Found",
                    f"The custom Excel register path does not exist: {self._register_path}\n\n"
                    "Would you like to browse for the correct location?",
                    icon='warning'
                ):
                    from tkinter import filedialog
                    selected_path = filedialog.askopenfilename(
                        title="Select Excel Register",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                    )
                    if selected_path:
                        self._register_path = selected_path
                        return selected_path
        
        # First make sure we have the Chip Tray Photos folder
        chip_tray_folder = self.get_chip_tray_folder_path()
        if not chip_tray_folder:
            return None
            
        # Check for required subfolder
        register_folder = os.path.join(chip_tray_folder, "1) Chip Tray Register and Images")
        if not os.path.exists(register_folder):
            if self.root is not None:
                messagebox.showerror(
                    "Folder Not Found", 
                    "The '1) Chip Tray Register and Images' folder was not found in the Chip Tray Photos folder."
                )
            self.logger.error(f"Required subfolder not found: {register_folder}")
            return None
            
        # Now look for the Excel register
        register_filename = "Chip Tray Photo Register (Automatic).xlsx"
        register_path = os.path.join(register_folder, register_filename)
        
        if not os.path.exists(register_path):
            if self.root is not None:
                messagebox.showerror(
                    "File Not Found", 
                    f"The '{register_filename}' file was not found in the Chip Tray Register and Images folder."
                )
            self.logger.error(f"Excel register not found: {register_path}")
            return None
            
        # Found it!
        self._register_path = register_path
        self.logger.info(f"Found Excel register: {register_path}")
        return register_path


class DuplicateHandler:
    """
    Manages detection and handling of duplicate image processing entries.
    
    Tracks processed entries to prevent unintentional duplicate processing.
    """
    # TODO - Can we make this able to identify if there are missing compartment images within the tray? Therefore we can reject some of the compartments and process another image only extracting the rejected compartments from the original image?
    def __init__(self, output_dir: str):
        """
        Initialize the duplicate handler.
        
        Args:
            output_dir: Directory where processed images are saved
        """
        self.output_dir = output_dir
        self.processed_entries = self._load_existing_entries()
        self.logger = logging.getLogger(__name__)
    
    def _generate_entry_key(self, hole_id: str, depth_from: float, depth_to: float) -> str:
        """
        Generate a unique key for a processed entry.
        
        Args:
            hole_id: Unique hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
        
        Returns:
            Standardized key representing the entry
        """
        return f"{hole_id.upper()}_{depth_from:.1f}-{depth_to:.1f}"
    
    def _load_existing_entries(self) -> Dict[str, List[str]]:
        """
        Load existing processed entries from saved files in the output directory.
        
        Returns:
            Dictionary of processed entries
        """
        entries = {}
        
        # Safety check for output directory
        if not os.path.exists(self.output_dir):
            return entries
        
        # Find all files in output directory that match pattern
        for filename in os.listdir(self.output_dir):
            # Try to extract metadata from filename - handle both new and old naming patterns
            match = re.search(r'([A-Z]{2}\d{4}).*?(\d+\.?\d*)-(\d+\.?\d*)', filename)
            if match:
                hole_id, depth_from, depth_to = match.groups()
                key = self._generate_entry_key(hole_id, float(depth_from), float(depth_to))
                
                # Track files for this key
                if key not in entries:
                    entries[key] = []
                entries[key].append(filename)
        
        return entries
    
    def check_debug_image_exists(self, 
                                hole_id: str, 
                                depth_from: float, 
                                depth_to: float) -> Optional[str]:
        """
        Check if a debug small image already exists for this hole ID and depth range.
        
        Args:
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            
        Returns:
            Path to existing debug image, or None if not found
        """
        # Build the path to where debug images would be stored
        debug_dir = os.path.join(self.output_dir, "Debug Images", hole_id)
        
        if not os.path.exists(debug_dir):
            return None
            
        # Look for debug small image files
        for filename in os.listdir(debug_dir):
            if f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Debug_small_image" in filename:
                return os.path.join(debug_dir, filename)
        
        return None
    
    def check_duplicate(self, 
                    hole_id: str, 
                    depth_from: float, 
                    depth_to: float,
                    small_image: np.ndarray,
                    full_filename: str) -> Union[bool, Dict[str, Any]]:
        """
        Check if an entry is a potential duplicate and prompt user.
        
        Args:
            hole_id: Unique hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
            small_image: Downsampled image for comparison
            full_filename: Full path to the original image file
        
        Returns:
            bool or Dict: True if processing should continue (replace),
                        False if processing should be skipped,
                        or a dictionary with modified metadata
        """
        # Reload existing entries to ensure we have the latest data
        self.processed_entries = self._load_existing_entries()
        
        # Generate key for current entry
        entry_key = self._generate_entry_key(hole_id, depth_from, depth_to)
        
        # Log for debugging
        self.logger.info(f"Checking for duplicates with key: {entry_key}")
        self.logger.info(f"Known entries: {list(self.processed_entries.keys())}")
        
        # Check if this exact entry exists
        is_duplicate = False
        existing_debug_image = None
        self.duplicate_files = []  # Store duplicate files as instance variable
        
        # Check for existing debug image
        existing_debug_image = self.check_debug_image_exists(hole_id, depth_from, depth_to)
        if existing_debug_image:
            is_duplicate = True
            self.duplicate_files.append(existing_debug_image)
            self.logger.info(f"Found existing debug image for {hole_id} {depth_from}-{depth_to}m")
            
        # Also check based on entry key in processed entries
        if entry_key in self.processed_entries:
            is_duplicate = True
            self.duplicate_files.extend(self.processed_entries[entry_key])
            self.logger.info(f"Found entry in processed list for {hole_id} {depth_from}-{depth_to}m: {len(self.processed_entries[entry_key])} files")
        
        # Check for overlapping depth ranges with the same hole ID
        for key, files in self.processed_entries.items():
            # If it's the same key we already checked above, so skip
            if entry_key == key:
                continue
                
            # Check if this entry has the same hole ID
            if key.startswith(f"{hole_id.upper()}_"):
                # Extract depth range from the key
                key_match = re.search(r'([A-Z]{2}\d{4})_(\d+\.?\d*)-(\d+\.?\d*)', key)
                if key_match:
                    key_hole_id, key_from, key_to = key_match.groups()
                    key_from = float(key_from)
                    key_to = float(key_to)
                    
                    # Check if ranges overlap
                    if (depth_from <= key_to and depth_to >= key_from):
                        is_duplicate = True
                        self.duplicate_files.extend(files)
                        self.logger.info(f"Found overlapping entry: {key} (current: {entry_key})")
        
        # Check Chip Compartment folder for existing compartments
        try:
            # Construct path to the chip compartment folder for this hole
            compartment_dir = os.path.join(self.output_dir, "Chip Compartments", hole_id)
            
            # Check if the directory exists
            if os.path.exists(compartment_dir) and os.path.isdir(compartment_dir):
                self.logger.info(f"Checking compartment directory: {compartment_dir}")
                
                # Calculate depth range covered by this image
                # Convert to integers for comparison
                start_depth = int(depth_from)
                end_depth = int(depth_to)
                
                # Calculate the compartment depths for this range
                # For a range like 40-60, compartments would be at 41, 42, ..., 60
                # Each compartment is labeled with its end depth
                compartment_depths = list(range(start_depth + 1, end_depth + 1))
                
                self.logger.info(f"Image covers compartments at depths: {compartment_depths}")
                
                # Get all compartment files in the directory
                missing_compartments = set(compartment_depths)  # Start with all compartments as missing
                existing_compartments = []
                
                for root, dirs, files in os.walk(compartment_dir):
                    for file in files:
                        if file.endswith(('.png', '.jpg', '.jpeg', '.tif', '.tiff')):
                            # Look for the pattern HoleID_CC_XXX
                            match = re.search(rf'{hole_id}_CC_(\d+)', file)
                            if match:
                                compartment_depth = int(match.group(1))
                                
                                # Check if this depth is in our compartment depths
                                if compartment_depth in compartment_depths:
                                    compartment_path = os.path.join(root, file)
                                    existing_compartments.append({
                                        'depth': compartment_depth,
                                        'path': compartment_path
                                    })
                                    missing_compartments.discard(compartment_depth)
                                    self.logger.info(f"Found existing compartment at depth {compartment_depth}: {file}")
                
                # If we found existing compartments, mark as duplicate but show only if we have all of them
                if existing_compartments:
                    if not missing_compartments:
                        # We have all compartments - mark as full duplicate
                        is_duplicate = True
                        self.duplicate_files.extend([comp['path'] for comp in existing_compartments])
                        self.logger.info(f"Found all {len(existing_compartments)} compartments for {hole_id} between {start_depth}-{end_depth}m")
                    else:
                        # We have some compartments but not all
                        # Mark as partial duplicate
                        self.is_partial_duplicate = True
                        self.existing_compartments = existing_compartments
                        self.missing_compartments = missing_compartments
                        self.logger.info(f"Found partial match: {len(existing_compartments)} existing compartments, {len(missing_compartments)} missing")
                        
                        # Return True but store partial info for QAQC to handle
                        if hasattr(self, 'parent') and hasattr(self.parent, 'visualization_cache'):
                            self.parent.visualization_cache['partial_duplicate_info'] = {
                                'existing_compartments': existing_compartments,
                                'missing_compartments': missing_compartments
                            }
                        
                        # We'll continue with processing but flag it for QAQC to handle differently
                        return True
        except Exception as e:
            self.logger.error(f"Error checking compartment directory: {str(e)}")
        
        # Log additional info about existing files
        if self.duplicate_files:  # Use self.duplicate_files instead of duplicate_files
            self.logger.info(f"Found {len(self.duplicate_files)} duplicate files")
            # Limit the number of files logged to avoid overwhelming the log
            for file in self.duplicate_files[:5]:  # Log up to 5 files
                self.logger.info(f"Existing file: {file}")
            if len(self.duplicate_files) > 5:
                self.logger.info(f"...and {len(self.duplicate_files) - 5} more files")
        
        if is_duplicate:
            # Create duplicate resolution dialog
            self.logger.info(f"Duplicate detected for {entry_key}, showing dialog")
            result = self._show_duplicate_dialog(
                hole_id, 
                depth_from, 
                depth_to, 
                small_image, 
                existing_debug_image
            )
            
            # Return the result directly - it will be either:
            # - False to skip processing
            # - True to continue processing (replacing existing)
            # - A dictionary with new metadata
            return result
        
        return True  # No duplicate found, continue processing
    

    def check_existing_compartment_statuses(self, hole_id, depth_from, depth_to):
        """
        Check Excel register for existing compartment statuses.
        
        Args:
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            
        Returns:
            Dictionary mapping compartment depths to their statuses
        """
        try:
            # Find the Excel register
            register_path = self.onedrive_manager.get_register_path()
            if not register_path or not os.path.exists(register_path):
                return {}
            
            # Try to open the Excel file
            import pandas as pd
            try:
                df = pd.read_excel(register_path, sheet_name='Processed Images')
            except Exception as e:
                self.logger.warning(f"Could not read Excel register: {str(e)}")
                return {}
            
            # Filter for this hole
            hole_data = df[df['HoleID'] == hole_id]
            if hole_data.empty:
                return {}
            
            # Calculate depths in this range
            start_depth = int(depth_from)
            end_depth = int(depth_to)
            depths = list(range(start_depth + 1, end_depth + 1))
            
            # Check for compartments in this range
            status_dict = {}
            for _, row in hole_data.iterrows():
                if 'From' in row and 'To' in row and 'Status' in row:
                    comp_from = row['From']
                    comp_to = row['To']
                    status = row['Status']
                    
                    # Check if this compartment is in our range
                    if int(comp_to) in depths:
                        status_dict[int(comp_to)] = status
            
            return status_dict
            
        except Exception as e:
            self.logger.error(f"Error checking compartment statuses: {str(e)}")
            return {}
    
    def _show_duplicate_dialog(self, 
                            hole_id: str, 
                            depth_from: float, 
                            depth_to: float, 
                            current_image: np.ndarray, 
                            existing_image_path: Optional[str] = None) -> Union[bool, Dict[str, Any]]:
        """
        Show dialog for duplicate resolution with improved UI.
        
        Args:
            hole_id: Unique hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
            current_image: Current image being processed
            existing_image_path: Path to existing image for comparison
            
        Returns:
            Either a boolean (False to skip, True to continue) or a dict with new metadata
        """
        # Create a Tkinter dialog for duplicate resolution
        dialog = tk.Toplevel()
        dialog.title("Duplicate Entry Detected")
        dialog.geometry("1200x900")
        
        # Store the result as a class attribute of the dialog for retrieval later
        dialog.result = None
        
        # Create modified_metadata dict to store any changes
        modified_metadata = {
            'hole_id': tk.StringVar(value=hole_id),
            'depth_from': tk.StringVar(value=str(depth_from)),
            'depth_to': tk.StringVar(value=str(depth_to))
        }
        
        # Create frames
        top_frame = ttk.Frame(dialog, padding=10)
        top_frame.pack(fill=tk.X)
        
        image_frame = ttk.Frame(dialog, padding=10)
        image_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add a frame for metadata editing (initially hidden)
        metadata_frame = ttk.Frame(dialog, padding=10)
        
        button_frame = ttk.Frame(dialog, padding=10)
        button_frame.pack(fill=tk.X)
        
        # Warning message
        warning_label = ttk.Label(
            top_frame, 
            text=f"Duplicate detected for Hole {hole_id}, Depth {depth_from}-{depth_to}m",
            font=("Arial", 12, "bold")
        )
        warning_label.pack(pady=10)
        
        # Load and display existing image if available
        from PIL import Image, ImageTk
        
        # Create a label for existing image
        existing_label = ttk.Label(
            image_frame,
            text="Existing Image:",
            font=("Arial", 10, "bold")
        )
        existing_label.pack(pady=(0, 5))
        
        # Flag to track if we successfully loaded an existing image
        existing_image_found = False
        existing_img_label = None
        
        # First try to find the original processed image in the Processed Originals directory
        processed_originals_dir = os.path.join(self.output_dir, "Processed Originals", hole_id)
        if os.path.exists(processed_originals_dir):
            try:
                # Look for an original file that matches the depth range
                for file in os.listdir(processed_originals_dir):
                    # Look for a pattern like "HoleID_Depth-Depth_Original.ext"
                    pattern = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Original"
                    if pattern in file:
                        original_path = os.path.join(processed_originals_dir, file)
                        # Load and display the original image
                        self.logger.info(f"Found existing original image: {original_path}")
                        
                        # Load image with OpenCV
                        original_img = cv2.imread(original_path)
                        if original_img is not None:
                            # Convert to RGB for PIL
                            original_rgb = cv2.cvtColor(original_img, cv2.COLOR_BGR2RGB)
                            
                            # Resize for display (original files can be quite large)
                            h, w = original_rgb.shape[:2]
                            max_height = 300
                            if h > max_height:
                                scale = max_height / h
                                new_width = int(w * scale)
                                original_rgb = cv2.resize(original_rgb, (new_width, max_height), interpolation=cv2.INTER_AREA)
                            
                            # Convert to PIL and then to ImageTk
                            pil_img = Image.fromarray(original_rgb)
                            tk_img = ImageTk.PhotoImage(image=pil_img)
                            
                            # Display image
                            existing_img_label = ttk.Label(image_frame, image=tk_img)
                            existing_img_label.image = tk_img  # Keep reference
                            existing_img_label.pack(pady=(0, 20))
                            
                            existing_image_found = True
                            break
            except Exception as e:
                self.logger.error(f"Error looking for original image: {str(e)}")
        
        # If no original file was found, fall back to compartment files
        if not existing_image_found and hasattr(self, 'duplicate_files') and self.duplicate_files:
            try:
                # Find first compartment image to show as preview
                for path in self.duplicate_files:
                    if path and os.path.exists(path):
                        # Load image with OpenCV
                        existing_img = cv2.imread(path)
                        if existing_img is not None:
                            # Convert to RGB for PIL
                            existing_rgb = cv2.cvtColor(existing_img, cv2.COLOR_BGR2RGB)
                            
                            # Resize if needed
                            h, w = existing_rgb.shape[:2]
                            max_height = 300
                            if h > max_height:
                                scale = max_height / h
                                new_width = int(w * scale)
                                existing_rgb = cv2.resize(existing_rgb, (new_width, max_height))
                            
                            # Convert to PIL and then to ImageTk
                            pil_img = Image.fromarray(existing_rgb)
                            tk_img = ImageTk.PhotoImage(image=pil_img)
                            
                            # Display image
                            existing_img_label = ttk.Label(image_frame, image=tk_img)
                            existing_img_label.image = tk_img  # Keep reference
                            existing_img_label.pack(pady=(0, 20))
                            
                            existing_image_found = True
                            
                            # Add text explaining this is a sample compartment
                            ttk.Label(
                                image_frame,
                                text=f"(Sample compartment - full original not available)",
                                font=("Arial", 9, "italic"),
                                foreground="gray"
                            ).pack(pady=(0, 20))
                            
                            break  # Only need one image for preview
                
                if not existing_image_found:
                    # If no image path provided, show message
                    ttk.Label(
                        image_frame,
                        text="No existing image available for preview",
                        foreground="gray"
                    ).pack(pady=(0, 20))
            except Exception as e:
                self.logger.error(f"Error loading existing image: {str(e)}")
                # Show error message if image can't be loaded
                ttk.Label(
                    image_frame,
                    text=f"Error loading existing image: {str(e)}",
                    foreground="red"
                ).pack(pady=(0, 20))
        elif not existing_image_found:
            # If no duplicate files found, show message
            ttk.Label(
                image_frame,
                text="No existing image available for preview",
                foreground="gray"
            ).pack(pady=(0, 20))
        
        # Create a label for current image
        current_label = ttk.Label(
            image_frame,
            text="Current Image:",
            font=("Arial", 10, "bold")
        )
        current_label.pack(pady=(0, 5))
        
        # Display current image
        try:
            # Resize current image for display
            h, w = current_image.shape[:2]
            max_height = 300
            if h > max_height:
                scale = max_height / h
                new_width = int(w * scale)
                current_resized = cv2.resize(current_image, (new_width, max_height))
            else:
                current_resized = current_image.copy()
            
            # Convert to RGB for PIL
            if len(current_resized.shape) == 3:
                current_rgb = cv2.cvtColor(current_resized, cv2.COLOR_BGR2RGB)
            else:
                current_rgb = cv2.cvtColor(current_resized, cv2.COLOR_GRAY2RGB)
            
            # Convert to PIL and then to ImageTk
            current_pil = Image.fromarray(current_rgb)
            current_tk = ImageTk.PhotoImage(image=current_pil)
            
            # Display current image
            current_img_label = ttk.Label(image_frame, image=current_tk)
            current_img_label.image = current_tk  # Keep reference
            current_img_label.pack()
        except Exception as e:
            self.logger.error(f"Error displaying current image: {str(e)}")
            ttk.Label(
                image_frame,
                text=f"Error displaying current image: {str(e)}",
                foreground="red"
            ).pack(pady=(0, 10))
        
        
        # Function to show/hide metadata editor
        def toggle_metadata_editor(show=False):
            if show:
                metadata_frame.pack(fill=tk.X, before=button_frame)
            else:
                metadata_frame.pack_forget()
        
        # Create metadata editor fields
        ttk.Label(
            metadata_frame,
            text="Edit Metadata:",
            font=("Arial", 10, "bold")
        ).pack(anchor="w", pady=(10, 5))
        
        # Hole ID field
        hole_id_frame = ttk.Frame(metadata_frame)
        hole_id_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(hole_id_frame, text="Hole ID:", width=15).pack(side=tk.LEFT)
        ttk.Entry(hole_id_frame, textvariable=modified_metadata['hole_id'], width=20).pack(side=tk.LEFT, padx=5)
        
        # Depth fields
        depth_frame = ttk.Frame(metadata_frame)
        depth_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(depth_frame, text="Depth Range:", width=15).pack(side=tk.LEFT)
        ttk.Entry(depth_frame, textvariable=modified_metadata['depth_from'], width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(depth_frame, text="-").pack(side=tk.LEFT)
        ttk.Entry(depth_frame, textvariable=modified_metadata['depth_to'], width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(depth_frame, text="m").pack(side=tk.LEFT)
        
        # Direct action functions for buttons
        def on_skip():
            dialog.result = False  # Skip processing
            dialog.destroy()
        
        def on_keep_image1():
            dialog.result = False  # Also skip processing (keep existing)
            dialog.destroy()
        
        def on_keep_image2():
            dialog.result = True  # Process this image (replace existing)
            dialog.destroy()
        
        def on_modify():
            # Show metadata editor
            toggle_metadata_editor(True)
        
        def on_apply_metadata():
            # Validate metadata
            try:
                new_hole_id = modified_metadata['hole_id'].get()
                new_depth_from = float(modified_metadata['depth_from'].get())
                new_depth_to = float(modified_metadata['depth_to'].get())
                
                # Basic validation
                if not new_hole_id or new_depth_from >= new_depth_to:
                    raise ValueError("Invalid metadata values")
                
                # Set result to modified metadata dictionary
                dialog.result = {
                    'hole_id': new_hole_id,
                    'depth_from': new_depth_from,
                    'depth_to': new_depth_to
                }
                dialog.destroy()
                
            except (ValueError, TypeError) as e:
                messagebox.showerror("Invalid Input", f"Invalid metadata values: {str(e)}")
        
        def on_cancel():
            dialog.result = False  # Default to skip if canceled
            dialog.destroy()
        
        # Create buttons with improved labeling
        skip_button = ColorButton(
            button_frame, 
            text="Skip Processing This Image", 
            background="#ffcccc",  # Pale red
            command=on_skip
        )
        skip_button.pack(side=tk.LEFT, padx=5, pady=10, fill=tk.X, expand=True)
        
        keep1_button = ColorButton(
            button_frame, 
            text="Keep Existing Image", 
            background="#ccffcc",  # Pale green
            command=on_keep_image1
        )
        keep1_button.pack(side=tk.LEFT, padx=5, pady=10, fill=tk.X, expand=True)
        
        keep2_button = ColorButton(
            button_frame, 
            text="Replace with New Image", 
            background="#ccccff",  # Pale blue
            command=on_keep_image2
        )
        keep2_button.pack(side=tk.LEFT, padx=5, pady=10, fill=tk.X, expand=True)
        
        modify_button = ColorButton(
            button_frame, 
            text="Modify Metadata", 
            background="#ffffcc",  # Pale yellow
            command=on_modify
        )
        modify_button.pack(side=tk.LEFT, padx=5, pady=10, fill=tk.X, expand=True)
        
        # Add confirm and cancel buttons to metadata frame
        metadata_button_frame = ttk.Frame(metadata_frame)
        metadata_button_frame.pack(fill=tk.X, pady=10)
        
        apply_button = ttk.Button(
            metadata_button_frame,
            text="Apply Changes",
            command=on_apply_metadata
        )
        apply_button.pack(side=tk.RIGHT, padx=5)
        
        cancel_button = ttk.Button(
            metadata_button_frame,
            text="Cancel",
            command=lambda: toggle_metadata_editor(False)
        )
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        # Wait for dialog
        dialog.wait_window()
        
        # Return the result
        return dialog.result

    def _set_decision_and_close(self, 
                              dialog: tk.Toplevel, 
                              decision_var: tk.StringVar, 
                              value: str) -> None:
        """
        Helper method to set the decision value and close the dialog.
        
        Args:
            dialog: Dialog window
            decision_var: StringVar to store decision
            value: Decision value to set
        """
        decision_var.set(value)
        dialog.destroy()
    
    def register_processed_entry(self, 
                               hole_id: str, 
                               depth_from: float, 
                               depth_to: float, 
                               output_files: List[str]) -> None:
        """
        Register a successfully processed entry.
        
        Args:
            hole_id: Unique hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
            output_files: List of files generated for this entry
        """
        entry_key = self._generate_entry_key(hole_id, depth_from, depth_to)
        self.processed_entries[entry_key] = output_files

class ChipTrayExtractor:
    """
    Main class for extracting chip tray compartments from panoramic images.
    """

    def __init__(self):
        """Initialize the chip tray extractor with default settings."""
        self.logger = getattr(self, 'logger', logging.getLogger(__name__))
        self.progress_queue = queue.Queue()
        self.processing_complete = False
        self.root = None
        
        # Initialize visualization cache for dialogs
        self.visualization_cache = {}

        # Initialize Tesseract manager
        self.tesseract_manager = TesseractManager()
        self.tesseract_manager.extractor = self  # Give it access to visualization_cache

        # Initialize File Manager
        self.file_manager = FileManager()

        # Don't initialize DuplicateHandler yet - it requires an output_dir
        # that is only known when processing an actual image
        self.duplicate_handler = None  # Will be initialized as needed
        
        # Initialize UpdateChecker
        self.update_checker = UpdateChecker()

        # Configuration settings (these can be modified via the GUI)
        self.config = {
            # Output settings
            'output_folder': 'extracted_compartments',
            'save_debug_images': False,
            'output_format': 'png',
            'jpeg_quality': 100,

                # Add blur detection settings to config
            'enable_blur_detection': True,
            'blur_threshold':207.24,  # Default threshold for Laplacian variance
            'blur_roi_ratio': 0.8,    # Use 80% of center image area for blur detection
            'flag_blurry_images': False,  # Whether to visually flag blurry images
            'blurry_threshold_percentage': 10.0,  # Percentage of blurry compartments to flag the tray
            'save_blur_visualizations': True,  # Whether to save blur detection visualizations
            
            # ArUco marker settings
            'aruco_dict_type': cv2.aruco.DICT_4X4_1000,
            'corner_marker_ids': [0, 1, 2, 3],  # Top-left, top-right, bottom-right, bottom-left
            'compartment_marker_ids': list(range(4, 24)),  # 20 markers for compartments
            'metadata_marker_id': 24,  # Marker ID between hole ID and depth labels
            
            # Processing settings
            'compartment_count': 20,
            'compartment_interval': 1.0,  # Default interval between compartments in meters
            
            # OCR settings
            'enable_ocr': self.tesseract_manager.is_available,
            'ocr_confidence_threshold': 70.0,  # Minimum confidence to accept OCR results
            'ocr_config_options': [
                '--psm 11 --oem 3',  # Sparse text, LSTM engine
                '--psm 6 --oem 3',   # Assume a single block of text
                '--psm 3 --oem 3',   # Fully automatic page segmentation
            ],
            'use_metadata_for_filenames': True,
            'metadata_filename_pattern': '{hole_id}_CC_{depth_from}-{depth_to}m',
            'prompt_for_metadata': True,  # Ask user to confirm/correct OCR results
            'valid_hole_prefixes': ['BA', 'NB', 'SB', 'KM'],  # Add your valid prefixes here
            'enable_prefix_validation': True,  # Toggle this feature on/off
        }
        
        # Pass config to tesseract manager
        self.tesseract_manager.config = self.config

        # Initialize ArUco detector
        self.aruco_dict = cv2.aruco.getPredefinedDictionary(self.config['aruco_dict_type'])
        self.aruco_params = cv2.aruco.DetectorParameters()
        self.aruco_detector = cv2.aruco.ArucoDetector(self.aruco_dict, self.aruco_params)

        # Initialize blur detector
        self.blur_detector = BlurDetector(
            threshold=self.config['blur_threshold'],
            roi_ratio=self.config['blur_roi_ratio']
        )

        # Initialize DrillholeTraceGenerator with FileManager reference
        self.trace_generator = DrillholeTraceGenerator(
            config=self.config, 
            progress_queue=self.progress_queue, 
            root=self.root,
            file_manager=self.file_manager  # Pass the FileManager instance
        )


    def select_folder(self) -> Optional[str]:
        """Open folder picker dialog and return selected folder path."""
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        folder_path = filedialog.askdirectory(title="Select folder with chip tray photos")
        return folder_path if folder_path else None

    
    def get_debug_dir(self, base_path: str) -> str:
        """
        Get the path to the debug directory based on output settings.
        
        Args:
            base_path: Base path of the input file or directory
            
        Returns:
            Path to the debug directory
        """
        # Always use the FileManager's debug directory structure
        if hasattr(self, 'file_manager'):
            return self.file_manager.dir_structure["debug_images"]
        
        # Log a warning if FileManager is not available (this should rarely happen)
        logger.warning("FileManager not available for debug directory. Using temporary location.")
        
        # Use a specific temp directory instead of creating one in the input folder
        temp_debug_dir = os.path.join(os.path.expanduser("~"), "TempChipTrayDebug")
        os.makedirs(temp_debug_dir, exist_ok=True)
        return temp_debug_dir

    def _handle_metadata_dialog_on_main_thread(self, ocr_metadata: Dict[str, Any], 
                                            result_queue: queue.Queue) -> None:
        """
        Helper method to show metadata dialog on the main thread and put result in queue.
        
        Args:
            ocr_metadata: OCR extracted metadata
            result_queue: Queue to put the result in
        """
        try:
            # Ensure we have a root window for the dialog
            if not hasattr(self, 'root') or self.root is None:
                logger.warning("No GUI root window available for metadata dialog")
                result_queue.put(None)
                return

            # Convert any depths to integers
            if 'depth_from' in ocr_metadata and ocr_metadata['depth_from'] is not None:
                ocr_metadata['depth_from'] = int(ocr_metadata['depth_from'])
            
            if 'depth_to' in ocr_metadata and ocr_metadata['depth_to'] is not None:
                ocr_metadata['depth_to'] = int(ocr_metadata['depth_to'])

            # ALWAYS ensure all required visualization images exist, regardless of debug settings
            if 'metadata_region_viz' not in ocr_metadata or ocr_metadata['metadata_region_viz'] is None:
                # Create a basic visualization if missing
                if 'metadata_region' in ocr_metadata and ocr_metadata['metadata_region'] is not None:
                    # Create a simple visualization from the metadata region
                    metadata_viz = ocr_metadata['metadata_region'].copy()
                    if len(metadata_viz.shape) == 2:  # Convert grayscale to BGR
                        metadata_viz = cv2.cvtColor(metadata_viz, cv2.COLOR_GRAY2BGR)
                    
                    # Add a border around it to make it clearer
                    metadata_viz = cv2.copyMakeBorder(metadata_viz, 20, 20, 20, 20, 
                                                    cv2.BORDER_CONSTANT, value=(0, 0, 255))
                    
                    # Add this visualization to the metadata
                    ocr_metadata['metadata_region_viz'] = metadata_viz
                elif hasattr(self, 'small_image') and self.small_image is not None:
                    # Use the small image as the visualization with annotations
                    metadata_viz = self.small_image.copy()
                    
                    # Add text to indicate that no metadata region was found
                    h, w = metadata_viz.shape[:2]
                    
                    # Add a semi-transparent overlay at the top for text
                    overlay = metadata_viz.copy()
                    cv2.rectangle(overlay, (0, 0), (w, 60), (0, 0, 255), -1)
                    
                    # Add text
                    cv2.putText(overlay, "No metadata region detected - using full image", 
                            (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
                    
                    # Blend the overlay with the original
                    alpha = 0.7
                    metadata_viz = cv2.addWeighted(overlay, alpha, metadata_viz, 1 - alpha, 0)
                    
                    ocr_metadata['metadata_region_viz'] = metadata_viz
                else:
                    # If we don't even have a small image, create a blank image with error text
                    metadata_viz = np.ones((300, 500, 3), dtype=np.uint8) * 255
                    cv2.putText(metadata_viz, "No metadata region detected", 
                            (50, 150), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                    ocr_metadata['metadata_region_viz'] = metadata_viz
                    
            # Now save these visualization images to an in-memory dictionary for later use
            # This ensures they're available for dialogs even if debug image saving is disabled
            if not hasattr(self, 'visualization_cache'):
                self.visualization_cache = {}
                
            cache_key = f"{ocr_metadata.get('hole_id', 'unknown')}_{ocr_metadata.get('depth_from', 0)}-{ocr_metadata.get('depth_to', 0)}"
            self.visualization_cache[cache_key] = {
                'metadata_region_viz': ocr_metadata.get('metadata_region_viz'),
                'metadata_region': ocr_metadata.get('metadata_region'),
                'small_image': getattr(self, 'small_image', None)
            }

            # Create dialog with both metadata_region and metadata_region_viz
            dialog = MetadataInputDialog(
                parent=self.root,
                image=ocr_metadata.get('metadata_region_viz'),  # Pass the visualization image
                metadata=ocr_metadata,
                tesseract_manager=self.tesseract_manager  # Pass the TesseractManager
            )
            
            # Show dialog and get result
            result = dialog.show()
            
            # If the user cancels, confirm cancellation
            if result is None:
                confirm = messagebox.askyesno(
                    "Confirm Cancellation",
                    "Are you sure you want to skip processing this image?\n\n"
                    "This will be logged as a failed processing attempt.",
                    icon='warning'
                )
                
                if confirm:
                    # User confirmed cancellation
                    logger.warning("User cancelled metadata input, skipping image processing")
                    result_queue.put(None)  # Signal that processing should be aborted
                else:
                    # User changed their mind, show the dialog again
                    self._handle_metadata_dialog_on_main_thread(ocr_metadata, result_queue)
                    return
            else:
                # User provided metadata
                result_queue.put(result)
            
        except Exception as e:
            logger.error(f"Error creating metadata dialog: {str(e)}")
            logger.error(traceback.format_exc())
            result_queue.put(None)

    def detect_blur_in_compartments(self, 
                                compartments: List[np.ndarray], 
                                base_filename: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Detect blur in extracted compartment images.
        
        Args:
            compartments: List of compartment images
            base_filename: Original filename for generating visualization filenames
            
        Returns:
            List of dictionaries with blur analysis results
        """
        if not self.config['enable_blur_detection']:
            # Return empty results if blur detection is disabled
            return [{'index': i, 'is_blurry': False, 'variance': 0.0} for i in range(len(compartments))]
        
        # Configure blur detector with current settings
        self.blur_detector.threshold = self.config['blur_threshold']
        self.blur_detector.roi_ratio = self.config['blur_roi_ratio']
        
        # Analyze all compartments - always generate visualizations
        # They will be saved later in save_compartments using FileManager
        generate_viz = self.config['save_blur_visualizations']
        blur_results = self.blur_detector.batch_analyze_images(compartments, generate_viz)
        
        # Log summary
        blurry_count = sum(1 for result in blur_results if result.get('is_blurry', False))
        logger.info(f"Blur detection: {blurry_count}/{len(compartments)} compartments are blurry")
        
        # Update progress queue if available
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((f"Blur detection: {blurry_count}/{len(compartments)} compartments are blurry", None))
        
        return blur_results
    
    def add_blur_indicators(self, 
                        compartment_images: List[np.ndarray], 
                        blur_results: List[Dict[str, Any]]) -> List[np.ndarray]:
        """
        Add visual indicators to blurry compartment images.
        
        Args:
            compartment_images: List of compartment images
            blur_results: List of blur analysis results
            
        Returns:
            List of compartment images with blur indicators added
        """
        if not self.config['flag_blurry_images']:
            return compartment_images
        
        result_images = []
        
        for i, image in enumerate(compartment_images):
            # Find the blur result for this image
            result = next((r for r in blur_results if r['index'] == i), None)
            
            if result and result.get('is_blurry', False):
                # Create a copy for modification
                marked_image = image.copy()
                
                # Get image dimensions
                h, w = marked_image.shape[:2]
                
                # Add a red border
                border_thickness = max(3, min(h, w) // 50)  # Scale with image size
                cv2.rectangle(
                    marked_image, 
                    (0, 0), 
                    (w - 1, h - 1), 
                    (0, 0, 255),  # Red in BGR
                    border_thickness
                )
                
                # Add "BLURRY" text
                font_scale = max(0.5, min(h, w) / 500)  # Scale with image size
                text_size, _ = cv2.getTextSize(
                    "BLURRY", 
                    cv2.FONT_HERSHEY_SIMPLEX, 
                    font_scale, 
                    2
                )
                
                # Position text in top-right corner
                text_x = w - text_size[0] - 10
                text_y = text_size[1] + 10
                
                # Add background rectangle for better visibility
                cv2.rectangle(
                    marked_image,
                    (text_x - 5, text_y - text_size[1] - 5),
                    (text_x + text_size[0] + 5, text_y + 5),
                    (0, 0, 0),
                    -1
                )
                
                # Add text
                cv2.putText(
                    marked_image,
                    "BLURRY",
                    (text_x, text_y),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    font_scale,
                    (0, 0, 255),  # Red in BGR
                    2
                )
                
                # Add variance value in smaller text
                variance_text = f"Var: {result.get('variance', 0):.1f}"
                cv2.putText(
                    marked_image,
                    variance_text,
                    (text_x, text_y + 20),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    font_scale * 0.7,
                    (0, 200, 255),  # Orange in BGR
                    1
                )
                
                result_images.append(marked_image)
            else:
                # Keep original image
                result_images.append(image)
        
        return result_images


    def improve_aruco_detection(self, image: np.ndarray) -> Tuple[Dict[int, np.ndarray], np.ndarray]:
        """
        Attempt to improve ArUco marker detection using various image preprocessing techniques.
        
        Args:
            image: Input image as numpy array
                
        Returns:
            Tuple of (dict mapping marker IDs to corner coordinates, visualization image)
        """
        # Try original detection first
        markers, viz_image = self.detect_aruco_markers(image)
        initial_marker_count = len(markers)
        
        # Log initial detection
        logger.info(f"Initial detection found {initial_marker_count} ArUco markers")
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((f"Initial detection found {initial_marker_count} ArUco markers", None))
        
        # Check if we need to improve detection
        expected_markers = len(self.config['corner_marker_ids']) + len(self.config['compartment_marker_ids'])
        if initial_marker_count >= expected_markers:
            # All markers detected, no need for improvement
            return markers, viz_image
        
        # Store the best result
        best_markers = markers
        best_viz = viz_image
        best_count = initial_marker_count
        
        # Initialize a combined marker dictionary
        combined_markers = markers.copy()
        
        # Try different preprocessing methods
        preprocessing_methods = [
            ("Adaptive thresholding", self._preprocess_adaptive_threshold),
            ("Histogram equalization", self._preprocess_histogram_equalization),
            ("Contrast enhancement", self._preprocess_contrast_enhancement),
            ("Edge enhancement", self._preprocess_edge_enhancement)
        ]
        
        for method_name, preprocess_func in preprocessing_methods:
            # Update status
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((f"Applying {method_name} to improve ArUco detection...", None))
            
            # Apply preprocessing
            processed_image = preprocess_func(image)
            
            # Detect markers on processed image
            new_markers, new_viz = self.detect_aruco_markers(processed_image)
            
            # Log results
            logger.info(f"{method_name}: detected {len(new_markers)} ArUco markers")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((f"{method_name}: detected {len(new_markers)} ArUco markers", None))
            
            # Check if this method found any new markers
            new_marker_ids = set(new_markers.keys()) - set(combined_markers.keys())
            if new_marker_ids:
                logger.info(f"{method_name} found {len(new_marker_ids)} new markers: {sorted(new_marker_ids)}")
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((f"{method_name} found {len(new_marker_ids)} new markers: {sorted(new_marker_ids)}", None))
                
                # Add new markers to combined set
                for marker_id in new_marker_ids:
                    combined_markers[marker_id] = new_markers[marker_id]
            
            # Update best result if this method found more markers
            if len(new_markers) > best_count:
                best_markers = new_markers
                best_viz = new_viz
                best_count = len(new_markers)
        
        # Log final results
        logger.info(f"Best single method found {best_count} markers")
        logger.info(f"Combined methods found {len(combined_markers)} markers")
        
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((f"Best single method found {best_count} markers", None))
            self.progress_queue.put((f"Combined methods found {len(combined_markers)} markers", None))
        
        # Use combined markers if better than any single method
        if len(combined_markers) > best_count:
            # We need to create a visualization for the combined markers
            combined_viz = image.copy()
            
            # Draw all the combined markers
            for marker_id, corners in combined_markers.items():
                # Draw marker outline
                cv2.polylines(combined_viz, [corners.astype(np.int32)], True, (0, 255, 0), 2)
                
                # Add marker ID label
                center_x = int(np.mean(corners[:, 0]))
                center_y = int(np.mean(corners[:, 1]))
                cv2.putText(combined_viz, f"ID:{marker_id}", (center_x, center_y), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            return combined_markers, combined_viz
        else:
            return best_markers, best_viz

    def _preprocess_adaptive_threshold(self, image: np.ndarray) -> np.ndarray:
        """Apply adaptive thresholding to improve marker detection."""
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        return cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
        )

    def _preprocess_histogram_equalization(self, image: np.ndarray) -> np.ndarray:
        """Apply histogram equalization to improve contrast."""
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        return cv2.equalizeHist(gray)

    def _preprocess_contrast_enhancement(self, image: np.ndarray) -> np.ndarray:
        """Enhance image contrast."""
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Apply CLAHE (Contrast Limited Adaptive Histogram Equalization)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        return clahe.apply(gray)

    def _preprocess_edge_enhancement(self, image: np.ndarray) -> np.ndarray:
        """Enhance edges to improve marker detection."""
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Apply Gaussian blur to reduce noise
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        
        # Apply Canny edge detection
        edges = cv2.Canny(blurred, 50, 150)
        
        # Dilate edges to connect potential markers
        kernel = np.ones((3, 3), np.uint8)
        dilated = cv2.dilate(edges, kernel, iterations=1)
        
        return dilated

    def detect_aruco_markers(self, image: np.ndarray) -> Tuple[Dict[int, np.ndarray], np.ndarray]:
        """
        Detect ArUco markers in the image with improved feedback.
        
        Args:
            image: Input image as numpy array
                
        Returns:
            Tuple of (dict mapping marker IDs to corner coordinates, visualization image)
        """
        # Create a copy for visualization
        viz_image = image.copy()
        
        # Convert to grayscale for marker detection
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Detect markers
        corners, ids, rejected = self.aruco_detector.detectMarkers(gray)
        
        # Create a dictionary to store detected markers
        markers = {}
        
        if ids is not None and len(ids) > 0:
            # Draw detected markers on visualization image
            cv2.aruco.drawDetectedMarkers(viz_image, corners, ids)
            
            # Store markers in dictionary with ID as key
            for i, marker_id in enumerate(ids.flatten()):
                markers[marker_id] = corners[i][0]  # Each corner is a 4x2 array
                
                # Add marker ID label
                center_x = int(np.mean(corners[i][0][:, 0]))
                center_y = int(np.mean(corners[i][0][:, 1]))
                cv2.putText(viz_image, f"ID:{marker_id}", (center_x, center_y), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            # Log detected markers and missing markers
            expected_markers = set(self.config['corner_marker_ids'] + self.config['compartment_marker_ids'])
            detected_markers = set(markers.keys())
            missing_markers = expected_markers - detected_markers
            
            logger.info(f"Detected {len(ids)} ArUco markers out of {len(expected_markers)} expected")
            logger.info(f"Missing markers: {sorted(missing_markers) if missing_markers else 'None'}")
            
            # Update status in GUI if available
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((f"Detected {len(ids)}/{len(expected_markers)} ArUco markers", None))
                if missing_markers:
                    self.progress_queue.put((f"Missing markers: {sorted(missing_markers)}", None))
        else:
            logger.warning("No ArUco markers detected in the image")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put(("No ArUco markers detected in the image", None))
        
        return markers, viz_image
    
    def extract_compartment_boundaries(self, 
                                    image: np.ndarray, 
                                    markers: Dict[int, np.ndarray]) -> Tuple[Optional[List[Tuple[int, int, int, int]]], np.ndarray]:
        """
        Extract compartment boundaries using detected ArUco markers.
        Each marker defines a single compartment directly beneath it.
        Improved to handle missing markers and maintain proper compartment order.
        
        Args:
            image: Input image
            markers: Dictionary mapping marker IDs to corner coordinates
            
        Returns:
            Tuple of (list of compartment boundaries as (x1, y1, x2, y2), visualization image)
        """
        # Create a copy for visualization
        viz_image = image.copy()
        
        # Log marker detection summary
        logger.info(f"Working with {len(markers)} markers: {sorted(markers.keys())}")
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((f"Working with {len(markers)} markers: {sorted(markers.keys())}", None))
        
        # Check if we have enough markers
        corner_markers = [markers.get(id) for id in self.config['corner_marker_ids']]
        corner_markers = [m for m in corner_markers if m is not None]
        
        if len(corner_markers) < 2:
            logger.error("Not enough corner markers detected to establish tray boundaries")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put(("Not enough corner markers detected", None))
            return None, viz_image
        
        # Find top and bottom boundaries using corner markers (same as before)
        corner_marker_ids = [id for id in self.config['corner_marker_ids'] if id in markers]
        
        # Calculate top boundary (y-min of top markers)
        top_marker_ids = [0, 1]  # IDs of markers that should be at top
        top_markers = [markers[id] for id in corner_marker_ids if id in top_marker_ids]
        
        if not top_markers:
            logger.warning("No top markers found, using top-most detected marker")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put(("No top markers found, using top-most detected marker", None))
            # Use the top-most marker(s) instead
            marker_y_means = {id: np.mean(markers[id][:, 1]) for id in corner_marker_ids}
            sorted_by_y = sorted(marker_y_means.items(), key=lambda x: x[1])
            top_markers = [markers[sorted_by_y[0][0]]]
        
        # Find lowest y-coordinate of top markers (highest in image)
        top_y = min([np.min(marker[:, 1]) for marker in top_markers])
        
        # Calculate bottom boundary (y-max of bottom markers)
        bottom_marker_ids = [2, 3]  # IDs of markers that should be at bottom
        bottom_markers = [markers[id] for id in corner_marker_ids if id in bottom_marker_ids]
        
        if not bottom_markers:
            logger.warning("No bottom markers found, using bottom-most detected marker")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put(("No bottom markers found, using bottom-most detected marker", None))
            # Use the bottom-most marker(s) instead
            marker_y_means = {id: np.mean(markers[id][:, 1]) for id in corner_marker_ids}
            sorted_by_y = sorted(marker_y_means.items(), key=lambda x: x[1], reverse=True)
            bottom_markers = [markers[sorted_by_y[0][0]]]
        
        # Find highest y-coordinate of bottom markers (lowest in image)
        bottom_y = max([np.max(marker[:, 1]) for marker in bottom_markers])
        
        # Draw horizontal boundary lines
        cv2.line(viz_image, (0, int(top_y)), (image.shape[1], int(top_y)), (0, 255, 0), 2)
        cv2.line(viz_image, (0, int(bottom_y)), (image.shape[1], int(bottom_y)), (0, 255, 0), 2)
        
        # ---- New improved compartment marker handling ----
        
        # Get expected compartment count from config
        expected_count = self.config['compartment_count']
        
        # Define expected compartment marker IDs
        expected_compartment_ids = [
            id for id in self.config['compartment_marker_ids'] 
            if id not in self.config['corner_marker_ids']
        ][:self.config['compartment_count']]
        
        # Identify which compartment markers are detected
        detected_compartment_ids = [id for id in expected_compartment_ids if id in markers]
        
        # Report detected compartments
        detected_msg = f"Detected {len(detected_compartment_ids)}/{len(expected_compartment_ids)} compartment markers"
        logger.info(detected_msg)
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((detected_msg, None))
        
        # Check if we're missing any markers
        missing_ids = [id for id in expected_compartment_ids if id not in markers]
        if missing_ids:
            missing_msg = f"Missing markers: {missing_ids}"
            logger.info(missing_msg)
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((missing_msg, None))
        
        # Calculate marker information for all detected compartment markers
        marker_info = {}
        for marker_id in detected_compartment_ids:
            marker_corners = markers[marker_id]
            
            # Calculate marker center
            center_x = int(np.mean(marker_corners[:, 0]))
            center_y = int(np.mean(marker_corners[:, 1]))
            
            # Calculate marker width - direct from marker
            left_x = int(np.min(marker_corners[:, 0]))
            right_x = int(np.max(marker_corners[:, 0]))
            width = right_x - left_x
            
            marker_info[marker_id] = {
                'center_x': center_x,
                'center_y': center_y,
                'left_x': left_x,
                'right_x': right_x,
                'width': width,
                'is_estimated': False
            }
            
            # Draw marker centers on visualization
            cv2.circle(viz_image, (center_x, center_y), 5, (255, 0, 0), -1)
            cv2.putText(viz_image, f"{marker_id}", (center_x - 10, center_y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 255), 2)
        
        # Estimate positions of missing markers by interpolation/extrapolation
        if missing_ids:
            logger.info(f"Estimating positions for {len(missing_ids)} missing markers")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((f"Estimating positions for {len(missing_ids)} missing markers", None))
            
            # Sort compartment IDs in order (assuming they correspond to compartment order from left to right)
            sorted_detected_ids = sorted(detected_compartment_ids)
            
            # If we have at least two detected markers, we can interpolate/extrapolate
            if len(sorted_detected_ids) >= 2:
                # Calculate the average width and spacing
                widths = [marker_info[id]['width'] for id in sorted_detected_ids]
                avg_width = sum(widths) / len(widths)
                
                # Calculate average spacing between consecutive markers
                spacings = []
                center_xs = [marker_info[id]['center_x'] for id in sorted_detected_ids]
                
                for i in range(len(sorted_detected_ids) - 1):
                    current_id = sorted_detected_ids[i]
                    next_id = sorted_detected_ids[i + 1]
                    
                    if next_id - current_id == 1:  # Only consider adjacent markers
                        current_center = marker_info[current_id]['center_x']
                        next_center = marker_info[next_id]['center_x']
                        spacing = next_center - current_center
                        spacings.append(spacing)
                
                avg_spacing = sum(spacings) / len(spacings) if spacings else avg_width * 1.5
                logger.info(f"Average marker width: {avg_width:.2f} pixels")
                logger.info(f"Average spacing between consecutive markers: {avg_spacing:.2f} pixels")
                
                # Estimate center_x for each missing marker
                for missing_id in missing_ids:
                    # Find nearest detected markers on both sides
                    left_markers = [id for id in sorted_detected_ids if id < missing_id]
                    right_markers = [id for id in sorted_detected_ids if id > missing_id]
                    
                    if left_markers and right_markers:
                        # Interpolate position
                        left_id = max(left_markers)
                        right_id = min(right_markers)
                        
                        left_center_x = marker_info[left_id]['center_x']
                        right_center_x = marker_info[right_id]['center_x']
                        
                        # Linear interpolation based on marker ID
                        ratio = (missing_id - left_id) / (right_id - left_id)
                        estimated_center_x = left_center_x + ratio * (right_center_x - left_center_x)
                        
                        estimated_center_y = marker_info[left_id]['center_y']  # Use same Y as neighbors
                        
                    elif left_markers:
                        # Extrapolate to the right
                        left_id = max(left_markers)
                        estimated_center_x = marker_info[left_id]['center_x'] + (missing_id - left_id) * avg_spacing
                        estimated_center_y = marker_info[left_id]['center_y']
                        
                    elif right_markers:
                        # Extrapolate to the left
                        right_id = min(right_markers)
                        estimated_center_x = marker_info[right_id]['center_x'] - (right_id - missing_id) * avg_spacing
                        estimated_center_y = marker_info[right_id]['center_y']
                        
                    else:
                        # This case shouldn't be reached, but just in case
                        logger.warning(f"Cannot estimate position for marker {missing_id}, no reference markers")
                        continue
                    
                    # Create estimated marker info
                    marker_info[missing_id] = {
                        'center_x': int(estimated_center_x),
                        'center_y': int(estimated_center_y),
                        'left_x': int(estimated_center_x - avg_width/2),
                        'right_x': int(estimated_center_x + avg_width/2),
                        'width': avg_width,
                        'is_estimated': True
                    }
                    
                    # Draw estimated marker position with different color
                    cv2.circle(viz_image, 
                            (int(estimated_center_x), int(estimated_center_y)), 
                            5, (0, 165, 255), -1)
                    cv2.putText(viz_image, 
                            f"E{missing_id}", 
                            (int(estimated_center_x - 10), int(estimated_center_y - 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 165, 255), 2)
        
        # Create compartment boundaries from ALL marker positions (actual + estimated)
        compartment_boundaries = []
        all_marker_ids = sorted(marker_info.keys())
        
        # If we have more detected markers than expected, only use the first expected_count
        all_marker_ids = all_marker_ids[:expected_count] if len(all_marker_ids) > expected_count else all_marker_ids
        
        # Create boundaries in consistent order (left to right)
        for i, marker_id in enumerate(all_marker_ids):
            info = marker_info[marker_id]
            
            # Use the marker's dimensions for the compartment
            x1 = info['left_x']
            x2 = info['right_x']
            y1 = int(top_y)
            y2 = int(bottom_y)
            
            # Add the compartment boundary with its original index to maintain order
            compartment_boundaries.append((x1, y1, x2, y2, marker_id))
            
            # Draw compartment boundary
            color = (0, 165, 255) if info.get('is_estimated', False) else (0, 255, 0)
            cv2.rectangle(viz_image, (x1, y1), (x2, y2), color, 2)
            
            # Draw marker ID and compartment number
            compartment_number = i + 1  # 1-based indexing for display
            cv2.putText(viz_image, f"C{compartment_number}", ((x1 + x2) // 2, (y1 + y2) // 2),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
        
        # Check if we found all expected compartments
        if len(compartment_boundaries) < expected_count:
            logger.warning(f"Only found {len(compartment_boundaries)}/{expected_count} compartments")
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((f"Only found {len(compartment_boundaries)}/{expected_count} compartments", "warning"))
        
        # Log summary of compartment detection
        logger.info(f"Successfully identified {len(compartment_boundaries)} compartment boundaries")
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((f"Successfully identified {len(compartment_boundaries)} compartment boundaries", None))
                
        # Sort compartments by x-position (left to right) and remove the marker_id from the final boundaries
        sorted_boundaries = [(x1, y1, x2, y2) for x1, y1, x2, y2, _ in 
                            sorted(compartment_boundaries, key=lambda x: x[0])]
        
        return sorted_boundaries, viz_image

    def correct_image_skew(self, image: np.ndarray, markers: Dict[int, np.ndarray], return_transform: bool = True) -> Union[np.ndarray, Tuple[np.ndarray, np.ndarray, float]]:
        """
        Correct image skew using corner ArUco markers.
        
        Args:
            image: Input image as numpy array
            markers: Dictionary mapping marker IDs to corner coordinates
            return_transform: Whether to return the rotation matrix and angle
            
        Returns:
            Deskewed image, and optionally rotation matrix and angle
        """
        try:
            # Get available corner markers
            corner_ids = [id for id in self.config['corner_marker_ids'] if id in markers]
            
            if len(corner_ids) < 2:
                logger.warning("Not enough corner markers for skew correction")
                if return_transform:
                    return image, None, 0.0
                return image
            
            # Get marker centers
            corner_centers = []
            for id in corner_ids:
                corners = markers[id]
                center_x = int(np.mean(corners[:, 0]))
                center_y = int(np.mean(corners[:, 1]))
                corner_centers.append((id, center_x, center_y))
            
            # Check if we have top markers for angle calculation
            top_markers = [m for m in corner_centers if m[0] in [0, 1]]
            if len(top_markers) >= 2:
                # Sort by x-coordinate
                top_markers.sort(key=lambda m: m[1])
                
                # Calculate angle
                _, x1, y1 = top_markers[0]
                _, x2, y2 = top_markers[1]
                
                angle_rad = np.arctan2(y2 - y1, x2 - x1)
                angle_deg = np.degrees(angle_rad)
                
                if abs(angle_deg) > 0.5:  # Only correct if angle is significant
                    logger.info(f"Correcting image skew of {angle_deg:.2f} degrees")
                    
                    # Get image center for rotation
                    h, w = image.shape[:2]
                    center = (w // 2, h // 2)
                    
                    # Get rotation matrix
                    rotation_matrix = cv2.getRotationMatrix2D(center, angle_deg, 1.0)
                    
                    # Apply rotation
                    corrected_image = cv2.warpAffine(image, rotation_matrix, (w, h), 
                                                flags=cv2.INTER_LINEAR, 
                                                borderMode=cv2.BORDER_CONSTANT,
                                                borderValue=(255, 255, 255))
                    
                    if return_transform:
                        return corrected_image, rotation_matrix, angle_deg
                    return corrected_image
            
            # If we can't correct or angle is minimal, return original
            if return_transform:
                return image, None, 0.0
            return image
        
        except Exception as e:
            logger.warning(f"Skew correction failed: {str(e)}")
            if return_transform:
                return image, None, 0.0
            return image

    def move_processed_image(self, source_path: str, metadata: Dict[str, Any]) -> Optional[str]:
        """
        Move a successfully processed image to the appropriate folder.
        
        Args:
            source_path: Path to the original image file
            metadata: Metadata with hole_id, depth_from, depth_to
            
        Returns:
            Optional[str]: Path to the new image location, or None if move fails
        """
        try:
            # Validate required metadata
            hole_id = metadata.get('hole_id', 'Unknown')
            depth_from = metadata.get('depth_from', 0)
            depth_to = metadata.get('depth_to', 0)
            
            # Move the file using the FileManager
            return self.file_manager.move_original_file(
                source_path,
                hole_id,
                depth_from,
                depth_to,
                is_processed=True
            )
            
        except Exception as e:
            logger.error(f"Error moving processed image {source_path}: {str(e)}")
            return None

    def process_image(self, image_path: str) -> bool:
        """
        Process a single chip tray image with enhanced feedback and improved detection.
        Uses a two-phase approach: first detects markers and metadata on a downsized image
        for speed, then applies the results to the full-resolution image for quality output.
        
        Args:
            image_path (str): Path to the image file to be processed
        
        Returns:
            bool: True if processing succeeded, False otherwise
        """
        # and reference to current instance for cache access
        self.duplicate_handler = DuplicateHandler(self.file_manager.processed_dir)
        self.duplicate_handler.parent = self  # Add reference to self for access to visualization cache

        try:
            # Try to read the image using PIL first
            try:
                from PIL import Image as PILImage
                import numpy as np
                
                # Open with PIL
                pil_img = PILImage.open(image_path)
                
                # Convert to numpy array for OpenCV processing
                original_image = np.array(pil_img)
                
                # Convert RGB to BGR for OpenCV compatibility if needed
                if len(original_image.shape) == 3 and original_image.shape[2] == 3:
                    original_image = cv2.cvtColor(original_image, cv2.COLOR_RGB2BGR)
                    
                logger.info(f"Successfully read image with PIL: {image_path}")
                
            except Exception as e:
                logger.warning(f"Failed to read image with PIL, trying OpenCV: {str(e)}")
                
                # Fallback to OpenCV if PIL fails
                original_image = cv2.imread(image_path)
                if original_image is None:
                    error_msg = f"Failed to read image with both PIL and OpenCV: {image_path}"
                    logger.error(error_msg)
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put((error_msg, None))
                    return False
            
            # Update progress
            base_name = os.path.basename(image_path)
            status_msg = f"Processing Image: {base_name}"
            logger.info(status_msg)
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((status_msg, None))
            
            # Calculate scale based on original size - aim for ~1-2MB file size
            h, w = original_image.shape[:2]
            original_pixels = h * w
            target_pixels = 2000000  # Target ~2 million pixels (e.g., 1414x1414)
            
            if original_pixels > target_pixels:
                scale = (target_pixels / original_pixels) ** 0.5
                new_width = int(w * scale)
                new_height = int(h * scale)
                
                # Create downsampled image
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((f"Downsizing image from {w}x{h} to {new_width}x{new_height} for processing", None))
                    
                small_image = cv2.resize(original_image, (new_width, new_height), interpolation=cv2.INTER_AREA)
            else:
                # Image is already small enough
                small_image = original_image.copy()
                logger.info(f"Image already small ({w}x{h}), using as is for processing")
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((f"Image already small ({w}x{h}), using as is for processing", None))
            
            # Store the small image as an instance variable so it's accessible to the metadata dialog
            self.small_image = small_image.copy()
            
            # Store visualization steps (all on the small image)
            viz_steps = []
            viz_steps.append(("Original Image (Small)", small_image.copy()))
            
            # in Class ChipTrayExtractor - Def process_image
            # Check if this file has been previously processed
            previously_processed = self.file_manager.check_original_file_processed(image_path)

            # Use existing metadata (e.g., from filename) or initialize empty
            metadata: Dict[str, Any] = getattr(self, 'metadata', {})


            # Extract metadata with OCR if enabled - using SMALL image
            if self.config['enable_ocr'] and self.tesseract_manager.is_available:
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put(("Checking for metadata...", None))
                
                # Initialize empty markers dictionary if we're going to skip OCR
                markers = {}
                
                try:
                    # Make sure TesseractManager has access to FileManager
                    self.tesseract_manager.file_manager = self.file_manager
                    
                    # Check if we already have metadata from the filename
                    if previously_processed:
                        # Use the metadata from the filename and skip OCR extraction
                        metadata = previously_processed
                        logger.info(f"Using metadata from filename: {metadata}")
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put((f"Using metadata from filename: Hole ID={metadata.get('hole_id')}, Depth={metadata.get('depth_from')}-{metadata.get('depth_to')}m", None))
                        
                        # Detect markers in the small image for subsequent steps
                        markers, markers_viz = self.improve_aruco_detection(small_image)
                        viz_steps.append(("ArUco Markers", markers_viz))
                        
                        # Create minimal OCR metadata for dialog display
                        ocr_metadata = {
                            'hole_id': metadata.get('hole_id'),
                            'depth_from': metadata.get('depth_from'),
                            'depth_to': metadata.get('depth_to'),
                            'confidence': metadata.get('confidence', 100.0),
                            'from_filename': True,
                            'metadata_region': None
                        }
                        
                        # Create a visualization image for the dialog
                        vis_image = small_image.copy()
                        h, w = vis_image.shape[:2]
                        
                        # Add text overlay
                        overlay = np.zeros((60, w, 3), dtype=np.uint8)
                        cv2.putText(overlay, f"Metadata from filename: {metadata.get('hole_id')} {metadata.get('depth_from')}-{metadata.get('depth_to')}m",
                                (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
                        
                        # Add overlay to top of image
                        if len(vis_image.shape) == 2:  # Convert grayscale to color
                            vis_image = cv2.cvtColor(vis_image, cv2.COLOR_GRAY2BGR)
                            
                        # Create combined image
                        combined = np.vstack([overlay, vis_image])
                        ocr_metadata['metadata_region_viz'] = combined
                        
                        # Still show the dialog for confirmation if enabled
                        needs_confirmation = self.config['prompt_for_metadata']
                        
                    else:
                        # No existing metadata, perform OCR
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put(("Extracting metadata with OCR...", None))
                        
                        # First try the composite method
                        ocr_metadata = self.tesseract_manager.extract_metadata_with_composite(
                            small_image, markers, original_filename=image_path, progress_queue=self.progress_queue
                        )
                        
                        # Check if we got good results from the composite method
                        composite_confidence = ocr_metadata.get('confidence', 0)
                        composite_has_data = (ocr_metadata.get('hole_id') is not None and 
                                            ocr_metadata.get('depth_from') is not None and 
                                            ocr_metadata.get('depth_to') is not None)
                        
                        # If the composite method didn't yield good results, just log it
                        if not composite_has_data or composite_confidence < self.config['ocr_confidence_threshold']:
                            logger.info(f"OCR confidence too low ({composite_confidence:.1f}%), will prompt user for metadata")
                            if hasattr(self, 'progress_queue'):
                                self.progress_queue.put((f"OCR confidence too low, user input required", None))

                        # Log OCR results of the method that was chosen
                        ocr_log_msg = f"OCR Results: Confidence={ocr_metadata.get('confidence', 0):.1f}%"
                        if ocr_metadata.get('hole_id'):
                            ocr_log_msg += f", Hole ID={ocr_metadata['hole_id']}"
                        if ocr_metadata.get('depth_from') is not None and ocr_metadata.get('depth_to') is not None:
                            ocr_log_msg += f", Depth={ocr_metadata['depth_from']}-{ocr_metadata['depth_to']}"
                        
                        logger.info(ocr_log_msg)
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put((ocr_log_msg, None))
                        
                        # Add visualization if available
                        if 'metadata_region_viz' in ocr_metadata:
                            viz_steps.append(("OCR Input", ocr_metadata['metadata_region_viz']))
                        
                        # Always show confirmation dialog if enabled
                        needs_confirmation = self.config['prompt_for_metadata']
                        
                        ocr_confidence_msg = f"OCR confidence: {ocr_metadata.get('confidence', 0):.1f}%"
                        logger.info(ocr_confidence_msg)
                        logger.info(f"Will prompt for metadata confirmation: {needs_confirmation}")

                    # Whether metadata is from OCR or filename, show the dialog if configured to do so
                    if needs_confirmation and self.root is not None:
                        # Handle user confirmation through dialog
                        if hasattr(self, 'progress_queue'):
                            dialog_msg = "Prompting for metadata confirmation..."
                            self.progress_queue.put((dialog_msg, None))
                        
                        # Create a temporary queue for receiving metadata
                        metadata_queue = queue.Queue()
                        
                        # Schedule dialog creation on main thread
                        self.root.after(0, self._handle_metadata_dialog_on_main_thread, 
                                    ocr_metadata, metadata_queue)
                        
                        # Wait for dialog result
                        try:
                            result = metadata_queue.get()  # Remove the timeout
                            if result:
                                metadata = result
                                logger.info(f"User confirmed metadata: {metadata}")
                            else:
                                # User canceled the dialog and confirmed cancellation
                                logger.warning("Image processing canceled by user")
                                if hasattr(self, 'progress_queue'):
                                    self.progress_queue.put(("Processing canceled by user", None))
                                return False  # Return early to skip further processing
                        except Exception as e:
                            logger.error(f"Metadata dialog error: {str(e)}")
                            logger.error(traceback.format_exc())
                            if hasattr(self, 'progress_queue'):
                                self.progress_queue.put((f"Metadata dialog error: {str(e)}", None))
                            return False
                    elif not self.config['prompt_for_metadata']:
                        # If prompt_for_metadata is disabled, use the metadata directly
                        if previously_processed:
                            logger.info(f"Using metadata from filename without confirmation (prompt disabled): {metadata}")
                        else:
                            # Use OCR metadata without prompting
                            metadata = {
                                'hole_id': ocr_metadata.get('hole_id'),
                                'depth_from': ocr_metadata.get('depth_from'),
                                'depth_to': ocr_metadata.get('depth_to')
                            }
                            logger.info(f"Using OCR metadata without confirmation (prompt disabled): {metadata}")
                    
                    # Check for potential duplicates AFTER OCR or manual metadata entry
                    if (metadata.get('hole_id') and 
                        metadata.get('depth_from') is not None and 
                        metadata.get('depth_to') is not None):
                        
                        # Loop for handling metadata modification
                        while True:
                            try:
                                # Check for duplicates
                                duplicate_result = self.duplicate_handler.check_duplicate(
                                    metadata['hole_id'], 
                                    metadata['depth_from'], 
                                    metadata['depth_to'], 
                                    small_image,  # Use the downsampled image
                                    image_path
                                )
                                
                                # Process the result based on its type
                                if isinstance(duplicate_result, dict):
                                    # User chose to modify metadata - update metadata and check again
                                    logger.info(f"User modified metadata from {metadata} to {duplicate_result}")
                                    metadata = duplicate_result
                                    # Continue the loop to check for duplicates again with new metadata
                                    continue
                                    
                                elif duplicate_result == False:
                                    # User chose to skip or keep the existing image
                                    logger.info(f"Skipping duplicate image: {os.path.basename(image_path)}")
                                    
                                    # Move to "Failed and Skipped Originals" folder
                                    try:
                                        self.file_manager.move_original_file(
                                            image_path,
                                            metadata['hole_id'],
                                            metadata['depth_from'],
                                            metadata['depth_to'],
                                            is_processed=False  # Mark as skipped
                                        )
                                    except Exception as move_error:
                                        logger.error(f"Error moving skipped file: {str(move_error)}")
                                    
                                    return False
                                    
                                else:
                                    # User chose to continue processing (replace existing)
                                    logger.info("Continuing with processing (replacing existing image)...")
                                    break
                                    
                            except Exception as e:
                                logger.error(f"Error checking for duplicates: {str(e)}")
                                logger.error(traceback.format_exc())
                                # Continue with processing despite duplicate check error
                                break

                    # Save small image for debugging if we have metadata
                    if metadata and metadata.get('hole_id') and self.config['save_debug_images']:
                        # Use FileManager to save the small debug image
                        self.file_manager.save_debug_image(
                            small_image,
                            metadata['hole_id'],
                            metadata.get('depth_from', 0),
                            metadata.get('depth_to', 0),
                            "small_image"
                        )
                        
                        # Save metadata region if available
                        if 'metadata_region_viz' in ocr_metadata:
                            self.file_manager.save_debug_image(
                                ocr_metadata['metadata_region_viz'],
                                metadata['hole_id'],
                                metadata.get('depth_from', 0),
                                metadata.get('depth_to', 0),
                                "ocr_input"
                            )
                    
                    # Update debug file names if we have metadata
                    if metadata and metadata.get('hole_id'):
                        try:
                            # Use the FileManager's method to rename debug files
                            self.file_manager.rename_debug_files(
                                image_path, 
                                metadata.get('hole_id'), 
                                metadata.get('depth_from'), 
                                metadata.get('depth_to')
                            )
                        except Exception as e:
                            logger.warning(f"Could not rename debug files: {str(e)}")
                            logger.debug(traceback.format_exc())
                except Exception as e:
                    logger.error(f"Error during OCR processing: {str(e)}")
                    logger.error(traceback.format_exc())

                # Correct image skew on the SMALL image if possible
                rotation_matrix = None  # Store rotation matrix for reuse
                rotation_angle = 0.0    # Store rotation angle

                try:
                    if markers:
                        # Capture all return values from correct_image_skew
                        result = self.correct_image_skew(small_image, markers)
                        
                        # Unpack the return values correctly
                        if isinstance(result, tuple):
                            corrected_small_image, rotation_matrix, rotation_angle = result
                        else:
                            corrected_small_image = result
                        
                        if corrected_small_image is not small_image:  # If correction was applied
                            # Re-detect markers on corrected image
                            markers, markers_viz = self.improve_aruco_detection(corrected_small_image)
                            small_image = corrected_small_image  # Update small image for further processing
                            viz_steps.append(("Corrected Image", markers_viz))
                            
                            logger.info(f"Re-detected {len(markers)} markers after skew correction")
                            if hasattr(self, 'progress_queue'):
                                self.progress_queue.put((f"Re-detected {len(markers)} markers after skew correction", None))
                except Exception as e:
                    logger.warning(f"Skew correction failed: {str(e)}")
                    logger.error(traceback.format_exc())  # Add this to see the full error stack trace
                # Report marker detection status
                expected_markers = set(self.config['corner_marker_ids'] + self.config['compartment_marker_ids'])
                detected_markers = set(markers.keys())
                missing_markers = expected_markers - detected_markers
                
                status_msg = f"Detected {len(detected_markers)}/{len(expected_markers)} ArUco markers"
                logger.info(status_msg)
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((status_msg, None))
                    
                    if missing_markers:
                        self.progress_queue.put((f"Missing markers: {sorted(missing_markers)}", None))
                    
            
                    # Apply the same skew correction to the original image if a rotation was applied
                    corrected_original_image = original_image
                    if rotation_matrix is not None and rotation_angle != 0.0:
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put((f"Applying skew correction of {rotation_angle:.2f} degrees to high-resolution image...", None))
                        
                        # Get image dimensions
                        h, w = original_image.shape[:2]
                        center = (w // 2, h // 2)
                        
                        # Apply same rotation to original image
                        corrected_original_image = cv2.warpAffine(
                            original_image, 
                            rotation_matrix, 
                            (w, h),
                            flags=cv2.INTER_LINEAR, 
                            borderMode=cv2.BORDER_CONSTANT,
                            borderValue=(255, 255, 255)
                        )
                        
                        # Use the corrected image for further processing
                        original_image = corrected_original_image
                        logger.info(f"Applied skew correction to high-resolution image")

                    # Extract compartment boundaries on the SMALL image
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put(("Extracting compartment boundaries...", None))
                    
                    try:
                        boundaries_result = self.extract_compartment_boundaries(small_image, markers)
                        if boundaries_result is None:
                            error_msg = "Failed to extract compartment boundaries"
                            logger.error(error_msg)
                            if hasattr(self, 'progress_queue'):
                                self.progress_queue.put((error_msg, None))
                            return False
                            
                        compartment_boundaries_small, boundaries_viz = boundaries_result
                    except Exception as e:
                        error_msg = f"Error in compartment boundary extraction: {str(e)}"
                        logger.error(error_msg)
                        logger.error(traceback.format_exc())
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put((error_msg, None))
                        return False
                    
                    viz_steps.append(("Compartment Boundaries", boundaries_viz))
                    
                    # Report number of compartments found
                    status_msg = f"Found {len(compartment_boundaries_small)}/{self.config['compartment_count']} compartments"
                    logger.info(status_msg)
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put((status_msg, None))
                    
                    # SCALED EXTRACTION: Scale up the coordinates from small image to original image
                    if small_image.shape != original_image.shape:
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put(("Scaling boundaries to original image size...", None))
                            
                        # Calculate scale factors
                        scale_x = original_image.shape[1] / small_image.shape[1]
                        scale_y = original_image.shape[0] / small_image.shape[0]
                        
                        # Scale up the coordinates
                        compartment_boundaries = []
                        for x1, y1, x2, y2 in compartment_boundaries_small:
                            # Round to integers
                            scaled_x1 = int(x1 * scale_x)
                            scaled_y1 = int(y1 * scale_y)
                            scaled_x2 = int(x2 * scale_x)
                            scaled_y2 = int(y2 * scale_y)
                            
                            compartment_boundaries.append((scaled_x1, scaled_y1, scaled_x2, scaled_y2))
                            
                        logger.info(f"Scaled {len(compartment_boundaries)} compartment boundaries to original image")
                    else:
                        # No scaling needed
                        compartment_boundaries = compartment_boundaries_small
                    
                    # Now extract compartments from the ORIGINAL high-resolution image
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put(("Extracting high-resolution compartments...", None))
                        
                    compartments, compartments_viz = self.extract_compartments(original_image, compartment_boundaries)
                    
                    # Add QAQC review process here
                    if hasattr(self, 'root') and self.root is not None:
                        # Create QAQC manager if it doesn't exist
                        if not hasattr(self, 'qaqc_manager'):
                            self.qaqc_manager = QAQCManager(self.root, self.file_manager, self)
                        
                        # Add tray for review
                        self.qaqc_manager.add_tray_for_review(
                            metadata['hole_id'],
                            metadata['depth_from'],
                            metadata['depth_to'],
                            image_path,
                            compartments
                        )
                        
                        # Skip standard saving since QAQC system will handle it
                        return True
                    
                    # Fallback to standard saving if we don't have a GUI
                    # Save the extracted compartments using the FileManager
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put(("Saving compartment images...", None))


                    # Create and save visualization image if required
                    if self.config['save_debug_images']:
                        if hasattr(self, 'progress_queue'):
                            self.progress_queue.put(("Saving debug images...", None))
                        
                        # Create the visualization image
                        viz_image = self.create_visualization_image(small_image, viz_steps)
                        
                        # Save via FileManager if we have valid metadata
                        if metadata and metadata.get('hole_id'):
                            hole_id = metadata['hole_id']
                            depth_from = metadata.get('depth_from', 0)
                            depth_to = metadata.get('depth_to', 0)
                            
                            # Save visualization image
                            self.file_manager.save_debug_image(
                                viz_image,
                                hole_id,
                                depth_from,
                                depth_to,
                                "visualization"
                            )
                            
                            logger.info(f"Saved debug images for {hole_id} {depth_from}-{depth_to}m")
                        else:
                            # Use FileManager to save in a general debug location
                            temp_debug_dir = os.path.join(self.file_manager.processed_dir, "Debug Images", "Unidentified")
                            os.makedirs(temp_debug_dir, exist_ok=True)
                            base_name = os.path.splitext(os.path.basename(image_path))[0]
                            viz_path = self.file_manager.save_temp_debug_image(viz_image, image_path, "visualization")
                            cv2.imwrite(viz_path, viz_image, [cv2.IMWRITE_JPEG_QUALITY, 90])
                            logger.info(f"Saved visualization to {viz_path} (no metadata available)")
                    
                    # Final summary
                    success_msg = f"Successfully processed {base_name}: saved {num_saved}/{len(compartments)} compartments"
                    if metadata.get('hole_id'):
                        success_msg += f" for hole {metadata['hole_id']}"
                        
                    logger.info(success_msg)
                    if hasattr(self, 'progress_queue'):
                        self.progress_queue.put((success_msg, None))
                        
                        # Add details about missing compartments if any
                        if num_saved < len(compartments):
                            warning_msg = f"Warning: Expected {len(compartments)} compartments, but saved {num_saved}"
                            logger.warning(warning_msg)
                            self.progress_queue.put((warning_msg, None))
                    
                    # Register with duplicate handler when processing is complete and successful
                    if (metadata.get('hole_id') and 
                        metadata.get('depth_from') is not None and 
                        metadata.get('depth_to') is not None):
                        # Register the processed entry with the duplicate handler
                        output_files = []
                        expected_count = self.config['compartment_count']
                        total_depth = float(metadata['depth_to']) - float(metadata['depth_from'])
                        depth_increment = total_depth / expected_count
                        
                        # Create the expected filenames
                        for i in range(expected_count):
                            comp_depth_to = float(metadata['depth_from']) + ((i + 1) * depth_increment)
                            filename = f"{metadata['hole_id']}_CC_{int(comp_depth_to)}.{self.config['output_format']}"
                            output_files.append(filename)
                        
                        self.duplicate_handler.register_processed_entry(
                            metadata['hole_id'], 
                            metadata['depth_from'], 
                            metadata['depth_to'], 
                            output_files
                        )

                    # Move the processed image to the appropriate folder
                    if (metadata and metadata.get('hole_id') and 
                        metadata.get('depth_from') is not None and 
                        metadata.get('depth_to') is not None):
                        
                        try:
                            # Use the FileManager to move the original file
                            new_path = self.file_manager.move_original_file(
                                image_path,
                                metadata['hole_id'],
                                metadata['depth_from'],
                                metadata['depth_to'],
                                is_processed=True
                            )
                            
                            logger.info(f"Moved processed image to: {new_path}")
                        except Exception as e:
                            # Log move error but don't interrupt overall processing result
                            logger.warning(f"Could not move processed image {image_path}: {str(e)}")
                    else:
                        logger.warning(f"Cannot move {image_path} to organized storage: missing metadata")
                    # TODO - check that this is correct so it isn't leaking metadata across images    
                    self.metadata = {}
                    return True  # Processing succeeded
            
        except Exception as e:
            # Handle any unexpected errors during processing
            error_msg = f"Error processing {image_path}: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            
            # Update progress queue if available
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((error_msg, None))
                self.metadata = {} # TODO - check that this is correct so it isn't leaking metadata across images 

            
            return False

    def process_folder(self, folder_path: str) -> Tuple[int, int]:
        """
        Process all images in a folder with enhanced progress reporting.
        
        Args:
            folder_path: Path to folder containing images
            
        Returns:
            Tuple of (number of images processed, number of images failed)
        """
        # Get all image files from the folder
        image_extensions = ('.jpg', '.jpeg', '.png')
        image_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path)
                    if os.path.isfile(os.path.join(folder_path, f)) and 
                    f.lower().endswith(image_extensions)]
        
        if not image_files:
            warning_msg = f"No image files found in {folder_path}"
            logger.warning(warning_msg)
            if hasattr(self, 'progress_queue'):
                self.progress_queue.put((warning_msg, None))
            return 0, 0
        
        # Log found images
        info_msg = f"Found {len(image_files)} image files"
        logger.info(info_msg)
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((info_msg, None))
        
        # Process each image
        successful = 0
        failed = 0
        
        for i, image_path in enumerate(image_files):
            try:
                # Update progress
                progress = ((i + 1) / len(image_files)) * 100
                progress_msg = f"Processing Image: {i+1}/{len(image_files)}: {os.path.basename(image_path)}"
                logger.info(progress_msg)
                
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((progress_msg, progress))
                
                # Process the image
                if self.process_image(image_path):
                    successful += 1
                else:
                    failed += 1
                    
            except Exception as e:
                error_msg = f"Error processing {image_path}: {str(e)}"
                logger.error(error_msg)
                logger.error(traceback.format_exc())
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((error_msg, None))
                failed += 1
        
        # Log summary
        summary_msg = f"Processing complete: {successful} successful, {failed} failed"
        logger.info(summary_msg)
        if hasattr(self, 'progress_queue'):
            self.progress_queue.put((summary_msg, 100))  # Set progress to 100%
        
        return successful, failed

    def extract_compartments(self, 
                            image: np.ndarray, 
                            compartment_boundaries: List[Tuple[int, int, int, int]]) -> Tuple[List[np.ndarray], np.ndarray]:
        # Create a copy for visualization
        viz_image = image.copy()
        
        # Extract each compartment
        compartments = []
        
        # Find the maximum width and height among all compartments
        max_width = 0
        max_height = 0
        extracted_compartments = []
        
        for i, (x1, y1, x2, y2) in enumerate(compartment_boundaries):
            # Ensure coordinates are within image bounds
            x1 = max(0, x1)
            y1 = max(0, y1)
            x2 = min(image.shape[1], x2)
            y2 = min(image.shape[0], y2)
            
            # Skip invalid or too small regions
            if x2 <= x1 or y2 <= y1 or (x2 - x1) < 10 or (y2 - y1) < 10:
                logger.warning(f"Skipping invalid compartment boundary: ({x1}, {y1}, {x2}, {y2})")
                continue
            
            # Extract compartment region
            compartment = image[y1:y2, x1:x2].copy()
            
            # Track maximum dimensions
            max_width = max(max_width, compartment.shape[1])
            max_height = max(max_height, compartment.shape[0])
            
            extracted_compartments.append((compartment, i))
        
        # Pad all compartments to uniform size without changing the color space
        for compartment, i in extracted_compartments:
            # Get the original color type
            is_grayscale = len(compartment.shape) == 2
            channels = 1 if is_grayscale else compartment.shape[2]
            
            # Create a properly sized container with the same type as original image
            if is_grayscale:
                # For grayscale, use white background (255)
                padded = np.ones((max_height, max_width), dtype=np.uint8) * 255
            elif channels == 3:  # BGR
                # For color, use white background [255, 255, 255]
                padded = np.ones((max_height, max_width, 3), dtype=np.uint8) * 255
            elif channels == 4:  # BGRA
                # For BGRA, use transparent white [255, 255, 255, 0]
                padded = np.ones((max_height, max_width, 4), dtype=np.uint8) * 255
                padded[:, :, 3] = 0  # Set alpha channel to transparent
            else:
                # Unexpected format, use BGR with white background
                padded = np.ones((max_height, max_width, 3), dtype=np.uint8) * 255
            
            # Calculate placement to center the compartment
            y_offset = (max_height - compartment.shape[0]) // 2
            x_offset = (max_width - compartment.shape[1]) // 2
            
            # Handle different image formats correctly
            if is_grayscale:
                # For grayscale images
                padded[y_offset:y_offset+compartment.shape[0], x_offset:x_offset+compartment.shape[1]] = compartment
            else:
                # For color images (BGR or BGRA)
                if channels == 4 and padded.shape[2] == 4:
                    # Set the alpha channel to fully opaque for the compartment area
                    padded[y_offset:y_offset+compartment.shape[0], x_offset:x_offset+compartment.shape[1], 3] = 255
                
                # Copy the compartment to the padded image
                padded[y_offset:y_offset+compartment.shape[0], x_offset:x_offset+compartment.shape[1], :channels] = compartment[:, :, :channels]
            
            # Add to final compartments list
            compartments.append(padded)
            
            # Draw rectangle on visualization image with compartment number
            cv2.rectangle(viz_image, 
                        (x1, y1), 
                        (x2, y2), 
                        (0, 255, 0), 2)
            cv2.putText(viz_image, 
                        f"{i+1}", 
                        (x1 + 10, y1 + 30), 
                        cv2.FONT_HERSHEY_SIMPLEX, 
                        1.0, 
                        (0, 0, 255), 2)
        
        return compartments, viz_image

    def on_generate_trace(self):
        """Handle the 'Generate Drillhole Trace' button click."""
        try:
            # Use the FileManager's directory structure
            compartment_dir = self.file_manager.dir_structure["chip_compartments"]
            
            # Check if the directory exists
            if not os.path.exists(compartment_dir):
                messagebox.showerror("Error", f"Compartment directory not found: {compartment_dir}")
                return

            # Select CSV file
            csv_path = filedialog.askopenfilename(
                title="Select CSV File",
                filetypes=[("CSV Files", "*.csv")]
            )
            if not csv_path:
                return

            # Initialize the trace generator with file manager reference
            trace_generator = DrillholeTraceGenerator(
                config=self.config, 
                progress_queue=self.progress_queue,
                root=self.root,
                file_manager=self.file_manager  # Pass the FileManager instance
            )

            # TODO - The optional columns don't do anything, they're not displayed anywhere
            # Let user select optional columns 
            csv_columns = trace_generator.get_csv_columns(csv_path)
            if not csv_columns:
                messagebox.showerror("CSV Error", "Could not read columns from CSV.")
                return

            selected_columns = trace_generator.select_csv_columns(csv_columns)
            
            # Get the drill traces directory to check for existing traces
            traces_dir = self.file_manager.dir_structure["drill_traces"]
            
            # Get list of existing trace files
            existing_traces = set()
            if os.path.exists(traces_dir):
                existing_traces = {os.path.splitext(f)[0].split('_')[0] for f in os.listdir(traces_dir) 
                                if f.lower().endswith(('.png', '.jpg')) and '_Trace' in f}
            
            # Get list of holes from compartment directories
            hole_dirs = [d for d in os.listdir(compartment_dir) if os.path.isdir(os.path.join(compartment_dir, d))]
            
            # Filter to holes that don't have traces
            holes_to_process = [hole for hole in hole_dirs if hole not in existing_traces]
            
            if not holes_to_process:
                messagebox.showinfo("Info", "All holes already have trace images.")
                return
            
            # Ask user for confirmation
            if not messagebox.askyesno("Confirm", f"Found {len(holes_to_process)} holes without trace images. Process them all?"):
                return
            
            # Run the trace generation for these specific holes
            generated_paths = trace_generator.process_selected_holes(
                compartment_dir=compartment_dir,
                csv_path=csv_path,
                selected_columns=selected_columns,
                hole_ids=holes_to_process
            )

            if generated_paths:
                messagebox.showinfo("Success", f"Generated {len(generated_paths)} drillhole trace images.")
                
                # Ask if the user wants to open the directory
                if messagebox.askyesno("Open Directory", "Would you like to open the directory containing the trace images?"):
                    # Get the directory from the first generated path
                    if os.path.isfile(generated_paths[0]):
                        trace_dir = os.path.dirname(generated_paths[0])
                    else:
                        trace_dir = generated_paths[0]  # In case it's already a directory
                        
                    # Open the directory in file explorer
                    try:
                        if platform.system() == "Windows":
                            os.startfile(trace_dir)
                        elif platform.system() == "Darwin":  # macOS
                            subprocess.Popen(["open", trace_dir])
                        else:  # Linux
                            subprocess.Popen(["xdg-open", trace_dir])
                    except Exception as e:
                        messagebox.showwarning("Error", f"Could not open directory: {str(e)}")
            else:
                messagebox.showwarning("No Output", "No drillhole trace images were generated.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            logger.error(f"Error in on_generate_trace: {str(e)}")
            logger.error(traceback.format_exc())
            
    def create_gui(self):
        """Create a GUI for chip tray extraction with enhanced status display."""
        self.root = tk.Tk()
        self.root.title("Chip Tray Extractor")
        
        # Set up the main frame with padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Chip Tray Extractor ", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 15))
        
        # Input frame - always expanded
        input_frame = ttk.LabelFrame(main_frame, text="Input", padding=10)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        folder_frame = ttk.Frame(input_frame)
        folder_frame.pack(fill=tk.X)
        
        folder_label = ttk.Label(folder_frame, text="Input Folder:", width=15, anchor='w')
        folder_label.pack(side=tk.LEFT)
        
        self.folder_var = tk.StringVar()
        folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_var)
        folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_button = ttk.Button(folder_frame, text="Browse", command=self.browse_folder)
        browse_button.pack(side=tk.RIGHT)
        
        # Add compartment interval setting
        interval_frame = ttk.Frame(input_frame)
        interval_frame.pack(fill=tk.X, pady=(10, 0))
        
        interval_label = ttk.Label(interval_frame, text="Compartment Interval (m):", width=20, anchor='w')
        interval_label.pack(side=tk.LEFT)
        
        # Create a dropdown for common interval choices
        self.interval_var = tk.DoubleVar(value=self.config['compartment_interval'])
        interval_choices = [1.0, 2.0]
        interval_dropdown = ttk.Combobox(interval_frame, textvariable=self.interval_var, 
                                    values=interval_choices, width=5)
        interval_dropdown.pack(side=tk.LEFT)
        
        # Output settings - always expanded
        output_frame = ttk.LabelFrame(main_frame, text="Output Settings", padding=10)
        output_frame.pack(fill=tk.X, pady=(0, 10))

        # Centralized storage information
        storage_frame = ttk.Frame(output_frame)
        storage_frame.pack(fill=tk.X, pady=2)

        storage_label = ttk.Label(storage_frame, text="Local Output Location:", width=15, anchor='w')
        storage_label.pack(side=tk.LEFT)

        # Display read-only output location (the centralized location)
        self.output_folder_var = tk.StringVar(value=self.file_manager.processed_dir)
        storage_entry = ttk.Entry(storage_frame, textvariable=self.output_folder_var, state='readonly')
        storage_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Add info button to explain file structure
        info_button = ttk.Button(storage_frame, text="?", width=2, command=self._show_file_structure_info)
        info_button.pack(side=tk.RIGHT, padx=(5, 0))

        # Add OneDrive Settings - initially collapsed
        onedrive_collapsible = CollapsibleFrame(main_frame, text="OneDrive Path Settings", expanded=False)
        onedrive_collapsible.pack(fill=tk.X, pady=(0, 10))

        # Create variables for OneDrive paths
        self.approved_path_var = tk.StringVar()
        self.processed_originals_path_var = tk.StringVar()
        self.drill_traces_path_var = tk.StringVar()

        # Initialize variables with default values from OneDrivePathManager
        onedrive_manager = OneDrivePathManager()
        default_project_path = "Gabon - Belinga - Exploration Drilling"
        default_project_path_alt = os.path.join("Shared Documents", "Exploration Drilling")

        # Set defaults to display in the GUI
        self.approved_path_var.set(os.path.join(default_project_path, "03 - Reverse Circulation", "Chip Tray Photos", "1) Chip Tray Register and Images", "Approved Compartment Images"))
        self.processed_originals_path_var.set(os.path.join(default_project_path, "03 - Reverse Circulation", "Chip Tray Photos", "4) Processed Originals"))
        self.drill_traces_path_var.set(os.path.join(default_project_path, "03 - Reverse Circulation", "Chip Tray Photos", "5) Drill Traces"))

        # Create path input fields
        self._create_onedrive_path_field(onedrive_collapsible.content_frame, "Approved Folder:", self.approved_path_var)
        self._create_onedrive_path_field(onedrive_collapsible.content_frame, "Processed Originals:", self.processed_originals_path_var)
        self._create_onedrive_path_field(onedrive_collapsible.content_frame, "Drill Traces:", self.drill_traces_path_var)

        # Add Register path input
        self.register_path_var = tk.StringVar()
        self.register_path_var.set(os.path.join(default_project_path, "03 - Reverse Circulation", "Chip Tray Photos", "1) Chip Tray Register and Images", "Chip Tray Photo Register (Automatic).xlsx"))
        self._create_onedrive_path_field(onedrive_collapsible.content_frame, "Excel Register:", self.register_path_var)

        # Add a button to save settings
        save_onedrive_button = ttk.Button(
            onedrive_collapsible.content_frame,
            text="Apply Path Settings",
            command=self._update_onedrive_paths
        )
        save_onedrive_button.pack(anchor="e", pady=(10, 5))

        # Output format
        format_frame = ttk.Frame(output_frame)
        format_frame.pack(fill=tk.X, pady=2)
        
        format_label = ttk.Label(format_frame, text="Output Format:", width=15, anchor='w')
        format_label.pack(side=tk.LEFT)
        
        self.format_var = tk.StringVar(value=self.config['output_format'])
        format_options = ['jpg', 'png', 'tiff']
        format_dropdown = ttk.OptionMenu(format_frame, self.format_var, self.config['output_format'], *format_options)
        format_dropdown.pack(side=tk.LEFT)
        


        # Save debug images option (moved to output frame)
        debug_frame = ttk.Frame(output_frame)
        debug_frame.pack(fill=tk.X, pady=2)
        
        self.debug_var = tk.BooleanVar(value=self.config['save_debug_images'])
        debug_check = ttk.Checkbutton(debug_frame, text="Save Debug Images", variable=self.debug_var)
        debug_check.pack(anchor='w')
        
        # Create collapsible frames for Blur Detection and OCR Settings
        # Blur Detection - initially collapsed
        blur_collapsible = CollapsibleFrame(main_frame, text="Blur Detection", expanded=False)
        blur_collapsible.pack(fill=tk.X, pady=(0, 10))
        
        # Move blur detection content to the collapsible frame's content_frame
        # Enable blur detection
        self.blur_enable_var = tk.BooleanVar(value=self.config['enable_blur_detection'])
        enable_check = ttk.Checkbutton(
            blur_collapsible.content_frame, 
            text="Enable Blur Detection", 
            variable=self.blur_enable_var,
            command=self._toggle_blur_settings
        )
        enable_check.pack(anchor='w', pady=(0, 5))
        
        # Blur threshold slider
        threshold_frame = ttk.Frame(blur_collapsible.content_frame)
        threshold_frame.pack(fill=tk.X, pady=2)
        
        threshold_label = ttk.Label(threshold_frame, text="Blur Threshold:", width=15, anchor='w')
        threshold_label.pack(side=tk.LEFT)
        
        self.blur_threshold_var = tk.DoubleVar(value=self.config['blur_threshold'])
        threshold_slider = ttk.Scale(
            threshold_frame, 
            from_=10.0, 
            to=500.0, 
            orient=tk.HORIZONTAL, 
            variable=self.blur_threshold_var
        )
        threshold_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        threshold_value = ttk.Label(threshold_frame, width=5)
        threshold_value.pack(side=tk.RIGHT)
        
        # Update threshold value label when slider changes
        def update_threshold_label(*args):
            threshold_value.config(text=f"{self.blur_threshold_var.get():.1f}")
        
        self.blur_threshold_var.trace_add("write", update_threshold_label)
        update_threshold_label()  # Initial update
        
        # ROI ratio slider
        roi_frame = ttk.Frame(blur_collapsible.content_frame)
        roi_frame.pack(fill=tk.X, pady=2)
        
        roi_label = ttk.Label(roi_frame, text="ROI Ratio:", width=15, anchor='w')
        roi_label.pack(side=tk.LEFT)
        
        self.blur_roi_var = tk.DoubleVar(value=self.config['blur_roi_ratio'])
        roi_slider = ttk.Scale(
            roi_frame, 
            from_=0.1, 
            to=1.0, 
            orient=tk.HORIZONTAL, 
            variable=self.blur_roi_var
        )
        roi_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        roi_value = ttk.Label(roi_frame, width=5)
        roi_value.pack(side=tk.RIGHT)
        
        # Update ROI value label when slider changes
        def update_roi_label(*args):
            roi_value.config(text=f"{self.blur_roi_var.get():.2f}")
        
        self.blur_roi_var.trace_add("write", update_roi_label)
        update_roi_label()  # Initial update
        
        # Flag blurry images checkbox
        self.flag_blurry_var = tk.BooleanVar(value=self.config['flag_blurry_images'])
        flag_check = ttk.Checkbutton(
            blur_collapsible.content_frame, 
            text="Flag Blurry Images", 
            variable=self.flag_blurry_var
        )
        flag_check.pack(anchor='w', pady=(5, 0))
        
        # Save blur visualizations checkbox
        self.save_blur_viz_var = tk.BooleanVar(value=self.config['save_blur_visualizations'])
        save_viz_check = ttk.Checkbutton(
            blur_collapsible.content_frame, 
            text="Save Blur Analysis Visualizations", 
            variable=self.save_blur_viz_var
        )
        save_viz_check.pack(anchor='w', pady=(5, 0))
        
        # Blurry threshold percentage
        threshold_pct_frame = ttk.Frame(blur_collapsible.content_frame)
        threshold_pct_frame.pack(fill=tk.X, pady=(5, 0))
        
        threshold_pct_label = ttk.Label(
            threshold_pct_frame, 
            text="Quality Alert Threshold:", 
            width=20, 
            anchor='w'
        )
        threshold_pct_label.pack(side=tk.LEFT)
        
        self.blur_threshold_pct_var = tk.DoubleVar(value=self.config['blurry_threshold_percentage'])
        threshold_pct_slider = ttk.Scale(
            threshold_pct_frame, 
            from_=5.0, 
            to=100.0, 
            orient=tk.HORIZONTAL, 
            variable=self.blur_threshold_pct_var
        )
        threshold_pct_slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        threshold_pct_value = ttk.Label(threshold_pct_frame, width=5)
        threshold_pct_frame.pack(side=tk.RIGHT)
        
        # Update threshold percentage value label when slider changes
        def update_threshold_pct_label(*args):
            threshold_pct_value.config(text=f"{self.blur_threshold_pct_var.get():.1f}%")
        
        self.blur_threshold_pct_var.trace_add("write", update_threshold_pct_label)
        update_threshold_pct_label()  # Initial update
        
        # Calibration button
        calibration_frame = ttk.Frame(blur_collapsible.content_frame)
        calibration_frame.pack(fill=tk.X, pady=(5, 0))
        
        calibrate_button = ttk.Button(
            calibration_frame,
            text="Calibrate Blur Detection",
            command=self._show_blur_calibration_dialog
        )
        calibrate_button.pack(side=tk.LEFT, padx=(0, 10))
        
        help_button = ttk.Button(
            calibration_frame,
            text="?",
            width=2,
            command=self._show_blur_help
        )
        help_button.pack(side=tk.RIGHT)
        
        # Store the blur_settings_frame for toggling
        self.blur_settings_frame = blur_collapsible.content_frame
        
        # OCR Settings - initially collapsed
        ocr_collapsible = CollapsibleFrame(main_frame, text="OCR Settings", expanded=False)
        ocr_collapsible.pack(fill=tk.X, pady=(0, 10))
        
        # Enable OCR checkbox
        self.ocr_enable_var = tk.BooleanVar(value=self.config['enable_ocr'])
        ocr_check = ttk.Checkbutton(
            ocr_collapsible.content_frame, 
            text="Enable OCR", 
            variable=self.ocr_enable_var,
            command=self._toggle_ocr_settings
        )
        ocr_check.pack(anchor='w')
        
        # Prefix validation checkbox
        self.prefix_validation_var = tk.BooleanVar(value=self.config.get('enable_prefix_validation', True))
        prefix_check = ttk.Checkbutton(
            ocr_collapsible.content_frame, 
            text="Validate Hole ID Prefixes", 
            variable=self.prefix_validation_var,
            command=self._toggle_prefix_settings
        )
        prefix_check.pack(anchor='w', padx=(20, 0))
        
        # Prefix list frame
        prefix_frame = ttk.Frame(ocr_collapsible.content_frame)
        prefix_frame.pack(fill=tk.X, pady=5)
        
        prefix_label = ttk.Label(prefix_frame, text="Valid Prefixes (comma separated):", width=25, anchor='w')
        prefix_label.pack(side=tk.LEFT, padx=(40, 5))
        
        # Convert list to comma-separated string for display
        prefix_str = ", ".join(self.config.get('valid_hole_prefixes', ['BA', 'NB', 'SB', 'KM']))
        self.prefix_var = tk.StringVar(value=prefix_str)
        prefix_entry = ttk.Entry(prefix_frame, textvariable=self.prefix_var)
        prefix_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Store reference to prefix frame for toggling
        self.prefix_frame = prefix_frame
        
        # Initialize the UI state based on checkboxes
        self._toggle_blur_settings()
        self._toggle_ocr_settings()
        self._toggle_prefix_settings()
        
        # Progress bar
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding=10)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                        orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X)
        
        # Status text - enlarged and improved for better visibility
        status_frame = ttk.LabelFrame(main_frame, text="Detailed Status", padding=10)
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Add a text widget with scrollbar
        self.status_text = tk.Text(status_frame, height=15, wrap=tk.WORD, font=("Consolas", 10))
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(status_frame, command=self.status_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.status_text.config(yscrollcommand=scrollbar.set)
        self.status_text.config(state=tk.DISABLED)
        
        # Add text tags for different status types (error, warning, success)
        self.status_text.tag_configure("error", foreground="red")
        self.status_text.tag_configure("warning", foreground="orange")
        self.status_text.tag_configure("success", foreground="green")
        self.status_text.tag_configure("info", foreground="blue")
        
        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.process_button = ttk.Button(button_frame, text="Process Photos", 
                                    command=self.start_processing)
        self.process_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        quit_button = ttk.Button(button_frame, text="Quit", command=self.quit_app)
        quit_button.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(5, 0))
        
        # Add Generate Drillhole Trace button below main buttons
        control_frame = ttk.Frame(main_frame, padding=(0, 5, 0, 0))
        control_frame.pack(fill=tk.X)
        
        self.trace_button = ttk.Button(
            control_frame,
            text="Generate Drillhole Trace",
            command=self.on_generate_trace
        )
        self.trace_button.pack(fill=tk.X)


        # Add Review button below the trace button
        self.review_button = ttk.Button(
            control_frame,
            text="Review Extracted Images",
            command=self._start_image_review
        )
        self.review_button.pack(fill=tk.X, pady=(5, 0))
        
        # Set up a timer to check for progress updates
        self.root.after(100, self.check_progress)
        
        # Add initial status message
        self.update_status("Ready. Select a folder and click 'Process Photos'.", "info")

        # After setting up all your existing UI components, add a menu bar
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Browse Folder...", command=self.browse_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit_app)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="Check for Updates", command=self.on_check_for_updates)
        help_menu.add_command(label="About", command=self._show_about_dialog)
        
        # Check for updates at startup if enabled
        if self.config.get('check_for_updates', True):
            # Schedule update check after GUI is fully loaded
            self.root.after(2000, self._check_updates_at_startup)


    def on_check_for_updates(self):
        self.update_checker.check_for_updates(parent=self.root)

    # Add helper methods for the About dialog and startup update check
    def _show_about_dialog(self):
        """Show information about the application."""
        version = self.update_checker.get_local_version() if hasattr(self, 'update_checker') else "Unknown"
        
        about_text = (
            f"Chip Tray Extractor v{version}\n\n"
            "A tool to extract individual compartment images from\n"
            "panoramic chip tray photos using ArUco markers.\n\n"
            "Author: George Symonds\n"
            "GitHub: https://github.com/Dragoarms/Geological-Chip-Tray-Compartment-Extractor"
        )
        
        messagebox.showinfo("About Chip Tray Extractor", about_text, parent=self.root)

    def _check_updates_at_startup(self):
        """Check for updates at startup without showing dialogs for up-to-date case."""
        if not hasattr(self, 'update_checker'):
            return
            
        try:
            result = self.update_checker.compare_versions()

            if result["update_available"]:
                if messagebox.askyesno("Update Available", f"A new version is available:\n{result['github_version']}.\n\nDownload and restart?"):
                    self.update_checker.download_and_replace_script(self.file_manager)

    def _create_onedrive_path_field(self, parent, label_text, string_var):
            """Create a field for OneDrive path input with browse button."""
            frame = ttk.Frame(parent)
            frame.pack(fill=tk.X, pady=5)
            
            label = ttk.Label(frame, text=label_text, width=15, anchor='w')
            label.pack(side=tk.LEFT)
            
            entry = ttk.Entry(frame, textvariable=string_var)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
            
            browse_button = ttk.Button(
                frame, 
                text="Browse", 
                command=lambda: self._browse_onedrive_path(string_var)
            )
            browse_button.pack(side=tk.RIGHT)

    def _browse_onedrive_path(self, string_var):
        """Open folder browser and update path variable."""
        folder_path = filedialog.askdirectory(title="Select OneDrive folder")
        if folder_path:
            string_var.set(folder_path)

    def _update_onedrive_paths(self):
        """Update OneDrive paths in the OneDrivePathManager."""
        if not hasattr(self, 'onedrive_manager'):
            self.onedrive_manager = OneDrivePathManager(self.root)
        
        # Update custom paths in OneDrivePathManager
        self.onedrive_manager._approved_folder_path = self.approved_path_var.get() if self.approved_path_var.get() else None
        self.onedrive_manager._processed_originals_path = self.processed_originals_path_var.get() if self.processed_originals_path_var.get() else None
        self.onedrive_manager._chip_tray_folder_path = os.path.dirname(os.path.dirname(self.processed_originals_path_var.get())) if self.processed_originals_path_var.get() else None
        self.onedrive_manager._register_path = self.register_path_var.get() if self.register_path_var.get() else None
        
        # Show confirmation message
        messagebox.showinfo(
            "OneDrive Paths Updated", 
            "OneDrive path settings have been updated."
        )


    def _toggle_ocr_settings(self):
        """Enable/disable OCR settings based on checkbox state."""
        if hasattr(self, 'prefix_check') and hasattr(self, 'prefix_validation_var'):
            if self.ocr_enable_var.get():
                # Enable all OCR settings
                self.prefix_check.configure(state='normal')
                self._toggle_prefix_settings()  # Update prefix entry state
            else:
                # Disable all OCR settings
                self.prefix_check.configure(state='disabled')
                for child in self.prefix_frame.winfo_children():
                    if isinstance(child, (ttk.Entry, ttk.Checkbutton)):
                        child.configure(state='disabled')

    def _toggle_prefix_settings(self):
        """Enable/disable prefix settings based on checkbox state."""
        if hasattr(self, 'prefix_frame') and hasattr(self, 'prefix_validation_var'):
            if self.ocr_enable_var.get() and self.prefix_validation_var.get():
                # Enable the prefix entry
                for child in self.prefix_frame.winfo_children():
                    if isinstance(child, ttk.Entry):
                        child.configure(state='normal')
            else:
                # Disable the prefix entry
                for child in self.prefix_frame.winfo_children():
                    if isinstance(child, ttk.Entry):
                        child.configure(state='disabled')


    def _start_image_review(self):
        """Start the image review process."""
        if hasattr(self, 'qaqc_manager'):
            if not self.qaqc_manager.pending_trays:
                messagebox.showinfo("No Images", "No images available for review. Process some images first.")
                return
            self.qaqc_manager.start_review_process()
        else:
            messagebox.showinfo("No Images", "No images available for review. Process some images first.")

    def _show_file_structure_info(self):
        """Show information about the new file structure."""
        info_message = (
            "Files are now saved in a centralized location with the following structure:\n\n"
            f"{self.file_manager.processed_dir}\n"
            "├── Blur Analysis\\\n"
            "├── Chip Compartments\\\n"
            "├── Debug Images\\\n"
            "├── Drill Traces\\\n"
            "├── Processed Originals\\\n"
            "└── Failed and Skipped Originals\\\n\n"
            "Each folder (except Drill Traces) contains subfolders for each Hole ID.\n\n"
            "File naming follows the pattern:\n"
            "- Debug: HoleID_From-To_Debug_[type].jpg\n"
            "- Compartments: HoleID_CC_[number].png\n"
            "- Blur Analysis: HoleID_[number]_blur_analysis.jpg\n"
            "- Originals: HoleID_From-To_Original.ext"
        )
        messagebox.showinfo("File Structure Information", info_message)
    
    def _toggle_blur_settings(self):
        """Enable or disable blur detection settings based on checkbox state."""
        if self.blur_enable_var.get():
            # Enable all blur settings
            for child in self.blur_settings_frame.winfo_children():
                for widget in child.winfo_children():
                    if isinstance(widget, (ttk.Scale, ttk.Checkbutton, ttk.Button, ttk.Entry)):
                        widget.configure(state='normal')
        else:
            # Disable all blur settings
            for child in self.blur_settings_frame.winfo_children():
                for widget in child.winfo_children():
                    if isinstance(widget, (ttk.Scale, ttk.Checkbutton, ttk.Button, ttk.Entry)):
                        widget.configure(state='disabled')

    def _show_blur_help(self):
        """Show help information about blur detection."""
        help_text = """
    Blur Detection Help

    The blur detector identifies blurry images using the Laplacian variance method:

    - Blur Threshold: Lower values make the detector more sensitive (detecting more images as blurry).
    Typical values range from 50-200 depending on image content.

    - ROI Ratio: Percentage of the central image to analyze. Use higher values for more complete analysis,
    lower values to focus on the center where subjects are typically sharpest.

    - Quality Alert Threshold: Percentage of blurry compartments that will trigger a quality alert.

    - Calibration: Use the calibration tool to automatically set an optimal threshold based on example
    sharp and blurry images.
    """
        messagebox.showinfo("Blur Detection Help", help_text)

    def _show_blur_calibration_dialog(self):
        """Show a dialog for calibrating blur detection."""
        # Create a dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Calibrate Blur Detection")
        dialog.geometry("600x500")
        dialog.grab_set()  # Make dialog modal
        
        # Main frame with padding
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Instructions
        instructions = ttk.Label(
            main_frame,
            text="Select example sharp and blurry images to calibrate the blur detection threshold.",
            wraplength=580,
            justify=tk.LEFT
        )
        instructions.pack(fill=tk.X, pady=(0, 10))
        
        # Frame for sharp images
        sharp_frame = ttk.LabelFrame(main_frame, text="Sharp (Good) Images", padding=10)
        sharp_frame.pack(fill=tk.X, pady=(0, 10))
        
        sharp_path_var = tk.StringVar()
        sharp_entry = ttk.Entry(sharp_frame, textvariable=sharp_path_var)
        sharp_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        sharp_button = ttk.Button(
            sharp_frame,
            text="Browse",
            command=lambda: self._select_calibration_images(sharp_path_var)
        )
        sharp_button.pack(side=tk.RIGHT)
        
        # Frame for blurry images
        blurry_frame = ttk.LabelFrame(main_frame, text="Blurry (Poor) Images", padding=10)
        blurry_frame.pack(fill=tk.X, pady=(0, 10))
        
        blurry_path_var = tk.StringVar()
        blurry_entry = ttk.Entry(blurry_frame, textvariable=blurry_path_var)
        blurry_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        blurry_button = ttk.Button(
            blurry_frame,
            text="Browse",
            command=lambda: self._select_calibration_images(blurry_path_var)
        )
        blurry_button.pack(side=tk.RIGHT)
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Calibration Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Results text widget
        results_text = tk.Text(results_frame, height=8, wrap=tk.WORD)
        results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar for results
        scrollbar = ttk.Scrollbar(results_frame, command=results_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        results_text.config(yscrollcommand=scrollbar.set)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        # Calibrate button
        def calibrate():
            # Get paths
            sharp_paths = sharp_path_var.get().split(';')
            blurry_paths = blurry_path_var.get().split(';')
            
            # Validate
            if not sharp_paths or not sharp_paths[0] or not blurry_paths or not blurry_paths[0]:
                messagebox.showerror("Error", "Please select both sharp and blurry images")
                return
            
            # Load images
            sharp_images = []
            blurry_images = []
            
            results_text.delete(1.0, tk.END)
            results_text.insert(tk.END, "Loading images...\n")
            
            try:
                # Load sharp images
                for path in sharp_paths:
                    if path and os.path.exists(path):
                        img = cv2.imread(path)
                        if img is not None:
                            sharp_images.append(img)
                            results_text.insert(tk.END, f"Loaded sharp image: {os.path.basename(path)}\n")
                
                # Load blurry images
                for path in blurry_paths:
                    if path and os.path.exists(path):
                        img = cv2.imread(path)
                        if img is not None:
                            blurry_images.append(img)
                            results_text.insert(tk.END, f"Loaded blurry image: {os.path.basename(path)}\n")
                
                if not sharp_images or not blurry_images:
                    results_text.insert(tk.END, "Error: Failed to load images\n")
                    return
                
                # Calibrate threshold
                results_text.insert(tk.END, "\nCalibrating...\n")
                
                old_threshold = self.blur_detector.threshold
                new_threshold = self.blur_detector.calibrate_threshold(sharp_images, blurry_images)
                
                # Update UI
                self.blur_threshold_var.set(new_threshold)
                
                # Log results
                results_text.insert(tk.END, f"\nResults:\n")
                results_text.insert(tk.END, f"Old threshold: {old_threshold:.2f}\n")
                results_text.insert(tk.END, f"New threshold: {new_threshold:.2f}\n")
                
                # Show variances
                results_text.insert(tk.END, "\nImage Variances:\n")
                
                for i, img in enumerate(sharp_images):
                    variance = self.blur_detector.get_laplacian_variance(img)
                    results_text.insert(tk.END, f"Sharp image {i+1}: {variance:.2f}\n")
                
                for i, img in enumerate(blurry_images):
                    variance = self.blur_detector.get_laplacian_variance(img)
                    results_text.insert(tk.END, f"Blurry image {i+1}: {variance:.2f}\n")
                
                # Remind user to save settings
                results_text.insert(tk.END, "\nRemember to click 'OK' to apply the new threshold.\n")
                
            except Exception as e:
                results_text.insert(tk.END, f"Error during calibration: {str(e)}\n")
                logger.error(f"Calibration error: {str(e)}")
                logger.error(traceback.format_exc())
        
        calibrate_button = ttk.Button(
            button_frame,
            text="Calibrate",
            command=calibrate
        )
        calibrate_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # OK button
        ok_button = ttk.Button(
            button_frame,
            text="OK",
            command=dialog.destroy
        )
        ok_button.pack(side=tk.RIGHT)

    
    
    def _select_calibration_images(self, path_var):
        """
        Show file dialog to select images for calibration.
        
        Args:
            path_var: StringVar to store selected paths
        """
        file_paths = filedialog.askopenfilenames(
            title="Select Images",
            filetypes=[
                ("Image files", "*.jpg *.jpeg *.png *.tif *.tiff"),
                ("All files", "*.*")
            ]
        )
        
        if file_paths:
            # Join paths with semicolons for display
            path_var.set(';'.join(file_paths))

        # Progress bar
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding=10)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                        orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X)
        
        # Status text - enlarged and improved for better visibility
        status_frame = ttk.LabelFrame(main_frame, text="Detailed Status", padding=10)
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Add a text widget with scrollbar
        self.status_text = tk.Text(status_frame, height=15, wrap=tk.WORD, font=("Consolas", 10))
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(status_frame, command=self.status_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.status_text.config(yscrollcommand=scrollbar.set)
        self.status_text.config(state=tk.DISABLED)
        
        # Add text tags for different status types (error, warning, success)
        self.status_text.tag_configure("error", foreground="red")
        self.status_text.tag_configure("warning", foreground="orange")
        self.status_text.tag_configure("success", foreground="green")
        self.status_text.tag_configure("info", foreground="blue")
        
        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 5))

        
        control_frame = ttk.Frame(self.root, padding="10")
        control_frame.pack(fill=tk.X, pady=5)

        # Generate Drillhole Trace
        self.trace_button = ttk.Button(
            control_frame,
            text="Generate Drillhole Trace",
            command=self.on_generate_trace
        )
        self.trace_button.pack(side=tk.LEFT, padx=5, pady=5)

        
        self.process_button = ttk.Button(button_frame, text="Process Photos", 
                                    command=self.start_processing)
        self.process_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        quit_button = ttk.Button(button_frame, text="Quit", command=self.quit_app)
        quit_button.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(5, 0))
        
        
        # Set up a timer to check for progress updates
        self.root.after(100, self.check_progress)
        
        # Add initial status message
        self.update_status("Ready. Select a folder and click 'Process Photos'.", "info")

    
    
    def update_status(self, message: str, status_type: str = None) -> None:
        """
        Add a message to the status text widget with optional formatting.
        
        Args:
            message: The message to display
            status_type: Optional status type for formatting (error, warning, success, info)
        """
        try:
            # Check if status_text exists
            if not hasattr(self, 'status_text') or self.status_text is None:
                # Just log the message
                logger.info(message)
                return
                
            self.status_text.config(state=tk.NORMAL)
            
            # Add timestamp
            import datetime
            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            self.status_text.insert(tk.END, f"[{timestamp}] ", "info")
            
            # Add message with appropriate tag if specified
            if status_type and status_type in ["error", "warning", "success", "info"]:
                self.status_text.insert(tk.END, f"{message}\n", status_type)
            else:
                self.status_text.insert(tk.END, f"{message}\n")
            
            self.status_text.see(tk.END)
            self.status_text.config(state=tk.DISABLED)
        except Exception as e:
            # Log error but don't raise it to avoid UI crashes
            logger.error(f"Error updating status: {str(e)}")
    
    
    
    def save_compartments(self, 
                        compartments: List[np.ndarray], 
                        output_dir: str, 
                        base_filename: str,
                        metadata: Optional[Dict[str, Any]] = None) -> int:
        """
        Save extracted compartment images to disk with correct depth ranges.
        
        Args:
            compartments: List of compartment images
            output_dir: Directory to save compartment images (legacy parameter, now using FileManager)
            base_filename: Base filename to use for compartment images
            metadata: Optional metadata for naming compartments
            
        Returns:
            Number of successfully saved compartments
        """
        # Detect blur in compartments
        blur_results = self.detect_blur_in_compartments(compartments, base_filename)
        
        # Add blur indicators to images if enabled
        if self.config['flag_blurry_images']:
            compartments = self.add_blur_indicators(compartments, blur_results)
        
        # Save each compartment
        saved_count = 0
        blurry_count = 0
        
        # Calculate expected number of compartments and depth range for each
        expected_count = self.config['compartment_count']
        
        # Get the compartment interval from config
        compartment_interval = self.config['compartment_interval']
        
        # Check if we have valid metadata
        if not metadata or not all(key in metadata for key in ['hole_id', 'depth_from', 'depth_to']):
            # If metadata is missing, this is a critical error - log and return
            logger.error("Missing required metadata for saving compartments. Cannot proceed.")
            return 0
        
        # Extract metadata values
        hole_id = metadata['hole_id']
        depth_from = float(metadata['depth_from'])
        depth_to = float(metadata['depth_to'])
        
        # Calculate depth increment per compartment using the configured interval
        depth_increment = compartment_interval
        
        # Save blur visualizations if enabled
        if self.config['enable_blur_detection'] and self.config['save_blur_visualizations']:
            for result in blur_results:
                if 'visualization' in result:
                    i = result['index']
                    
                    # Calculate depth for this compartment using the interval
                    comp_depth_from = depth_from + (i * depth_increment)
                    comp_depth_to = comp_depth_from + depth_increment
                    
                    # End depth is used as the compartment number (when interval is 1.0)
                    compartment_depth = int(comp_depth_to)
                    
                    if is_blurry := result.get('is_blurry', False):
                        blurry_count += 1
                        
                    # Save blur analysis visualization
                    self.file_manager.save_blur_analysis(
                        result['visualization'],
                        hole_id,
                        compartment_depth
                    )
        
        # Calculate depth increment per compartment using the configured interval
        compartment_interval = self.config['compartment_interval']

        # Process each compartment
        for i, compartment in enumerate(compartments):
            try:
                # Calculate depth for this compartment using the interval
                # For depths like 40-60, first compartment (i=0) should be at depth 41
                comp_depth_from = depth_from + (i * compartment_interval)
                comp_depth_to = comp_depth_from + compartment_interval
                
                # The compartment is identified by its end depth
                compartment_depth = int(comp_depth_to)
                
                # Save using the FileManager with depth-based numbering
                self.file_manager.save_compartment(
                    compartment,
                    hole_id,
                    compartment_depth,
                    False,  # has_data
                    self.config['output_format']
                )
                
                saved_count += 1
                    
            except Exception as e:
                logger.error(f"Error saving compartment {i+1}: {str(e)}")
        
        # Log blur detection summary
        if blurry_count > 0:
            logger.warning(f"Detected {blurry_count}/{len(compartments)} blurry compartments")
            
            # Calculate percentage of blurry compartments
            blurry_percentage = (blurry_count / len(compartments)) * 100
            threshold_percentage = self.config['blurry_threshold_percentage']
            
            if blurry_percentage >= threshold_percentage:
                logger.warning(f"QUALITY ALERT: {blurry_percentage:.1f}% of compartments are blurry (threshold: {threshold_percentage}%)")
                if hasattr(self, 'progress_queue'):
                    self.progress_queue.put((f"QUALITY ALERT: {blurry_percentage:.1f}% of compartments are blurry!", "warning"))
        
        # Return just the saved count
        return saved_count
        
    def process_folder_with_progress(self, folder_path: str):
        """
        Process folder with progress updates for GUI.
        
        Args:
            folder_path: Path to folder containing images
        """
        try:
            # Update progress
            self.progress_queue.put(("Starting processing...", 0))
            
            # Process the folder
            successful, failed = self.process_folder(folder_path)
            
            # Update progress
            self.progress_queue.put((f"Processing complete: {successful} successful, {failed} failed", 100))
            
        except Exception as e:
            self.progress_queue.put((f"Error: {str(e)}", None))
            logger.error(f"Error processing folder: {str(e)}")
        finally:
            self.processing_complete = True

    def check_progress(self):
        """Check for progress updates from the processing thread with enhanced status handling."""
        while not self.progress_queue.empty():
            try:
                message, progress = self.progress_queue.get_nowait()
                
                # Update progress bar if progress value provided
                if progress is not None:
                    self.progress_var.set(progress)
                
                # Update status message with appropriate status type
                if message:
                    # Determine message type based on content
                    if any(error_term in message.lower() for error_term in ["error", "failed", "not enough"]):
                        self.update_status(message, "error")
                    elif any(warning_term in message.lower() for warning_term in ["warning", "missing"]):
                        self.update_status(message, "warning")
                    elif any(success_term in message.lower() for success_term in ["success", "complete", "saved"]):
                        self.update_status(message, "success")
                    else:
                        self.update_status(message, "info")
                        
                # If processing is complete, re-enable the process button
                if self.processing_complete:
                    self.process_button.config(state=tk.NORMAL)
                    
            except queue.Empty:
                pass
            
        # Schedule the next check
        self.root.after(100, self.check_progress)

    def create_visualization_image(self, 
                                    original_image: np.ndarray, 
                                    processing_images: List[Tuple[str, np.ndarray]]) -> np.ndarray:
        """
        Create a collage showing the original image alongside processing steps.
        
        Args:
            original_image: The original input image
            processing_images: List of (step_name, image) tuples
            
        Returns:
            Visualization collage as numpy array
        """
        # Helper function to ensure 3-channel image
        def ensure_3channel(img: np.ndarray) -> np.ndarray:
            if len(img.shape) == 2:  # Grayscale
                return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
            elif len(img.shape) == 3:
                if img.shape[2] == 1:  # Single channel
                    return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
                elif img.shape[2] == 3:  # Already color
                    return img
                elif img.shape[2] == 4:  # RGBA
                    return cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
            
            # Fallback - create a blank color image
            return np.zeros((img.shape[0], img.shape[1], 3), dtype=np.uint8)

        # Determine the number of images in the collage
        num_images = 1 + len(processing_images)
        
        # Determine layout (1 row if ≤ 3 images, 2 rows otherwise)
        if num_images <= 3:
            rows, cols = 1, num_images
        else:
            rows = 2
            cols = (num_images + 1) // 2  # Ceiling division
        
        # Resize images to a standard height
        target_height = 600
        resized_images = []
        
        # Resize original image - ensure 3 channels
        original_image = ensure_3channel(original_image)
        h, w = original_image.shape[:2]
        scale = target_height / h
        resized_original = cv2.resize(original_image, (int(w * scale), target_height))
        resized_images.append(("Original Image", resized_original))
        
        # Resize processing images
        for name, img in processing_images:
            # Ensure 3 channels before resizing
            img = ensure_3channel(img)
            h, w = img.shape[:2]
            scale = target_height / h
            resized = cv2.resize(img, (int(w * scale), target_height))
            resized_images.append((name, resized))
        
        # Calculate canvas size
        max_width = max(img[1].shape[1] for img in resized_images)
        canvas_width = max_width * cols + 20 * (cols + 1)  # Add padding
        canvas_height = target_height * rows + 60 * (rows + 1)  # Add space for titles
        
        # Create canvas
        canvas = np.ones((canvas_height, canvas_width, 3), dtype=np.uint8) * 255
        
        # Place images on canvas
        for i, (name, img) in enumerate(resized_images):
            row = i // cols
            col = i % cols
            
            # Calculate position
            x = 20 + col * (max_width + 20)
            y = 60 + row * (target_height + 60)
            
            # Place image
            h, w = img.shape[:2]
            canvas[y:y+h, x:x+w] = img
            
            # Add title
            cv2.putText(canvas, name, (x, y - 20), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 0), 2)
        
        # Add main title
        cv2.putText(canvas, "Chip Tray Extraction Process", 
                (canvas_width // 2 - 200, 30), 
                cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 2)
        
        return canvas

    
    def browse_folder(self):
        """Open folder browser dialog and update the folder entry."""
        folder_path = filedialog.askdirectory(title="Select folder with chip tray photos")
        if folder_path:
            self.folder_var.set(folder_path)
            # Check if status_text exists before updating
            if hasattr(self, 'status_text') and self.status_text:
                self.update_status(f"Selected folder: {folder_path}")

    def start_processing(self):
        """Start processing in a separate thread."""
        folder_path = self.folder_var.get()
        if not folder_path:
            messagebox.showerror("Error", "Please select a folder")
            return
        
        if not os.path.isdir(folder_path):
            messagebox.showerror("Error", "Selected path is not a valid folder")
            return
        
        # Update config with current GUI settings
        self.config['output_format'] = self.format_var.get()
        self.config['save_debug_images'] = self.debug_var.get()
        self.config['compartment_interval'] = self.interval_var.get()  # Use interval_var.get() instead of interval
        
        # Update blur detection settings
        self.config['enable_blur_detection'] = self.blur_enable_var.get()
        self.config['blur_threshold'] = self.blur_threshold_var.get()
        self.config['blur_roi_ratio'] = self.blur_roi_var.get()
        self.config['flag_blurry_images'] = self.flag_blurry_var.get()
        self.config['save_blur_visualizations'] = self.save_blur_viz_var.get()
        self.config['blurry_threshold_percentage'] = self.blur_threshold_pct_var.get()
        
        # Update OCR settings
        self.config['enable_ocr'] = self.ocr_enable_var.get()
        self.config['enable_prefix_validation'] = self.prefix_validation_var.get()
        
        # Parse the prefix string into a list
        prefix_str = self.prefix_var.get()
        if prefix_str:
            # Split by comma and strip whitespace
            prefixes = [p.strip().upper() for p in prefix_str.split(',')]
            # Filter out any empty or invalid entries
            self.config['valid_hole_prefixes'] = [p for p in prefixes if p and len(p) == 2 and p.isalpha()]
        
        # Update blur detector with new settings
        self.blur_detector.threshold = self.config['blur_threshold']
        self.blur_detector.roi_ratio = self.config['blur_roi_ratio']
        
        # Ensure TesseractManager has the updated config
        self.tesseract_manager.config = self.config
        
        # Inform user about the new file structure
        message = (f"Files will be saved in organized folders under:\n"
                f"{self.file_manager.processed_dir}\n\n"
                f"Do you want to continue?")
        
        if not messagebox.askyesno("Centralized File Storage", message):
            return
        
        # Clear status
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.config(state=tk.DISABLED)
        
        self.progress_var.set(0)
        self.processing_complete = False
        self.process_button.config(state=tk.DISABLED)
        
        # Start processing thread
        processing_thread = threading.Thread(
            target=self.process_folder_with_progress, 
            args=(folder_path,)
        )
        processing_thread.daemon = True
        processing_thread.start()
        
        self.update_status(f"Started processing folder: {folder_path}")
    
    
    def quit_app(self):
        """Close the application."""
        if self.root:
            self.root.destroy()

class CollapsibleFrame(ttk.Frame):
    """A frame that can be expanded or collapsed with a toggle button."""
    
    def __init__(self, parent, text="", expanded=False, **kwargs):
        """
        Initialize the collapsible frame.
        
        Args:
            parent: Parent widget
            text: Title text for the frame
            expanded: Whether the frame is initially expanded
            **kwargs: Additional arguments for ttk.Frame
        """
        ttk.Frame.__init__(self, parent, **kwargs)
        
        # Create a header frame
        self.header_frame = ttk.Frame(self)
        self.header_frame.pack(fill=tk.X, expand=False)
        
        # Create toggle button with arrow
        self.toggle_button = ttk.Label(
            self.header_frame, 
            text="▼ " if expanded else "▶ ",
            cursor="hand2"
        )
        self.toggle_button.pack(side=tk.LEFT, padx=(5, 0))
        self.toggle_button.bind("<Button-1>", self.toggle)
        
        # Create header label
        self.header_label = ttk.Label(
            self.header_frame, 
            text=text,
            cursor="hand2",
            font=("Arial", 10, "bold")
        )
        self.header_label.pack(side=tk.LEFT, padx=(5, 0))
        self.header_label.bind("<Button-1>", self.toggle)
        
        # Add a separator line
        separator = ttk.Separator(self, orient=tk.HORIZONTAL)
        separator.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        # Content frame
        self.content_frame = ttk.Frame(self, padding=(15, 5, 5, 5))
        
        # Set initial state
        self.expanded = expanded
        if expanded:
            self.content_frame.pack(fill=tk.X, expand=True, padx=5, pady=5)
        
    def toggle(self, event=None):
        """Toggle the expanded/collapsed state."""
        if self.expanded:
            self.content_frame.pack_forget()
            self.toggle_button.configure(text="▶ ")
            self.expanded = False
        else:
            self.content_frame.pack(fill=tk.X, expand=True, padx=5, pady=5)
            self.toggle_button.configure(text="▼ ")
            self.expanded = True

class DrillholeTraceGenerator:
    """
    A class to generate drillhole trace images by stitching together chip tray compartment images.
    Integrates with ChipTrayExtractor to create complete drillhole visualization.
    """
    
    def __init__(self, 
                config: Dict[str, Any] = None, 
                progress_queue: Optional[Any] = None,
                root: Optional[tk.Tk] = None,
                file_manager: Optional[Any] = None):
        """
        Initialize the Drillhole Trace Generator.
        
        Args:
            config: Configuration dictionary
            progress_queue: Optional queue for reporting progress
            root: Optional Tkinter root for dialog windows
        """
        self.progress_queue = progress_queue
        self.root = root
        self.file_manager = file_manager
        
        # Default configuration
        self.config = {
            'output_folder': 'drillhole_traces',
            'metadata_box_color': (200, 200, 200, 150),  # BGRA (light gray with transparency)
            'metadata_text_color': (0, 0, 0),  # BGR (black)
            'metadata_font_scale': 0.7,
            'metadata_font_thickness': 1,
            'metadata_font_face': cv2.FONT_HERSHEY_SIMPLEX,
            'metadata_box_padding': 10,
            'metadata_pattern': r'(.+)_CC_(\d+\.?\d*)-(\d+\.?\d*)m',  # Pattern to extract metadata from filename
            'max_width': 2000,  # Maximum width for output image
            'min_width': 800,   # Minimum width for output image
            'box_alpha': 0.7,    # Transparency of metadata box (0-1)
            'additional_columns': [],  # Additional columns from CSV to include
            'save_individual_compartments': True  # Save un-stitched images
        }
        
        # Update with provided config if any
        if config:
            self.config.update(config)
            
        # Logger
        self.logger = logging.getLogger(__name__)

    def process_selected_holes(self, 
                        compartment_dir: str,
                        csv_path: Optional[str] = None,
                        selected_columns: Optional[List[str]] = None,
                        hole_ids: Optional[List[str]] = None) -> List[str]:
        """
        Process specific holes to create trace images.
        
        Args:
            compartment_dir: Directory containing compartment images
            csv_path: Optional path to CSV file with additional data
            selected_columns: Optional list of columns to include from CSV
            hole_ids: Optional list of specific hole IDs to process
            
        Returns:
            List of paths to generated trace images
        """
        # Load CSV data
        csv_data = None
        if csv_path and os.path.exists(csv_path):
            try:
                # Load the CSV
                csv_data = pd.read_csv(csv_path)
                
                # Convert column names to lowercase for case-insensitive matching
                csv_data.columns = [col.lower() for col in csv_data.columns]
                
                # Ensure numeric columns are properly typed
                for col in csv_data.columns:
                    if col not in ['holeid', 'cutoffs1', 'cutoffs2']:  # Don't convert text columns
                        try:
                            csv_data[col] = pd.to_numeric(csv_data[col], errors='coerce')
                        except:
                            self.logger.warning(f"Could not convert column {col} to numeric")
                
                self.logger.info(f"Loaded CSV with {len(csv_data)} rows and {len(csv_data.columns)} columns")
                if selected_columns:
                    self.logger.info(f"Selected columns: {selected_columns}")
                    
            except Exception as e:
                self.logger.error(f"Error loading CSV data: {str(e)}")
                if self.progress_queue:
                    self.progress_queue.put((f"Error loading CSV data: {str(e)}", None))
                    
        # Get the output directory from FileManager
        output_dir = self.file_manager.dir_structure["drill_traces"] if self.file_manager else None
        
        if not output_dir:
            self.logger.error("No output directory available")
            return []
        
        # Process each hole separately based on subdirectories
        generated_traces = []
        
        # Process only specified holes if provided
        for i, hole_id in enumerate(hole_ids or []):
            try:
                # Update progress
                if self.progress_queue:
                    progress = ((i + 1) / len(hole_ids)) * 100
                    self.progress_queue.put((f"Processing hole {i+1}/{len(hole_ids)}: {hole_id}", progress))
                
                # Path to hole directory inside compartment_dir
                hole_dir = os.path.join(compartment_dir, hole_id)
                
                if not os.path.isdir(hole_dir):
                    self.logger.warning(f"Directory not found for hole {hole_id}")
                    continue
                    
                # Collect compartment images for this hole
                compartments = []
                
                # Look for compartment images
                image_files = [f for f in os.listdir(hole_dir) 
                            if os.path.isfile(os.path.join(hole_dir, f)) and 
                            f.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff'))]
                
                # Process each compartment image
                for filename in image_files:
                    # Parse metadata from filename
                    hole_id_from_file, depth_from, depth_to = self.parse_filename_metadata(filename)
                    
                    if hole_id_from_file and depth_from is not None and depth_to is not None:
                        # Check if this belongs to the current hole
                        if hole_id_from_file.upper() == hole_id.upper():
                            file_path = os.path.join(hole_dir, filename)
                            compartments.append((hole_id, depth_from, depth_to, file_path))
                
                # Sort compartments by depth
                compartments = sorted(compartments, key=lambda x: x[1])
                
                if not compartments:
                    self.logger.warning(f"No valid compartment images found for hole {hole_id}")
                    continue
                    
                # Generate trace
                trace_path = self.generate_drillhole_trace_cv2(
                    hole_id, compartments, csv_data, output_dir
                )
                
                if trace_path:
                    generated_traces.append(trace_path)
                    self.logger.info(f"Generated trace for hole {hole_id}")
                else:
                    self.logger.warning(f"Failed to generate trace for hole {hole_id}")
                    
            except Exception as e:
                self.logger.error(f"Error processing hole {hole_id}: {str(e)}")
                self.logger.error(traceback.format_exc())
        
        # Final status update
        status_msg = f"Completed trace generation: {len(generated_traces)}/{len(hole_ids) if hole_ids else 0} successful"
        self.logger.info(status_msg)
        if self.progress_queue:
            self.progress_queue.put((status_msg, 100))
        
        return generated_traces

    def get_csv_columns(self, csv_path: str) -> List[str]:
        """
        Get column names from a CSV file.
        
        Args:
            csv_path: Path to CSV file
            
        Returns:
            List of column names
        """
        try:
            # Read just the header row for efficiency
            df = pd.read_csv(csv_path, nrows=0)
            columns = df.columns.tolist()
            
            # Check for required columns
            required_columns = ['holeid', 'from', 'to']
            missing_columns = [col for col in required_columns if col.lower() not in [c.lower() for c in columns]]
            
            if missing_columns:
                msg = f"CSV missing required columns: {', '.join(missing_columns)}"
                self.logger.error(msg)
                if self.progress_queue:
                    self.progress_queue.put((msg, None))
                return []
            
            return columns
        except Exception as e:
            self.logger.error(f"Error reading CSV columns: {str(e)}")
            if self.progress_queue:
                self.progress_queue.put((f"Error reading CSV: {str(e)}", None))
            return []

    def select_csv_columns(self, columns: List[str]) -> List[str]:
        """
        Open a dialog to let the user select columns from a CSV.
        
        Args:
            columns: List of column names from the CSV
            
        Returns:
            List of selected column names
        """
        if not self.root:
            self.logger.warning("No Tkinter root available for column selection dialog")
            return []
            
        # Create a dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Select CSV Columns")
        dialog.geometry("500x400")
        dialog.grab_set()  # Make dialog modal
        
        # Explanatory text
        header_label = ttk.Label(
            dialog, 
            text="Select additional columns to display in the metadata box (max 5):",
            wraplength=480,
            justify=tk.LEFT,
            padding=(10, 10)
        )
        header_label.pack(fill=tk.X)
        
        # Required columns notice
        required_label = ttk.Label(
            dialog, 
            text="Note: 'holeid', 'from', and 'to' are always included.",
            font=("Arial", 9, "italic"),
            foreground="gray",
            padding=(10, 0, 10, 10)
        )
        required_label.pack(fill=tk.X)
        
        # Create a frame for the column checkboxes
        column_frame = ttk.Frame(dialog, padding=10)
        column_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create a canvas with scrollbar for many columns
        canvas = tk.Canvas(column_frame)
        scrollbar = ttk.Scrollbar(column_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Dictionary to track which columns are selected
        selected_columns = {}
        
        # Exclude required columns from the selection list
        required_columns = ['holeid', 'from', 'to']
        selectable_columns = [col for col in columns if col.lower() not in [c.lower() for c in required_columns]]
        
        # Add checkboxes for each column
        for column in selectable_columns:
            var = tk.BooleanVar(value=False)
            selected_columns[column] = var
            
            checkbox = ttk.Checkbutton(
                scrollable_frame,
                text=column,
                variable=var,
                command=lambda: self._update_selection_count(selected_columns, selection_label)
            )
            checkbox.pack(anchor="w", padx=10, pady=5)
        
        # Label to show how many columns are selected
        selection_label = ttk.Label(
            dialog,
            text="0 columns selected (max 5)",
            padding=(10, 10)
        )
        selection_label.pack()
        
        # Buttons frame
        button_frame = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        button_frame.pack(fill=tk.X)
        
        # Result container
        result = []
        
        # OK button handler
        def on_ok():
            nonlocal result
            result = [col for col, var in selected_columns.items() if var.get()]
            if len(result) > 5:
                messagebox.showwarning("Too Many Columns", "Please select at most 5 columns.")
                return
            dialog.destroy()
        
        # Cancel button handler
        def on_cancel():
            nonlocal result
            result = []
            dialog.destroy()
        
        # Add buttons
        ok_button = ttk.Button(button_frame, text="OK", command=on_ok)
        ok_button.pack(side=tk.RIGHT, padx=(5, 0))
        
        cancel_button = ttk.Button(button_frame, text="Cancel", command=on_cancel)
        cancel_button.pack(side=tk.RIGHT)
        
        # Wait for dialog to close
        dialog.wait_window()
        return result
        
    def _update_selection_count(self, selected_columns, label_widget):
        """Helper method to update the selection count label."""
        count = sum(var.get() for var in selected_columns.values())
        color = "red" if count > 5 else "black"
        label_widget.config(text=f"{count} columns selected (max 5)", foreground=color)

    def load_csv_data(self, csv_path: str, selected_columns: List[str]) -> pd.DataFrame:
        """
        Load CSV data for metadata integration.
        
        Args:
            csv_path: Path to CSV file
            selected_columns: Columns to include
            
        Returns:
            DataFrame with the CSV data
        """
        try:
            # Determine which columns to load - ensure required columns are included
            required_columns = ['holeid', 'from', 'to']
            columns_to_load = list(set(required_columns + selected_columns))
            
            # Load the CSV
            df = pd.read_csv(csv_path, usecols=columns_to_load)
            
            # Normalize column names to lowercase for case-insensitive matching
            df.columns = [col.lower() for col in df.columns]
            
            # Ensure numeric columns are properly typed
            if 'from' in df.columns:
                df['from'] = pd.to_numeric(df['from'], errors='coerce')
            if 'to' in df.columns:
                df['to'] = pd.to_numeric(df['to'], errors='coerce')
                
            # Drop rows with missing required values
            df = df.dropna(subset=['holeid', 'from', 'to'])
            
            self.logger.info(f"Loaded CSV with {len(df)} rows and {len(df.columns)} columns")
            return df
            
        except Exception as e:
            self.logger.error(f"Error loading CSV data: {str(e)}")
            if self.progress_queue:
                self.progress_queue.put((f"Error loading CSV data: {str(e)}", None))
            return pd.DataFrame()

    def get_csv_data_for_interval(self, 
                                df: pd.DataFrame, 
                                hole_id: str, 
                                depth_from: float, 
                                depth_to: float) -> Dict[str, Any]:
        """
        Get relevant CSV data for a specific interval.
        
        Args:
            df: DataFrame with CSV data
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            
        Returns:
            Dictionary with column values for the interval or empty dict if no match
        """
        if df.empty:
            return {}
            
        try:
            # Normalize hole ID for matching
            hole_id_norm = hole_id.upper()
            
            # Filter for matching hole ID
            hole_data = df[df['holeid'].str.upper() == hole_id_norm]
            
            if hole_data.empty:
                return {}
                
            # Find intervals that overlap with our target interval
            matching_intervals = hole_data[
                ((hole_data['from'] <= depth_from) & (hole_data['to'] > depth_from)) |
                ((hole_data['from'] >= depth_from) & (hole_data['from'] < depth_to)) |
                ((hole_data['from'] <= depth_from) & (hole_data['to'] >= depth_to))
            ].copy()  # Use .copy() to avoid the SettingWithCopyWarning
            
            if matching_intervals.empty:
                # Return empty dict - no approximation
                self.logger.info(f"No matching interval found for {hole_id} {depth_from}-{depth_to}m")
                return {}
                
            # If multiple matches, take the one with the highest overlap
            if len(matching_intervals) > 1:
                # Calculate overlap for each interval
                def calculate_overlap(row):
                    overlap_start = max(row['from'], depth_from)
                    overlap_end = min(row['to'], depth_to)
                    return max(0, overlap_end - overlap_start)
                    
                matching_intervals.loc[:, 'overlap'] = matching_intervals.apply(calculate_overlap, axis=1)
                best_match = matching_intervals.loc[matching_intervals['overlap'].idxmax()]
                
                # Convert the Series to a dictionary
                result = best_match.to_dict()
                
                # Clean up the result
                if 'overlap' in result:
                    del result['overlap']
                    
                return result
            else:
                # Just one match, convert to dictionary
                return matching_intervals.iloc[0].to_dict()
                
        except Exception as e:
            self.logger.error(f"Error getting CSV data for interval: {str(e)}")
            return {}

            
    def rotate_image(self, image: np.ndarray) -> np.ndarray:
        """
        Rotate an image 90 degrees clockwise.
        
        Args:
            image: Input image
            
        Returns:
            Rotated image
        """
        # Rotate 90 degrees clockwise
        return cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)

    
    def parse_filename_metadata(self, filename: str) -> Tuple[Optional[str], Optional[float], Optional[float]]:
        """
        Parse metadata from a filename based on various patterns including leading zeros.
        
        Args:
            filename: Filename to parse
            
        Returns:
            Tuple of (hole_id, depth_from, depth_to)
        """
        try:
            # Get compartment interval from config
            compartment_interval = self.config.get('compartment_interval', 1.0)
            
            # Try the pattern with any number of digits: HoleID_CC_EndDepth
            match = re.search(r'([A-Za-z]{2}\d{4})_CC_(\d{1,3})(?:\..*)?$', filename)
            if match:
                hole_id = match.group(1)
                depth_to = float(match.group(2))
                # Use the configured interval
                depth_from = depth_to - compartment_interval
                self.logger.info(f"Parsed from CC format: {hole_id}, {depth_from}-{depth_to}m (interval: {compartment_interval}m)")
                return hole_id, depth_from, depth_to
            
            # Try with 3-digit format explicitly
            match = re.search(r'([A-Za-z]{2}\d{4})_CC_(\d{3})(?:\..*)?$', filename)
            if match:
                hole_id = match.group(1)
                depth_to = float(match.group(2))
                # Use the configured interval
                depth_from = depth_to - compartment_interval
                self.logger.info(f"Parsed from 3-digit format: {hole_id}, {depth_from}-{depth_to}m (interval: {compartment_interval}m)")
                return hole_id, depth_from, depth_to
            
            # Try the old pattern as fallback
            match = re.match(self.config['metadata_pattern'], filename)
            if match:
                hole_id = match.group(1)
                depth_from = float(match.group(2))
                depth_to = float(match.group(3))
                return hole_id, depth_from, depth_to
            
            # Try another generic pattern if needed
            match = re.search(r'([A-Za-z]{2}\d{4}).*?(\d+\.?\d*)-(\d+\.?\d*)m', filename)
            if match:
                hole_id = match.group(1)
                depth_from = float(match.group(2))
                depth_to = float(match.group(3))
                return hole_id, depth_from, depth_to
                
            # No match found
            self.logger.warning(f"Could not parse metadata from filename: {filename}")
            return None, None, None
            
        except Exception as e:
            self.logger.error(f"Error parsing filename metadata: {str(e)}")
            return None, None, None

    def add_metadata_to_image(self, 
                        image: np.ndarray, 
                        hole_id: str, 
                        depth_from: float, 
                        depth_to: float,
                        csv_data: Optional[Dict[str, Any]] = None,
                        is_missing: bool = False) -> np.ndarray:
        """
        Add metadata box that overlaps with the left side of the image.
        
        Args:
            image: Input image
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            csv_data: Optional CSV data to include
            is_missing: Whether this interval is missing (a gap)
            
        Returns:
            Image with metadata box overlaid on the left side
        """
        # Get image dimensions
        h, w = image.shape[:2]
        
        # Make a copy of the image to overlay metadata on
        result_image = image.copy()
        
        # Determine metadata box width (25% of image width)
        metadata_width = int(w * 0.25)
        
        # Create semi-transparent background for metadata
        overlay = result_image.copy()
        
        # Create metadata section with semi-transparent background
        if is_missing:
            # Dark red background for missing intervals with alpha
            background_color = (50, 50, 150)  # BGR dark red
        else:
            # Gray background for normal intervals with alpha
            background_color = (200, 200, 200)  # BGR gray
        
        # Fill the metadata section with the background color
        cv2.rectangle(
            overlay,
            (0, 0),
            (metadata_width, h),
            background_color,
            -1
        )
        
        # Add a separator line between metadata and rest of image
        line_color = (100, 100, 100)  # Dark gray
        line_thickness = 2
        
        cv2.line(
            overlay,
            (metadata_width, 0),
            (metadata_width, h),
            line_color,
            line_thickness
        )
        
        # Apply the overlay with transparency
        alpha = 0.8  # 80% opacity
        cv2.addWeighted(overlay, alpha, result_image, 1 - alpha, 0, result_image)
        
        # Add text to metadata section
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.5  # Smaller font
        font_thickness = 1
        padding = 10  # Padding from left edge
        
        # Use white text for missing intervals, black for normal
        if is_missing:
            text_color = (255, 255, 255)  # White for contrast with dark red
        else:
            text_color = (0, 0, 0)  # Black for normal intervals
        
        # Format depth as integers rather than floats
        depth_from_int = int(depth_from)
        depth_to_int = int(depth_to)
        
        # Create the title line (combined hole ID and interval with bold font)
        title_text = f"{hole_id}_{depth_from_int}-{depth_to_int}m"
        
        # Prepare the list of text lines with appropriate spacing
        text_lines = []
        
        # Add indicator for missing interval if needed
        if is_missing:
            text_lines.append(("** MISSING INTERVAL **", True))  # (text, is_bold)
        
        # Add the title as first line (or second if missing interval)
        text_lines.append((title_text, True))  # Bold for title
        
        # Add blank line after title
        text_lines.append(("", False))
        
        # Add CSV data if available
        if csv_data and isinstance(csv_data, dict):
            # Skip holeid, from, and to as they're already included above
            skip_columns = ['holeid', 'from', 'to']
            
            # Process each column
            for key, value in csv_data.items():
                if key.lower() not in skip_columns:
                    # Format numeric values appropriately
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        if value == int(value):  # It's a whole number
                            formatted_value = f"{int(value)}"
                        else:
                            formatted_value = f"{value:.2f}"
                    elif pd.isna(value):
                        formatted_value = "N/A"
                    else:
                        formatted_value = str(value)
                    
                    # Add to text lines with proper capitalization
                    display_key = key.replace('_', ' ').capitalize() 
                    text_lines.append((f"{display_key}: {formatted_value}", False))
        
        # Draw each line of text
        y_offset = 30  # Start a bit down from the top
        line_spacing = 20  # Increased spacing between lines
        
        for text, is_bold in text_lines:
            if not text:  # Skip empty lines, just add space
                y_offset += 10
                continue
            
            # Handle text that's too long for the metadata box by wrapping
            thickness = font_thickness + 1 if is_bold else font_thickness
            available_width = metadata_width - (padding * 2)
            
            # Calculate if this text needs wrapping
            (text_width, text_height), _ = cv2.getTextSize(text, font, font_scale, thickness)
            
            if text_width > available_width:
                # Text needs wrapping - split into words
                words = text.split()
                line = ""
                for word in words:
                    test_line = line + word + " "
                    (test_width, _), _ = cv2.getTextSize(test_line, font, font_scale, thickness)
                    
                    if test_width <= available_width:
                        line = test_line
                    else:
                        # Draw current line and start a new one
                        cv2.putText(
                            result_image,
                            line.strip(),
                            (padding, y_offset),
                            font,
                            font_scale,
                            text_color,
                            thickness
                        )
                        y_offset += line_spacing
                        line = word + " "
                
                # Draw the last line
                if line:
                    cv2.putText(
                        result_image,
                        line.strip(),
                        (padding, y_offset),
                        font,
                        font_scale,
                        text_color,
                        thickness
                    )
            else:
                # Text fits on one line
                cv2.putText(
                    result_image,
                    text,
                    (padding, y_offset),
                    font,
                    font_scale,
                    text_color,
                    thickness
                )
            
            # Move to next line
            y_offset += line_spacing
        
        return result_image
    
    
    def collect_compartment_images(self, compartment_dir: str) -> Dict[str, List[Tuple[str, float, float, str]]]:
        """
        Collect and organize compartment images by hole ID.
        
        Args:
            compartment_dir: Directory containing compartment images
            
        Returns:
            Dictionary mapping hole IDs to lists of (hole_id, depth_from, depth_to, file_path) tuples
        """
        # Dictionary to store compartment info by hole ID
        hole_compartments: Dict[str, List[Tuple[str, float, float, str]]] = {}
        
        # Valid image extensions
        valid_extensions = ('.jpg', '.jpeg', '.png', '.tif', '.tiff')
        
        try:
            # Get all image files
            image_files = [f for f in os.listdir(compartment_dir) 
                        if os.path.isfile(os.path.join(compartment_dir, f)) and 
                        f.lower().endswith(valid_extensions)]
            
            if not image_files:
                self.logger.warning(f"No image files found in {compartment_dir}")
                return hole_compartments
                
            self.logger.info(f"Found {len(image_files)} image files to process")
            
            # Process each file
            for filename in image_files:
                # Parse metadata from filename
                hole_id, depth_from, depth_to = self.parse_filename_metadata(filename)
                
                if hole_id and depth_from is not None and depth_to is not None:
                    # Add to appropriate hole ID list
                    if hole_id not in hole_compartments:
                        hole_compartments[hole_id] = []
                        
                    file_path = os.path.join(compartment_dir, filename)
                    hole_compartments[hole_id].append((hole_id, depth_from, depth_to, file_path))
            
            # Sort each hole's compartments by depth
            for hole_id, compartments in hole_compartments.items():
                hole_compartments[hole_id] = sorted(compartments, key=lambda x: x[1])  # Sort by depth_from
                
            self.logger.info(f"Organized compartments for {len(hole_compartments)} holes")
            
            # Log some statistics
            for hole_id, compartments in hole_compartments.items():
                self.logger.info(f"Hole {hole_id}: {len(compartments)} compartments")
                
            return hole_compartments
            
        except Exception as e:
            self.logger.error(f"Error collecting compartment images: {str(e)}")
            if self.progress_queue:
                self.progress_queue.put((f"Error collecting images: {str(e)}", None))
            return {}
 
    def generate_compartment_with_data(self,
                                    image: np.ndarray,
                                    hole_id: str,
                                    depth_from: float,
                                    depth_to: float,
                                    fe_value: Optional[float] = None,
                                    sio2_value: Optional[float] = None,
                                    al2o3_value: Optional[float] = None,
                                    p_value: Optional[float] = None,
                                    cutoffs1_value: Optional[str] = None,
                                    cutoffs2_value: Optional[str] = None,
                                    is_missing: bool = False,
                                    data_column_width: int = 100) -> np.ndarray:
        # Get the image dimensions
        h, w = image.shape[:2]
        
        # Use the provided fixed width for data columns
        column_width = data_column_width
        
        # Total width of all data columns
        data_width = column_width * 6
        
        # Total width for the new image: original width + data columns
        total_width = w + data_width
        
        # Create new image with exact original dimensions plus data columns
        if len(image.shape) == 3:
            if image.shape[2] == 4:  # BGRA
                composite = np.ones((h, total_width, 4), dtype=np.uint8) * 255
                # Set alpha channel to fully opaque
                composite[:, :, 3] = 255
            else:  # BGR
                composite = np.ones((h, total_width, 3), dtype=np.uint8) * 255
        else:  # Grayscale
            composite = np.ones((h, total_width), dtype=np.uint8) * 255
        
        # Copy the image to the left portion with exact dimensions
        if is_missing:
            # For missing intervals, create a black box with diagonal stripes
            composite[:, :w] = 0  # Black box
            
            # Add diagonal stripes
            for i in range(0, h + w, 20):
                cv2.line(
                    composite[:, :w],
                    (0, i),
                    (i, 0),
                    (50, 50, 150),  # Dark red stripes
                    2
                )
            
            # Add MISSING text
            font = cv2.FONT_HERSHEY_SIMPLEX
            text_size = cv2.getTextSize("MISSING", font, 1.0, 2)[0]
            text_x = (w - text_size[0]) // 2
            text_y = h // 2
            
            cv2.putText(
                composite,
                "MISSING",
                (text_x, text_y),
                font,
                1.0,
                (255, 255, 255),  # White text
                2
            )
        else:
            # Normal image - copy the full original image 
            composite[:, :w] = image
        
        # Calculate x positions for data columns - each starts after the original image
        x_positions = [
            w,                      # Fe
            w + column_width,       # SiO2
            w + column_width * 2,   # Al2O3
            w + column_width * 3,   # P
            w + column_width * 4,   # Cutoffs1
            w + column_width * 5    # Cutoffs2
        ]

        # Add column headers
        header_bg = (220, 220, 220)  # Light gray
        header_height = 30
        
        # Draw header backgrounds and add header text
        headers = ["Fe %", "SiO2 %", "Al2O3 %", "P %", "Cutoffs1", "Cutoffs2"]
        font = cv2.FONT_HERSHEY_SIMPLEX  # Define font for later text
        
        for i, header in enumerate(headers):
            x = x_positions[i]
            width = column_width
            
            # Draw header background
            cv2.rectangle(composite, (x, 0), (x + width, header_height), header_bg, -1)
            
            # Add header text
            cv2.putText(composite, header, (x + 10, 20), font, 0.6, (0, 0, 0), 1)

        # Function to draw column
        def draw_column(x_pos, width, value, color_func, format_str="{:.2f}"):
            # If value is None or NaN, keep column white
            if value is None or pd.isna(value):
                return
                
            # Get color for the value
            color = color_func(value)
            
            # Draw the colored column
            cv2.rectangle(
                composite,
                (x_pos, header_height),
                (x_pos + width, h),
                color,
                -1  # Fill
            )
            
            # Add value text if not None
            if value is not None and not pd.isna(value):
                # Black or white text depending on background
                text_color = (255, 255, 255) if sum(color) < 380 else (0, 0, 0)
                
                # Format the value
                if isinstance(value, (int, float)):
                    text = format_str.format(value)
                else:
                    text = str(value)
                    
                # Get text size
                text_size = cv2.getTextSize(text, font, 0.8, 2)[0]
                text_x = x_pos + (width - text_size[0]) // 2
                text_y = h // 2
                
                cv2.putText(
                    composite,
                    text,
                    (text_x, text_y),
                    font,
                    0.8,
                    text_color,
                    2
                )
        
        # Define color functions for numeric data
        def get_fe_color(value):
            # Fe color scale (modified from matplotlib version)
            if value is None or np.isnan(value) or value < 30:
                return (224, 224, 224)  # Light gray (BGRs)
            elif value < 40:
                return (48, 48, 48)      # Dark gray
            elif value < 45:
                return (255, 0, 0)       # Blue (BGR)
            elif value < 50:
                return (255, 255, 0)     # Cyan
            elif value < 54:
                return (0, 255, 0)       # Green  
            elif value < 56:
                return (0, 255, 255)     # Yellow
            elif value < 58:
                return (0, 128, 255)     # Orange
            elif value < 60:
                return (0, 0, 255)       # Red
            else:
                return (255, 0, 255)     # Magenta
        
        def get_sio2_color(value):
            # SiO2 color scale (modified from matplotlib version)
            if value is None or np.isnan(value) or value < 5:
                return (255, 255, 255)  # White (BGR)
            elif value < 15:
                return (0, 255, 0)      # Green
            elif value < 35:
                return (255, 0, 0)      # Blue
            else:
                return (0, 0, 255)      # Red

        def get_al2o3_color(value):
            # Al2O3 scale based on your Image 1
            if value is None or np.isnan(value):
                return (200, 200, 200)  # Light gray for no value
            elif value < 2:
                return (255, 0, 255)    # Magenta (0)
            elif value < 5:
                return (0, 0, 255)      # Red (2)
            elif value < 10:
                return (0, 165, 255)    # Orange (5) 
            elif value < 15:
                return (0, 255, 255)    # Yellow (10)
            else:
                return (0, 0, 255)      # Blue (15+)

        def get_p_color(value):
            # P scale based on your Image 2 (3 gradient scheme)
            if value is None or np.isnan(value):
                return (200, 200, 200)  # Light gray for no value
            elif value < 0.02:
                return (255, 0, 255)    # Magenta
            elif value < 0.08:
                return (0, 255, 0)      # Green
            else:
                return (0, 0, 255)      # Red
        
        def get_cutoffs_color(value):
            # Color mapping for Cutoffs text
            cutoffs_colors = {
                'Other': (31, 161, 217),    # #D9A11F
                'BID/Fs?': (0, 255, 255),   # #FFFF00
                'BIFf': (255, 77, 77),      # #4DFFFF
                'BIFf?': (0, 165, 255),     # #FFA500
                'BIFhm': (0, 0, 255),       # #FF0101
                'Mineralised': (200, 0, 254),# #FE00C4
                'High Confidence': (0, 255, 0),  # Green
                'Potential BID/Fs': (255, 255, 0)  # Cyan
            }
            
            # If value is None or empty, return default color
            if value is None or pd.isna(value) or str(value).strip() == "":
                return (200, 200, 200)  # Light gray for no value
                
            # Convert value to string just in case it's a number
            value_str = str(value).strip()
            
            # Check if value matches any of our known labels
            return cutoffs_colors.get(value_str, (200, 200, 200))  # Default to light gray

        # Draw each data column
        draw_column(x_positions[0], column_width, fe_value, get_fe_color)
        draw_column(x_positions[1], column_width, sio2_value, get_sio2_color)
        draw_column(x_positions[2], column_width, al2o3_value, get_al2o3_color)
        draw_column(x_positions[3], column_width, p_value, get_p_color, format_str="{:.3f}")
        
        # Draw cutoffs columns with text
        if cutoffs1_value and not pd.isna(cutoffs1_value) and str(cutoffs1_value).lower() != 'nan':
            x_pos = x_positions[4]
            width = column_width
            color = get_cutoffs_color(cutoffs1_value)
            
            # Draw colored background
            cv2.rectangle(composite, (x_pos, header_height), (x_pos + width, h), color, -1)
            
            # Add text
            text_color = (255, 255, 255) if sum(color) < 380 else (0, 0, 0)
            
            # Word wrap for longer labels
            text = str(cutoffs1_value)
            if len(text) > 10:
                # Split into multiple lines
                words = text.split()
                lines = []
                current_line = ""
                
                for word in words:
                    test_line = current_line + " " + word if current_line else word
                    if len(test_line) <= 10:
                        current_line = test_line
                    else:
                        lines.append(current_line)
                        current_line = word
                
                if current_line:
                    lines.append(current_line)
                    
                # Draw each line
                for i, line in enumerate(lines):
                    y_pos = h // 2 - (len(lines) - 1) * 15 + i * 30
                    cv2.putText(
                        composite,
                        line,
                        (x_pos + width//2 - 25, y_pos),
                        font,
                        0.7,
                        text_color,
                        2
                    )
            else:
                # Single line text
                cv2.putText(
                    composite,
                    text,
                    (x_pos + width//2 - 25, h//2),
                    font,
                    0.8,
                    text_color,
                    2
                )
        
        # Same for cutoffs2
        if cutoffs2_value and not pd.isna(cutoffs2_value) and str(cutoffs2_value).lower() != 'nan':
            x_pos = x_positions[5]
            width = column_width
            color = get_cutoffs_color(cutoffs2_value)
            
            # Draw colored background
            cv2.rectangle(composite, (x_pos, header_height), (x_pos + width, h), color, -1)
            
            # Add text
            text_color = (255, 255, 255) if sum(color) < 380 else (0, 0, 0)
            
            # Word wrap for longer labels
            text = str(cutoffs2_value)
            if len(text) > 10:
                # Split into multiple lines
                words = text.split()
                lines = []
                current_line = ""
                
                for word in words:
                    test_line = current_line + " " + word if current_line else word
                    if len(test_line) <= 10:
                        current_line = test_line
                    else:
                        lines.append(current_line)
                        current_line = word
                
                if current_line:
                    lines.append(current_line)
                    
                # Draw each line
                for i, line in enumerate(lines):
                    y_pos = h // 2 - (len(lines) - 1) * 15 + i * 30
                    cv2.putText(
                        composite,
                        line,
                        (x_pos + width//2 - 25, y_pos),
                        font,
                        0.7,
                        text_color,
                        2
                    )
            else:
                # Single line text
                cv2.putText(
                    composite,
                    text,
                    (x_pos + width//2 - 25, h//2),
                    font,
                    0.8,
                    text_color,
                    2
                )
        
        # Add depth information at the top of the compartment in HoleID - From-To format
        depth_text = f"{hole_id} - {int(depth_from)}-{int(depth_to)}m"
        text_size = cv2.getTextSize(depth_text, font, 0.7, 2)[0]  # Bold text (thickness=2)

        # Add a dark background for better visibility
        bg_margin = 5
        cv2.rectangle(
            composite,
            (10 - bg_margin, 30 - text_size[1] - bg_margin),
            (10 + text_size[0] + bg_margin, 30 + bg_margin),
            (0, 0, 0),  # Black background
            -1  # Fill
        )

        # Add white bold text
        cv2.putText(
            composite,
            depth_text,
            (10, 30),
            font,
            0.7,
            (255, 255, 255),  # White text
            2  # Bold
        )
        
        # Add separator lines between data columns
        for x in x_positions[1:]:
            cv2.line(composite, (x, 0), (x, h), (0, 0, 0), 1)
        
        return composite

        
    def create_value_legend(self, width: int = 800, height: int = 300) -> np.ndarray:
        """
        Create a comprehensive legend image for all data columns.
        
        Args:
            width: Width of the legend image
            height: Height of the legend image
            
        Returns:
            Legend image as numpy array
        """
        # Create white image for legend
        legend = np.ones((height, width, 3), dtype=np.uint8) * 255
        
        # Define color scales and boundaries for each column
        legend_configs = [
            {
                'title': 'Fe %',
                'colors': [
                    (224, 224, 224),  # Light gray (BGR)
                    (48, 48, 48),     # Dark gray
                    (255, 0, 0),      # Blue
                    (255, 255, 0),    # Cyan
                    (0, 255, 0),      # Green
                    (0, 255, 255),    # Yellow
                    (0, 128, 255),    # Orange
                    (0, 0, 255),      # Red
                    (255, 0, 255)     # Magenta
                ],
                'bounds': ["<30", "30-40", "40-45", "45-50", "50-54", "54-56", "56-58", "58-60", ">60"]
            },
            {
                'title': 'SiO2 %',
                'colors': [
                    (255, 255, 255),  # White
                    (0, 255, 0),      # Green
                    (255, 0, 0),      # Blue
                    (0, 0, 255)       # Red
                ],
                'bounds': ["<5", "5-15", "15-35", ">35"]
            },
            {
                'title': 'Al2O3 %',
                'colors': [
                    (255, 255, 255),  # White
                    (200, 200, 255),  # Light blue
                    (120, 120, 255),  # Medium blue
                    (50, 50, 255),    # Dark blue
                    (0, 0, 255)       # Darkest blue
                ],
                'bounds': ["<2", "2-4", "4-6", "6-8", ">8"]
            },
            {
                'title': 'P %',
                'colors': [
                    (255, 255, 255),  # White
                    (200, 255, 200),  # Light green
                    (100, 255, 100),  # Medium green
                    (0, 255, 0),      # Green
                    (0, 200, 0)       # Dark green
                ],
                'bounds': ["<0.1", "0.1-0.3", "0.3-0.5", "0.5-0.7", ">0.7"]
            },
            {
                'title': 'Cutoffs',
                'colors': [
                    (31, 161, 217),   # Other - #D9A11F
                    (0, 255, 255),    # BID/Fs? - #FFFF00
                    (255, 77, 77),    # BIFf - #4DFFFF
                    (0, 165, 255),    # BIFf? - #FFA500
                    (0, 0, 255),      # BIFhm - #FF0101
                    (200, 0, 254)     # Mineralised - #FE00C4
                ],
                'bounds': [
                    "Other", 
                    "BID/Fs?", 
                    "BIFf", 
                    "BIFf?", 
                    "BIFhm", 
                    "Mineralised"
                ]
            }
        ]
        
        # Draw legend
        font = cv2.FONT_HERSHEY_SIMPLEX
        
        # Title
        cv2.putText(legend, "Geochemical Data Legend", (20, 30), font, 1.0, (0, 0, 0), 2)
        
        # Positioning
        box_width = 50
        box_height = 25
        x_start = 20
        y_pos = 50
        
        # Render each column's legend
        for config in legend_configs:
            # Column title
            cv2.putText(legend, config['title'], (x_start, y_pos), font, 0.7, (0, 0, 0), 2)
            y_pos += 30
            
            # Render color boxes and labels
            for i in range(len(config['colors'])):
                # Draw color box
                cv2.rectangle(
                    legend,
                    (x_start, y_pos),
                    (x_start + box_width, y_pos + box_height),
                    config['colors'][i],
                    -1
                )
                
                # Add border
                cv2.rectangle(
                    legend,
                    (x_start, y_pos),
                    (x_start + box_width, y_pos + box_height),
                    (0, 0, 0),
                    1
                )
                
                # Add label
                cv2.putText(
                    legend,
                    config['bounds'][i],
                    (x_start + box_width + 10, y_pos + 20),
                    font,
                    0.5,
                    (0, 0, 0),
                    1
                )
                
                # Move to next position
                y_pos += box_height + 5
            
            # Add spacing between columns
            y_pos += 20
            
            # Move to next column if running out of vertical space
            if y_pos > height - 100:
                x_start += box_width + 150
                y_pos = 50
        
        return legend

        
    def generate_drillhole_trace_cv2(self, 
                                hole_id: str, 
                                compartments: List[Tuple[str, float, float, str]],
                                csv_data: Optional[pd.DataFrame] = None,
                                output_dir: Optional[str] = None) -> Optional[str]:
        """
        Generate a drillhole trace using OpenCV instead of matplotlib.
        This approach preserves image quality and creates a more accurate representation.
        
        Args:
            hole_id: Hole ID
            compartments: List of (hole_id, depth_from, depth_to, file_path) tuples
            csv_data: Optional DataFrame with CSV data
            output_dir: Optional directory (uses FileManager's directory if None)
            
        Returns:
            Path to the generated image file, or None if failed
        """
        if not compartments:
            self.logger.warning(f"No compartments provided for hole {hole_id}")
            return None
            
        try:
            # Use output_dir parameter if provided, otherwise use FileManager
            if output_dir is None and hasattr(self, 'file_manager') and self.file_manager is not None:
                output_dir = self.file_manager.dir_structure["drill_traces"]
                
            # Make sure we have a valid output directory
            if not output_dir:
                self.logger.error("No output directory available for saving trace images")
                return None
                
            # Sort compartments by depth
            sorted_compartments = sorted(compartments, key=lambda x: x[1])
            
            # Extract depth range
            min_depth = sorted_compartments[0][1]
            max_depth = sorted_compartments[-1][2]
            
            # Prepare chemical data if available
            hole_csv_data = {}
            if csv_data is not None and not csv_data.empty:
                # Filter for this hole
                hole_data = csv_data[csv_data['holeid'].str.upper() == hole_id.upper()]
                
                # Convert to dictionary format for easier lookup
                for _, row in hole_data.iterrows():
                    depth_from = row.get('from', 0)
                    depth_to = row.get('to', 0)
                    if depth_from is not None and depth_to is not None:
                        # Store interval with all available data
                        interval_data = {
                            'from': depth_from,
                            'to': depth_to
                        }
                        
                        # Add all other columns to the interval data
                        for col in hole_data.columns:
                            if col not in ['holeid', 'from', 'to']:
                                interval_data[col] = row.get(col)
                        
                        # Use midpoint as key for interval
                        interval_key = (depth_from + depth_to) / 2
                        hole_csv_data[interval_key] = interval_data
            
            # Process each compartment image - store results for concatenation
            processed_images = []
            
            # Create a list to track missing intervals
            missing_intervals = []
            processed_depths = []
            
            # Track all depth ranges we have processed
            for _, depth_from, depth_to, _ in sorted_compartments:
                processed_depths.append((depth_from, depth_to))
            
            # Load first image to determine standard width
            standard_width = None
            for _, _, _, file_path in sorted_compartments:
                try:
                    test_img = cv2.imread(file_path)
                    if test_img is not None:
                        # Rotate to get proper dimensions
                        test_rotated = cv2.rotate(test_img, cv2.ROTATE_90_CLOCKWISE)
                        h, w = test_rotated.shape[:2]
                        
                        # Calculate data column widths - fixed size
                        data_column_width = 100  # Fixed width for each data column
                        total_data_width = data_column_width * 6  # 6 data columns
                        
                        # Get maximum width we'll need to accommodate both the image and data columns
                        standard_width = w + total_data_width
                        
                        # No need to keep checking - we've got our reference width
                        self.logger.info(f"Using standard width of {standard_width} pixels based on first valid image")
                        break
                except Exception:
                    continue

            # Set default if no images loaded
            if standard_width is None:
                standard_width = 800
                self.logger.warning("No valid images found to determine width, using default width of 800 pixels")
            
            # Find missing intervals by checking gaps between processed depths
            processed_depths.sort()  # Ensure they're in order
            prev_depth = min_depth
            missing_intervals = []  # Initialize an empty list for missing intervals
            
            for depth_from, depth_to in processed_depths:
                if depth_from > prev_depth:
                    # Found a gap
                    missing_intervals.append((prev_depth, depth_from))
                prev_depth = max(prev_depth, depth_to)
            
            # Also check if there's a gap at the end
            if prev_depth < max_depth:
                missing_intervals.append((prev_depth, max_depth))
                
            # Process each compartment
            for i, (hole_id, depth_from, depth_to, file_path) in enumerate(sorted_compartments):
                try:
                    # Load and process image
                    img = cv2.imread(file_path)
                    if img is None:
                        self.logger.warning(f"Couldn't read image: {file_path}")
                        continue
                    
                    # Rotate image 90 degrees clockwise
                    rotated_img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
                    
                    # Get the original image width before any processing
                    original_width = rotated_img.shape[1]
                    
                    # Look up chemical data for this depth interval
                    fe_value = None
                    sio2_value = None
                    al2o3_value = None
                    p_value = None
                    cutoffs1_value = None
                    cutoffs2_value = None

                    
                    # Find best matching interval in CSV data
                    if hole_csv_data:
                        # Calculate midpoint of current interval
                        midpoint = (depth_from + depth_to) / 2
                        
                        # Find closest interval in CSV data
                        closest_interval = None
                        min_distance = float('inf')
                        
                        for interval_mid, interval_data in hole_csv_data.items():
                            interval_from = interval_data['from']
                            interval_to = interval_data['to']
                            
                            # Check if midpoint is within interval
                            if interval_from <= midpoint <= interval_to:
                                closest_interval = interval_data
                                break
                            
                            # Otherwise check if it's the closest interval
                            distance = min(abs(midpoint - interval_from), abs(midpoint - interval_to))
                            if distance < min_distance:
                                min_distance = distance
                                closest_interval = interval_data
                        
                        # Extract chemistry and cutoff1 and cutoffs 2 values if available
                        if closest_interval:
                            fe_value = closest_interval.get('fe_pct')
                            sio2_value = closest_interval.get('sio2_pct')
                            al2o3_value = closest_interval.get('al2o3_pct')
                            p_value = closest_interval.get('p_pct')
                            
                            # Properly handle text columns for cutoffs - don't display 'nan'
                            cutoffs1_value = closest_interval.get('cutoffs1')
                            if pd.isna(cutoffs1_value) or cutoffs1_value == 'nan':
                                cutoffs1_value = None
                                
                            cutoffs2_value = closest_interval.get('cutoffs2')
                            if pd.isna(cutoffs2_value) or cutoffs2_value == 'nan':
                                cutoffs2_value = None
                    
                    # Create composite image with data
                    data_column_width = 100  # Fixed width for data columns
                    composite = self.generate_compartment_with_data(
                        rotated_img, 
                        hole_id, 
                        depth_from, 
                        depth_to, 
                        fe_value=fe_value, 
                        sio2_value=sio2_value, 
                        al2o3_value=al2o3_value, 
                        p_value=p_value,
                        cutoffs1_value=cutoffs1_value, 
                        cutoffs2_value=cutoffs2_value,
                        data_column_width=data_column_width
                    )
                    
                    # Add to processed images
                    processed_images.append(composite)
                    
                except Exception as e:
                    self.logger.error(f"Error processing compartment {file_path}: {str(e)}")
                    self.logger.error(traceback.format_exc())

            # Now process missing intervals - only if there are any
            if missing_intervals:
                self.logger.info(f"Processing {len(missing_intervals)} missing intervals")
                for gap_from, gap_to in missing_intervals:
                    # Log the gap being processed
                    self.logger.info(f"Processing gap from {gap_from}m to {gap_to}m")

                    # For each 1-meter interval in the gap
                    for meter_start in range(int(gap_from), int(gap_to)):
                        # The compartment at meter_start is identified by its end depth: meter_start + 1
                        meter_end = meter_start + 1
                        compartment_depth = meter_end
                        self.logger.info(f"Creating missing compartment for {meter_start}-{meter_end}m (depth {compartment_depth})")
                        
                        # Determine a reasonable height for the missing interval
                        if processed_images:
                            avg_height = sum(img.shape[0] for img in processed_images) // len(processed_images)
                        else:
                            avg_height = 300  # Default height if no other images available
                            
                        # Create black image with appropriate dimensions and type
                        if processed_images and len(processed_images[0].shape) == 3:
                            if processed_images[0].shape[2] == 4:  # With alpha
                                blank_image = np.zeros((avg_height, standard_width, 4), dtype=np.uint8)
                            else:  # Regular BGR
                                blank_image = np.zeros((avg_height, standard_width, 3), dtype=np.uint8)
                        else:
                            blank_image = np.zeros((avg_height, standard_width, 3), dtype=np.uint8)
                        
                        # Initialize all values properly
                        fe_value = None
                        sio2_value = None
                        al2o3_value = None
                        p_value = None
                        cutoffs1_value = None
                        cutoffs2_value = None
                            
                        # Here we need to look up data in the CSV for this specific interval
                        if hole_csv_data:
                            # Calculate midpoint of current interval
                            midpoint = (meter_start + meter_end) / 2
                            
                            # Only use exact interval matches
                            matching_interval = None
                            
                            for interval_mid, interval_data in hole_csv_data.items():
                                interval_from = interval_data['from']
                                interval_to = interval_data['to']
                                
                                # Check if midpoint is EXACTLY within interval
                                if interval_from <= midpoint <= interval_to:
                                    matching_interval = interval_data
                                    break
                            
                            # Extract values if available (only if exact match found)
                            if matching_interval:
                                fe_value = matching_interval.get('fe_pct')
                                sio2_value = matching_interval.get('sio2_pct')
                                al2o3_value = matching_interval.get('al2o3_pct')
                                p_value = matching_interval.get('p_pct')
                                
                                # Properly handle text columns for cutoffs - don't display 'nan'
                                cutoffs1_value = matching_interval.get('cutoffs1')
                                if pd.isna(cutoffs1_value) or str(cutoffs1_value).lower() == 'nan':
                                    cutoffs1_value = None
                                    
                                cutoffs2_value = matching_interval.get('cutoffs2')
                                if pd.isna(cutoffs2_value) or str(cutoffs2_value).lower() == 'nan':
                                    cutoffs2_value = None
                        
                        # Create missing box with the data
                        black_box = self.generate_compartment_with_data(
                            blank_image, hole_id, meter_start, meter_end,
                            fe_value=fe_value,
                            sio2_value=sio2_value,
                            al2o3_value=al2o3_value,
                            p_value=p_value,
                            cutoffs1_value=cutoffs1_value,
                            cutoffs2_value=cutoffs2_value,
                            is_missing=True,
                            data_column_width=data_column_width
                        )

                        # Find where to insert based on depth
                        insert_index = 0
                        for i, (comp_hole_id, comp_from, comp_to, _) in enumerate(sorted_compartments):
                            if meter_start < comp_from:
                                insert_index = i
                                break
                            insert_index = i + 1

                        # Insert the black box in the processed_images list
                        if insert_index < len(processed_images):
                            processed_images.insert(insert_index, black_box)
                        else:
                            processed_images.append(black_box)

                        # Also insert into sorted_compartments so the insert_index keeps working
                        sorted_compartments.insert(insert_index, (hole_id, meter_start, meter_end, None))
            else:
                self.logger.info("No missing intervals detected - all depths are continuous")

            # Check if we should save individual compartments
            if self.config.get('save_individual_compartments', True):
                # Use FileManager if available
                if hasattr(self, 'file_manager') and self.file_manager is not None:
                    # Save each individual compartment with data columns through FileManager
                    for i, (h_id, d_from, d_to, _) in enumerate(sorted_compartments):
                        if i < len(processed_images):
                            # Calculate compartment depth
                            compartment_depth = int(d_to)
                            
                            # Use the FileManager method to save the compartment with data
                            self.file_manager.save_compartment_with_data(
                                processed_images[i],
                                hole_id,
                                compartment_depth,
                                output_format="png"
                            )
                else:
                    # Fallback for cases without FileManager (should be rare)
                    self.logger.warning("No FileManager available for saving individual compartments")

            # Create a legend image
            legend = self.create_value_legend(width=standard_width)
            
            # Create hole info header
            header_height = 80
            header = np.ones((header_height, standard_width, 3), dtype=np.uint8) * 255
            
            # Add hole ID and depth range
            font = cv2.FONT_HERSHEY_SIMPLEX
            cv2.putText(
                header,
                f"Drillhole: {hole_id}",
                (20, 40),
                font,
                1.2,
                (0, 0, 0),
                2
            )
            
            cv2.putText(
                header,
                f"Depth range: {int(min_depth)}-{int(max_depth)}m",
                (20, 70),
                font,
                0.8,
                (0, 0, 0),
                1
            )
            
            # Verify all images have the same width and type for vconcat
            all_images = [header] + processed_images + [legend]

            # Debug widths before concat
            for i, img in enumerate(all_images):
                img_type = "header" if i == 0 else "legend" if i == len(all_images)-1 else f"compartment {i-1}"
                self.logger.info(f"Image {i} ({img_type}) shape: {img.shape}, width diff from standard: {img.shape[1] - standard_width}")
                
                # Force resize to standard width if needed
                if img.shape[1] != standard_width:
                    self.logger.warning(f"Resizing {img_type} from width {img.shape[1]} to {standard_width}")
                    all_images[i] = cv2.resize(img, (standard_width, img.shape[0]), interpolation=cv2.INTER_AREA)
            
            # Combine all images vertically - verify all have same width
            if processed_images:
                try:
                    # Try concatenating in chunks to avoid potential memory issues
                    chunk_size = 10  # Adjust as needed
                    final_chunks = []
                    
                    for i in range(0, len(all_images), chunk_size):
                        chunk = all_images[i:i+chunk_size]
                        try:
                            chunk_image = cv2.vconcat(chunk)
                            final_chunks.append(chunk_image)
                        except Exception as e:
                            self.logger.error(f"Error concatenating chunk {i//chunk_size}: {str(e)}")
                            # Try individual images if chunk fails
                            for j, img in enumerate(chunk):
                                try:
                                    final_chunks.append(img)
                                except Exception:
                                    self.logger.error(f"Error with image {i+j}")
                    
                    # Final concatenation of chunks
                    final_image = cv2.vconcat(final_chunks)
                    
                    # Save the result
                    output_filename = f"{hole_id}_drillhole_trace.png"
                    output_path = os.path.join(output_dir, output_filename)
                    
                    # Save using cv2.imwrite without creating directories
                    cv2.imwrite(output_path, final_image)
                    
                    self.logger.info(f"Successfully created drillhole trace for {hole_id} at {output_path}")
                    return output_path
                    
                except Exception as e:
                    self.logger.error(f"Error in final concatenation: {str(e)}")
                    self.logger.error(traceback.format_exc())
                    
                    # Use FileManager for fallback if available
                    if hasattr(self, 'file_manager') and self.file_manager is not None:
                        # Get a path from FileManager's drill_traces directory
                        fallback_dir = os.path.join(self.file_manager.dir_structure["drill_traces"], f"{hole_id}_sections")
                    else:
                        # Fallback to output_dir if provided
                        fallback_dir = os.path.join(output_dir or ".", f"{hole_id}_sections")
                    
                    self.logger.info(f"Falling back to saving individual sections in {fallback_dir}")
                    
                    # Save individual sections
                    for i, img in enumerate(all_images):
                        section_path = os.path.join(fallback_dir, f"{hole_id}_section_{i:03d}.png")
                        
                        # Save without creating directories
                        cv2.imwrite(section_path, img)
                    
                    return fallback_dir
            else:
                self.logger.error(f"No valid images processed for hole {hole_id}")
                return None
                
        except Exception as e:
            self.logger.error(f"Error generating drillhole trace for {hole_id}: {str(e)}")
            self.logger.error(traceback.format_exc())
            return None

    def process_all_drillholes(self, 
                            compartment_dir: str,
                            csv_path: Optional[str] = None,
                            selected_columns: Optional[List[str]] = None) -> List[str]:
        """
        Process all drillholes in a directory to create trace images.
        
        Args:
            compartment_dir: Directory containing compartment images
            csv_path: Optional path to CSV file with additional data
            selected_columns: Optional list of columns to include from CSV
            
        Returns:
            List of paths to generated trace images
        """
        # Load CSV data
        csv_data = None
        if csv_path and os.path.exists(csv_path):
            try:
                # Load the CSV
                csv_data = pd.read_csv(csv_path)
                
                # Convert column names to lowercase for case-insensitive matching
                csv_data.columns = [col.lower() for col in csv_data.columns]
                
                # Ensure numeric columns are properly typed, but keep text columns as strings
                for col in csv_data.columns:
                    if col not in ['holeid', 'cutoffs1', 'cutoffs2']:  # Don't convert text columns
                        try:
                            csv_data[col] = pd.to_numeric(csv_data[col], errors='coerce')
                        except:
                            self.logger.warning(f"Could not convert column {col} to numeric")
                
                self.logger.info(f"Loaded CSV with {len(csv_data)} rows and {len(csv_data.columns)} columns")
                
            except Exception as e:
                self.logger.error(f"Error loading CSV data: {str(e)}")
        
        # Collect compartment images
        hole_compartments = self.collect_compartment_images(compartment_dir)
        
        if not hole_compartments:
            self.logger.warning("No valid compartment images found")
            if self.progress_queue:
                self.progress_queue.put(("No valid compartment images found", None))
            return []
        
        # Get the output directory from FileManager if available
        output_dir = None
        if hasattr(self, 'file_manager') and self.file_manager is not None:
            output_dir = self.file_manager.dir_structure["drill_traces"]
            self.logger.info(f"Using FileManager directory for drill traces: {output_dir}")
        else:
            # Fall back to a centralized location if no FileManager available
            # This should almost never happen in normal operation
            output_dir = os.path.join("C:/Excel Automation Local Outputs/Chip Tray Photo Processor", 
                                    "Processed", "Drill Traces")
            self.logger.warning(f"FileManager not available, using fallback directory: {output_dir}")
        
        # Get compartment interval for logging
        compartment_interval = self.config.get('compartment_interval', 1.0)
        
        # Process each hole
        generated_traces = []
        
        for i, (hole_id, compartments) in enumerate(hole_compartments.items()):
            # Update progress
            if self.progress_queue:
                progress = ((i + 1) / len(hole_compartments)) * 100
                self.progress_queue.put((f"Processing hole {i+1}/{len(hole_compartments)}: {hole_id} (interval: {compartment_interval}m)", progress))
            
            # Generate trace using the OpenCV-based method
            trace_path = self.generate_drillhole_trace_cv2(
                hole_id, compartments, csv_data, output_dir
            )
            
            if trace_path:
                generated_traces.append(trace_path)
        
        # Final status update
        status_msg = f"Completed drillhole trace generation: {len(generated_traces)}/{len(hole_compartments)} successful"
        self.logger.info(status_msg)
        if self.progress_queue:
            self.progress_queue.put((status_msg, 100))
        
        return generated_traces

class BlurDetector:
    """
    A class to detect blurry images using the Laplacian variance method.
    
    The Laplacian operator is used to measure the second derivative of an image.
    The variance of the Laplacian is a simple measure of the amount of edges 
    present in an image - blurry images tend to have fewer edges and thus lower variance.
    """

    def __init__(self, threshold: float = 100.0, roi_ratio: float = 0.8):
        """
        Initialize the blur detector with configurable parameters.
        
        Args:
            threshold: Laplacian variance threshold below which an image is considered blurry
            roi_ratio: Ratio of central image area to use for blur detection (0.0-1.0)
        """
        self.threshold = threshold
        self.roi_ratio = max(0.1, min(1.0, roi_ratio))  # Clamp between 0.1 and 1.0
        self.logger = logging.getLogger(__name__)
    
    def get_laplacian_variance(self, image: np.ndarray) -> float:
        """
        Calculate the variance of the Laplacian for an image.
        
        Args:
            image: Input image as numpy array
            
        Returns:
            Variance of the Laplacian as a measure of blurriness (lower = more blurry)
        """
        try:
            # Convert to grayscale if needed
            if len(image.shape) == 3:
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            else:
                gray = image.copy()
            
            # Extract region of interest if ratio < 1.0
            if self.roi_ratio < 1.0:
                h, w = gray.shape
                center_h, center_w = h // 2, w // 2
                roi_h, roi_w = int(h * self.roi_ratio), int(w * self.roi_ratio)
                start_h, start_w = center_h - (roi_h // 2), center_w - (roi_w // 2)
                
                # Ensure ROI is within image bounds
                start_h = max(0, start_h)
                start_w = max(0, start_w)
                end_h = min(h, start_h + roi_h)
                end_w = min(w, start_w + roi_w)
                
                # Extract ROI
                gray = gray[start_h:end_h, start_w:end_w]
            
            # Calculate Laplacian
            laplacian = cv2.Laplacian(gray, cv2.CV_64F)
            
            # Calculate variance
            variance = laplacian.var()
            return variance
            
        except Exception as e:
            self.logger.error(f"Error calculating Laplacian variance: {str(e)}")
            return 0.0
    
    def is_blurry(self, image: np.ndarray) -> Tuple[bool, float]:
        """
        Determine if an image is blurry.
        
        Args:
            image: Input image as numpy array
            
        Returns:
            Tuple of (is_blurry, variance_score)
        """
        variance = self.get_laplacian_variance(image)
        return variance < self.threshold, variance
    
    def analyze_image_with_visualization(self, image: np.ndarray) -> Tuple[bool, float, np.ndarray]:
        """
        Analyze an image and create a visualization of the blur detection.
        
        Args:
            image: Input image as numpy array
            
        Returns:
            Tuple of (is_blurry, variance_score, visualization_image)
        """
        # Make a copy for visualization
        viz_image = image.copy()
        
        # Calculate blur metrics
        is_blurry, variance = self.is_blurry(image)
        
        # Add text with blur metrics
        status = "BLURRY" if is_blurry else "SHARP"
        color = (0, 0, 255) if is_blurry else (0, 255, 0)  # Red for blurry, green for sharp
        
        # Add a background box for better text visibility
        h, w = viz_image.shape[:2]
        cv2.rectangle(viz_image, (10, h - 60), (w - 10, h - 10), (0, 0, 0), -1)
        
        # Add text
        cv2.putText(
            viz_image,
            f"Status: {status}", 
            (20, h - 40),
            cv2.FONT_HERSHEY_SIMPLEX, 
            0.6, 
            color, 
            2
        )
        
        cv2.putText(
            viz_image,
            f"Laplacian Variance: {variance:.2f} (threshold: {self.threshold:.2f})", 
            (20, h - 15),
            cv2.FONT_HERSHEY_SIMPLEX, 
            0.5, 
            (255, 255, 255), 
            1
        )
        
        return is_blurry, variance, viz_image
    
    def batch_analyze_images(self, 
                           images: List[np.ndarray],
                           generate_visualizations: bool = False) -> List[Dict[str, Any]]:
        """
        Analyze a batch of images for blurriness.
        
        Args:
            images: List of input images
            generate_visualizations: Whether to create visualization images
            
        Returns:
            List of dictionaries with analysis results
        """
        results = []
        
        for i, image in enumerate(images):
            try:
                # Perform analysis
                if generate_visualizations:
                    is_blurry, variance, viz_image = self.analyze_image_with_visualization(image)
                    result = {
                        'index': i,
                        'is_blurry': is_blurry,
                        'variance': variance,
                        'visualization': viz_image
                    }
                else:
                    is_blurry, variance = self.is_blurry(image)
                    result = {
                        'index': i,
                        'is_blurry': is_blurry,
                        'variance': variance
                    }
                
                results.append(result)
                
            except Exception as e:
                self.logger.error(f"Error analyzing image {i}: {str(e)}")
                # Add a placeholder result for the failed image
                results.append({
                    'index': i,
                    'is_blurry': False,
                    'variance': 0.0,
                    'error': str(e)
                })
        
        return results
    
    def calibrate_threshold(self, 
                           sharp_images: List[np.ndarray], 
                           blurry_images: List[np.ndarray],
                           safety_factor: float = 1.5) -> float:
        """
        Calibrate the blur threshold based on example images.
        
        Args:
            sharp_images: List of sharp (good quality) images
            blurry_images: List of blurry (poor quality) images
            safety_factor: Factor to apply to the calculated threshold
            
        Returns:
            Calibrated threshold value
        """
        if not sharp_images or not blurry_images:
            self.logger.warning("Not enough sample images for calibration, using default threshold")
            return self.threshold
        
        try:
            # Calculate variances for sharp images
            sharp_variances = [self.get_laplacian_variance(img) for img in sharp_images]
            
            # Calculate variances for blurry images
            blurry_variances = [self.get_laplacian_variance(img) for img in blurry_images]
            
            # Find the minimum variance of sharp images
            min_sharp = min(sharp_variances)
            
            # Find the maximum variance of blurry images
            max_blurry = max(blurry_variances)
            
            # Calculate the threshold
            if min_sharp > max_blurry:
                # Clear separation - use the midpoint
                threshold = (min_sharp + max_blurry) / 2
            else:
                # Overlap - use a weighted average
                threshold = (min_sharp * 0.7 + max_blurry * 0.3)
            
            # Apply safety factor (lower the threshold to err on the side of detecting blurry images)
            threshold = threshold / safety_factor
            
            self.logger.info(f"Calibrated blur threshold: {threshold:.2f}")
            
            # Update the instance threshold
            self.threshold = threshold
            
            return threshold
            
        except Exception as e:
            self.logger.error(f"Error during threshold calibration: {str(e)}")
            return self.threshold


class FileManager:
    """
    Manages file operations for the Chip Tray Extractor.
    
    Handles directory creation, file naming conventions, and saving operations.

    C:\\Excel Automation Local Outputs\\Chip Tray Photo Processor\\
    ├── Program Resources

        C:\\Excel Automation Local Outputs\\Chip Tray Photo Processor\\Processed\
    ├── Blur Analysis\\[HoleID]\
    ├── Chip Compartments\\[HoleID]\\With Assays
    ├── Debug Images\\[HoleID]\
    ├── Drill Traces\
    ├── Processed Originals\\[HoleID]\
    └── Failed and Skipped Originals\\[HoleID]\
    """
    
    def __init__(self, base_dir: str = None):
        """
        Initialize the file manager with a base directory.
        
        Args:
            base_dir: Base directory for all outputs, defaults to "C:\\Excel Automation Local Outputs\\Chip Tray Photo Processor"
        """
        # Logger setup first
        self.logger = logging.getLogger(__name__)
        
        # Set up base directory before using it
        self.base_dir = base_dir or Path("C:/Excel Automation Local Outputs/Chip Tray Photo Processor")

        self.processed_dir = os.path.join(self.base_dir, "Processed")
        
        # Now that processed_dir exists, we can define directory structure
        self.dir_structure = {
            "blur_analysis": os.path.join(self.processed_dir, "Blur Analysis"),
            "chip_compartments": os.path.join(self.processed_dir, "Chip Compartments"),
            "debug_images": os.path.join(self.processed_dir, "Debug Images"),
            "drill_traces": os.path.join(self.processed_dir, "Drill Traces"),
            "processed_originals": os.path.join(self.processed_dir, "Processed Originals"),
            "failed_originals": os.path.join(self.processed_dir, "Failed and Skipped Originals")
        }
        
        # Create base directories after everything is set up
        self.create_base_directories()


    # in FileManager Class    
    def check_original_file_processed(self, original_filename: str) -> Optional[Dict[str, Any]]:
        """
        Check if an original file has already been processed by extracting metadata from filename.
        
        Args:
            original_filename: Original input filename
            
        Returns:
            Dictionary with extracted metadata if found, None otherwise
        """
        try:
            # Get the base filename without path and extension
            base_name = os.path.splitext(os.path.basename(original_filename))[0]
            
            # Define a pattern to extract metadata from filename
            # Pattern matches formats like "BB1234_40-60_Original" or "BB1234_40-60_Original_Skipped"
            pattern = r'([A-Z]{2}\d{4})_(\d+)-(\d+)_Original(?:_Skipped)?'
            
            # Try to extract metadata directly from filename
            match = re.search(pattern, base_name)
            if match:
                hole_id = match.group(1)
                depth_from = float(match.group(2))
                depth_to = float(match.group(3))
                
                # Check against valid prefixes if the configuration exists and prefix validation is enabled
                valid_prefixes = getattr(self, 'config', {}).get('valid_hole_prefixes', [])
                enable_prefix_validation = getattr(self, 'config', {}).get('enable_prefix_validation', False)
                
                # If prefix validation is enabled, check against valid prefixes
                if enable_prefix_validation and valid_prefixes:
                    prefix = hole_id[:2].upper()
                    if prefix not in valid_prefixes:
                        self.logger.warning(f"Hole ID prefix {prefix} not in valid prefixes: {valid_prefixes}")
                        return None
                
                # Determine if the file was previously skipped
                is_skipped = '_Skipped' in base_name
                
                self.logger.info(f"Found metadata in filename: {hole_id}, {depth_from}-{depth_to}")
                return {
                    'hole_id': hole_id,
                    'depth_from': depth_from,
                    'depth_to': depth_to,
                    'confidence': 100.0 if not is_skipped else 90.0,  # Slightly lower confidence for skipped files
                    'from_filename': True,
                    'previously_skipped': is_skipped
                }
            
            # No match found
            return None
            
        except Exception as e:
            self.logger.error(f"Error checking for previously processed file: {str(e)}")
            return None
    
    def save_compartment(self, 
                        image: np.ndarray, 
                        hole_id: str, 
                        compartment_num: int,
                        has_data: bool = False,
                        output_format: str = "png") -> str:
        """
        Save a compartment image.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            compartment_num: Compartment number
            has_data: Whether the image has data columns
            output_format: Output image format
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory
            save_dir = self.get_hole_dir("chip_compartments", hole_id)
            
            # Create filename with 3-digit compartment number (001, 002, etc.)
            if has_data:
                filename = f"{hole_id}_CC_{compartment_num:03d}_Data.{output_format}"
            else:
                filename = f"{hole_id}_CC_{compartment_num:03d}.{output_format}"
            
            # Full path
            file_path = os.path.join(save_dir, filename)
            
            # Save the image
            if output_format.lower() == "jpg":
                cv2.imwrite(file_path, image, [cv2.IMWRITE_JPEG_QUALITY, 100])
            else:
                cv2.imwrite(file_path, image)
            
            self.logger.info(f"Saved compartment image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving compartment image: {str(e)}")
            return None

    def save_compartment_with_data(self, 
                                image: np.ndarray, 
                                hole_id: str, 
                                compartment_num: int,
                                output_format: str = "png") -> str:
        """
        Save a compartment image with data columns to a 'With_Data' subfolder.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            compartment_num: Compartment number
            output_format: Output image format
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory and create a "With_Data" subfolder
            base_dir = self.get_hole_dir("chip_compartments", hole_id)
            with_data_dir = os.path.join(base_dir, "With_Data")
            os.makedirs(with_data_dir, exist_ok=True)
            
            # Create filename with 3-digit compartment number (001, 002, etc.)
            filename = f"{hole_id}_CC_{compartment_num:03d}_Data.{output_format}"
            
            # Full path
            file_path = os.path.join(with_data_dir, filename)
            
            # Save the image
            if output_format.lower() == "jpg":
                cv2.imwrite(file_path, image, [cv2.IMWRITE_JPEG_QUALITY, 100])
            else:
                cv2.imwrite(file_path, image)
            
            self.logger.info(f"Saved compartment with data image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving compartment with data image: {str(e)}")
            return None
    
    
    
    def rename_debug_files(self, 
                        original_filename: str, 
                        hole_id: Optional[str], 
                        depth_from: Optional[float], 
                        depth_to: Optional[float]) -> None:
        """
        Rename debug files for a specific image after successful metadata extraction.
        
        Args:
            original_filename: Original input image filename
            hole_id: Extracted hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
        """
        try:
            if not (hole_id and depth_from is not None and depth_to is not None):
                return
            
            # Get the debug directory for this hole ID
            debug_dir = self.get_hole_dir("debug_images", hole_id)
            
            # Find all debug files for this image
            base_name = os.path.splitext(os.path.basename(original_filename))[0]
            
            try:
                debug_files = [
                    f for f in os.listdir(debug_dir) 
                    if f.startswith(base_name) and f.endswith('.jpg')
                ]
            except FileNotFoundError:
                # Directory might not exist yet
                os.makedirs(debug_dir, exist_ok=True)
                debug_files = []
                self.logger.info(f"Created debug directory for {hole_id}")
            
            # Also look for temp debug images in the Unidentified folder
            temp_debug_dir = os.path.join(self.dir_structure["debug_images"], "Unidentified")
            if os.path.exists(temp_debug_dir):
                try:
                    temp_files = [
                        f for f in os.listdir(temp_debug_dir)
                        if f.startswith(base_name) and f.endswith('.jpg')
                    ]
                    
                    # Move and rename temp files
                    for old_filename in temp_files:
                        # Extract the step name from the old filename
                        if '_' in old_filename:
                            step_name = old_filename.split('_', 1)[1].replace('.jpg', '')
                            
                            # Generate new filename with metadata
                            new_filename = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Debug_{step_name}.jpg"
                            
                            old_path = os.path.join(temp_debug_dir, old_filename)
                            new_path = os.path.join(debug_dir, new_filename)
                            
                            try:
                                # Ensure debug directory exists
                                os.makedirs(debug_dir, exist_ok=True)
                                
                                # Copy the file (use copy instead of move for safety)
                                shutil.copy2(old_path, new_path)
                                
                                # Only remove original after successful copy
                                os.remove(old_path)
                                
                                self.logger.info(f"Moved debug file from temp location: {old_filename} -> {new_filename}")
                            except Exception as e:
                                self.logger.error(f"Error moving temp debug file {old_filename}: {e}")
                except Exception as e:
                    self.logger.error(f"Error processing temp debug files: {e}")
            
            # Rename files in the debug directory
            for old_filename in debug_files:
                try:
                    # Extract the step name from the old filename
                    step_parts = old_filename.split('_')
                    if len(step_parts) >= 2:
                        step_name = step_parts[-1].replace('.jpg', '')
                        
                        # Generate new filename with metadata
                        new_filename = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Debug_{step_name}.jpg"
                        
                        old_path = os.path.join(debug_dir, old_filename)
                        new_path = os.path.join(debug_dir, new_filename)
                        
                        # Rename the file
                        os.rename(old_path, new_path)
                        self.logger.info(f"Renamed debug file: {old_filename} -> {new_filename}")
                except Exception as e:
                    self.logger.error(f"Error renaming debug file {old_filename}: {e}")
        except Exception as e:
            self.logger.error(f"Error in rename_debug_files: {e}")

    def save_temp_debug_image(self, image: np.ndarray, original_filename: str, debug_type: str) -> str:
        """
        Save a debug image without proper hole/depth metadata.
        
        Args:
            image: Image to save
            original_filename: Original input filename used to generate base name
            debug_type: Type of debug image
            
        Returns:
            Full path to the saved image
        """
        try:
            # Use the centralized directory structure
            temp_dir = os.path.join(self.dir_structure["debug_images"], "Unidentified")
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generate filename from original file
            base_name = os.path.splitext(os.path.basename(original_filename))[0]
            filename = f"{base_name}_{debug_type}.jpg"
            full_path = os.path.join(temp_dir, filename)
            
            # Save the image with moderate compression
            cv2.imwrite(full_path, image, [cv2.IMWRITE_JPEG_QUALITY, 90])
            
            self.logger.info(f"Saved temporary debug image: {full_path}")
            return full_path
            
        except Exception as e:
            self.logger.error(f"Error saving temporary debug image: {str(e)}")
            
            # Create a fallback directory if needed
            fallback_dir = os.path.join(self.base_dir, "Temp_Debug")
            os.makedirs(fallback_dir, exist_ok=True)
            
            # Generate a unique fallback filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_filename = f"debug_{timestamp}_{debug_type}.jpg"
            fallback_path = os.path.join(fallback_dir, fallback_filename)
            
            # Try to save to fallback location
            try:
                cv2.imwrite(fallback_path, image, [cv2.IMWRITE_JPEG_QUALITY, 90])
                self.logger.info(f"Saved debug image to fallback location: {fallback_path}")
                return fallback_path
            except Exception as fallback_error:
                self.logger.error(f"Failed to save debug image to fallback location: {str(fallback_error)}")
                return ""
    
    def create_base_directories(self) -> None:
        """Create the base directory structure."""
        try:
            # Create the processed directory
            os.makedirs(self.processed_dir, exist_ok=True)
            
            # Create each subdirectory
            for dir_path in self.dir_structure.values():
                os.makedirs(dir_path, exist_ok=True)
                
            self.logger.info(f"Created base directory structure in {self.base_dir}")
        except Exception as e:
            self.logger.error(f"Error creating directory structure: {str(e)}")
            raise
    
    def get_hole_dir(self, dir_type: str, hole_id: str) -> str:
        """
        Get the directory path for a specific hole ID.
        
        Args:
            dir_type: Directory type (must be one of the keys in dir_structure)
            hole_id: Hole ID
            
        Returns:
            Full path to the hole directory
        """
        if dir_type not in self.dir_structure:
            raise ValueError(f"Invalid directory type: {dir_type}")
        
        # Create hole-specific directory
        hole_dir = os.path.join(self.dir_structure[dir_type], hole_id)
        os.makedirs(hole_dir, exist_ok=True)
        
        return hole_dir
    
    
    def save_blur_analysis(self, 
                        image: np.ndarray, 
                        hole_id: str, 
                        compartment_num: int) -> str:
        """
        Save a blur analysis image.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            compartment_num: Compartment number
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory
            save_dir = self.get_hole_dir("blur_analysis", hole_id)
            
            # Create filename
            filename = f"{hole_id}_{compartment_num}_blur_analysis.jpg"
            
            # Full path
            file_path = os.path.join(save_dir, filename)
            
            # Save the image
            cv2.imwrite(file_path, image, [cv2.IMWRITE_JPEG_QUALITY, 90])
            
            self.logger.info(f"Saved blur analysis image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving blur analysis image: {str(e)}")
            return None

    def save_compartment(self, 
                        image: np.ndarray, 
                        hole_id: str, 
                        compartment_num: int,
                        has_data: bool = False,
                        output_format: str = "png") -> str:
        """
        Save a compartment image.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            compartment_num: Compartment number
            has_data: Whether the image has data columns
            output_format: Output image format
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory
            save_dir = self.get_hole_dir("chip_compartments", hole_id)
            
            # Create filename with 3-digit compartment number (001, 002, etc.)
            if has_data:
                filename = f"{hole_id}_CC_{compartment_num:03d}_Data.{output_format}"
            else:
                filename = f"{hole_id}_CC_{compartment_num:03d}.{output_format}"
            
            # Full path
            file_path = os.path.join(save_dir, filename)
            
            # Save the image
            if output_format.lower() == "jpg":
                cv2.imwrite(file_path, image, [cv2.IMWRITE_JPEG_QUALITY, 100])
            else:
                cv2.imwrite(file_path, image)
            
            self.logger.info(f"Saved compartment image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving compartment image: {str(e)}")
            return None
    
    def save_debug_image(self, 
                       image: np.ndarray, 
                       hole_id: str, 
                       depth_from: float,
                       depth_to: float,
                       image_type: str) -> str:
        """
        Save a debug image.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            image_type: Type of debug image
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory
            save_dir = self.get_hole_dir("debug_images", hole_id)
            
            # Create filename
            filename = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Debug_{image_type}.jpg"
            
            # Full path
            file_path = os.path.join(save_dir, filename)
            
            # Save the image
            cv2.imwrite(file_path, image, [cv2.IMWRITE_JPEG_QUALITY, 85])
            
            self.logger.info(f"Saved debug image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving debug image: {str(e)}")
            return None
    
    def save_drill_trace(self, 
                       image: np.ndarray, 
                       hole_id: str) -> str:
        """
        Save a drill trace image.
        
        Args:
            image: Image to save
            hole_id: Hole ID
            
        Returns:
            Path to the saved file
        """
        try:
            # Get the appropriate directory
            save_dir = self.dir_structure["drill_traces"]
            
            # Create filename
            filename = f"{hole_id}_Trace.png"
            
            # Full path
            file_path = os.path.join(save_dir, filename)
            
            # Save the image
            cv2.imwrite(file_path, image)
            
            self.logger.info(f"Saved drill trace image: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error saving drill trace image: {str(e)}")
            return None
    
    
    def move_original_file(self, 
                        source_path: str, 
                        hole_id: str, 
                        depth_from: float,
                        depth_to: float,
                        is_processed: bool = True) -> str:
        """
        Move an original file to the appropriate directory.
        
        Args:
            source_path: Path to the source file
            hole_id: Hole ID
            depth_from: Starting depth
            depth_to: Ending depth
            is_processed: Whether the file was successfully processed
            
        Returns:
            Path to the new file location
        """
        try:
            # Determine the target directory
            if is_processed:
                target_dir = self.get_hole_dir("processed_originals", hole_id)
            else:
                target_dir = self.get_hole_dir("failed_originals", hole_id)
            
            # Get original file extension
            _, ext = os.path.splitext(source_path)
            
            # Create new filename with appropriate suffix
            status_suffix = "" if is_processed else "_Skipped"
            new_filename = f"{hole_id}_{int(depth_from)}-{int(depth_to)}_Original{status_suffix}{ext}"
            
            # Full target path
            target_path = os.path.join(target_dir, new_filename)
            
            # Check if file already exists, add number if needed
            counter = 1
            base_name = new_filename
            while os.path.exists(target_path):
                name_parts = os.path.splitext(base_name)
                new_filename = f"{name_parts[0]}_{counter}{name_parts[1]}"
                target_path = os.path.join(target_dir, new_filename)
                counter += 1
            
            # Move the file
            shutil.move(source_path, target_path)
            
            self.logger.info(f"Moved original file to: {target_path}")
            return target_path
            
        except Exception as e:
            self.logger.error(f"Error moving original file: {str(e)}")
            return None



def main():
    """Run the application."""
    # Parse command line arguments
    import argparse
    
    parser = argparse.ArgumentParser(description='Chip Tray Extractor')
    parser.add_argument('--input', '-i', help='Input folder containing chip tray photos')
    parser.add_argument('--output', '-o', help='Output folder for extracted compartments')
    parser.add_argument('--no-gui', action='store_true', help='Run in command-line mode without GUI')
    parser.add_argument('--format', choices=['jpg', 'png', 'tiff'], default='png', 
                        help='Output image format')
    parser.add_argument('--quality', type=int, default=100, help='JPEG quality (0-100)')
    parser.add_argument('--debug', action='store_true', help='Save debug images')
    
    args = parser.parse_args()
    
    # Create the extractor
    extractor = ChipTrayExtractor()
    
    # Update config based on command-line arguments
    if args.output:
        extractor.config['output_folder'] = args.output
    if args.format:
        extractor.config['output_format'] = args.format
    if args.quality:
        extractor.config['jpeg_quality'] = args.quality
    if args.debug:
        extractor.config['save_debug_images'] = False
    
    # Run in appropriate mode
    if args.no_gui or args.input:
        # Command-line mode
        if args.input:
            successful, failed = extractor.process_folder(args.input)
            logger.info(f"Processing complete: {successful} successful, {failed} failed")
        else:
            logger.error("No input folder specified. Use --input to specify a folder to process.")
            return 1
    else:
        # GUI mode
        extractor.create_gui()
        extractor.root.mainloop()
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
